#!/usr/bin/env python3
"""
PubMed v2 Step 4: Generate 2 Final Excel Files

Source: 1,899 PMIDs from PubMed2 sheet in [0211] IEEE John_Screening 1 & 2.xlsx

File 1: PubMed_FullText_Screened.xlsx — papers that passed both T/A + full-text screening
File 2: PubMed_NoFullText.xlsx — papers that passed T/A only (no full text available)
"""

import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = '/home/user/Bias-Review-Paper'
INPUT_FILE = f'{BASE_DIR}/pubmed_v2_step3_results.json'

# ============================================================
# COLUMN EXTRACTION (same maps as before)
# ============================================================

METHOD_MAP = {
    'Deep Learning': ['deep learning', 'deep neural', 'cnn', 'convolutional neural', 'resnet', 'densenet', 'vgg', 'inception'],
    'NLP/LLM': ['natural language processing', 'nlp', 'large language model', 'llm', 'chatgpt', 'gpt-4', 'gpt-3', 'bert', 'transformer', 'text mining', 'text classification'],
    'Random Forest': ['random forest'],
    'Logistic Regression': ['logistic regression'],
    'XGBoost/Gradient Boosting': ['xgboost', 'gradient boosting', 'lightgbm', 'catboost'],
    'Support Vector Machine': ['support vector', 'svm'],
    'Neural Network': ['neural network', 'mlp', 'perceptron'],
    'Decision Tree': ['decision tree', 'cart'],
    'Ensemble Methods': ['ensemble', 'bagging', 'stacking'],
    'Federated Learning': ['federated learning'],
    'Reinforcement Learning': ['reinforcement learning'],
    'Computer Vision/Imaging AI': ['computer vision', 'image classification', 'image segmentation', 'object detection', 'medical imaging'],
    'Clinical Prediction Model': ['clinical prediction', 'risk prediction', 'risk score', 'prediction model', 'predictive model', 'prognostic model'],
    'Clinical Decision Support': ['clinical decision support', 'decision support system', 'cdss'],
    'Foundation Model': ['foundation model'],
    'Generative AI': ['generative ai', 'generative model', 'gan', 'generative adversarial'],
    'Clustering': ['clustering', 'k-means', 'unsupervised'],
    'Regression': ['linear regression', 'regression model'],
    'Survival Analysis': ['survival analysis', 'cox regression', 'hazard model'],
}

DOMAIN_MAP = {
    'Radiology/Medical Imaging': ['radiology', 'radiograph', 'x-ray', 'ct scan', 'mri', 'mammograph', 'ultrasound', 'medical imaging', 'chest radiograph'],
    'Dermatology': ['dermatology', 'skin lesion', 'skin cancer', 'melanoma', 'skin disease', 'dermatoscop'],
    'Cardiology': ['cardiology', 'cardiovascular', 'heart', 'cardiac', 'ecg', 'ekg', 'echocardiogram', 'atrial fibrillation', 'coronary'],
    'Oncology': ['oncology', 'cancer', 'tumor', 'tumour', 'neoplasm', 'malignant', 'chemotherapy', 'radiation therapy'],
    'Ophthalmology': ['ophthalmology', 'retinal', 'retinopathy', 'diabetic retinopathy', 'glaucoma', 'fundus', 'optic'],
    'Mental Health/Psychiatry': ['mental health', 'psychiatry', 'depression', 'anxiety', 'suicide', 'psychiatric', 'behavioral health', 'substance abuse', 'opioid'],
    'Emergency Medicine': ['emergency department', 'emergency medicine', 'trauma', 'triage', 'acute care'],
    'ICU/Critical Care': ['icu', 'intensive care', 'critical care', 'sepsis', 'mechanical ventilation'],
    'EHR/Health Informatics': ['electronic health record', 'ehr', 'emr', 'health information', 'clinical informatics', 'health informatics'],
    'Primary Care': ['primary care', 'family medicine', 'general practice', 'ambulatory'],
    'Surgery': ['surgery', 'surgical', 'operative', 'postoperative'],
    'Pathology': ['pathology', 'histopathology', 'cytology', 'biopsy'],
    'Pediatrics': ['pediatric', 'paediatric', 'child', 'neonatal', 'newborn'],
    'Obstetrics/Maternal Health': ['obstetric', 'maternal', 'pregnancy', 'prenatal', 'perinatal', 'childbirth'],
    'Nephrology': ['kidney', 'renal', 'nephrology', 'dialysis'],
    'Pulmonology': ['pulmonary', 'respiratory', 'lung', 'pneumonia', 'covid-19', 'asthma', 'copd'],
    'Neurology': ['neurology', 'neurological', 'stroke', 'alzheimer', 'dementia', 'brain', 'epilepsy'],
    'Genomics/Genetics': ['genomic', 'genetic', 'genome', 'gene expression', 'variant', 'sequencing'],
    'Public Health': ['public health', 'population health', 'epidemiology', 'surveillance', 'community health'],
    'Drug Discovery/Pharmacology': ['drug discovery', 'pharmacology', 'pharmaceutical', 'drug development'],
    'Infectious Disease': ['infectious disease', 'infection', 'hiv', 'tuberculosis', 'malaria', 'hepatitis'],
    'Endocrinology/Diabetes': ['diabetes', 'diabetic', 'endocrin', 'insulin', 'glucose', 'hba1c'],
    'Wearables/Remote Monitoring': ['wearable', 'pulse oximetry', 'spo2', 'remote monitoring', 'mobile health', 'mhealth'],
}

BIAS_AXIS_MAP = {
    'Race/Ethnicity': ['race', 'racial', 'ethnicity', 'ethnic', 'african american', 'black', 'white', 'hispanic', 'latino', 'asian', 'indigenous', 'minority', 'skin color', 'skin tone', 'skin pigment'],
    'Gender/Sex': ['gender', 'sex', 'male', 'female', 'women', 'men', 'transgender', 'sex-based', 'gender-based'],
    'Age': ['age', 'aging', 'elderly', 'older adult', 'pediatric', 'geriatric', 'age-related', 'age group'],
    'Socioeconomic Status': ['socioeconomic', 'income', 'poverty', 'insurance', 'uninsured', 'underinsured', 'medicaid', 'sociodemographic', 'education level', 'employment'],
    'Language': ['language', 'non-english', 'english proficiency', 'limited english', 'linguistic'],
    'Disability': ['disability', 'disabled', 'impairment', 'handicap'],
    'Geographic': ['geographic', 'rural', 'urban', 'region', 'low-resource', 'low-income country', 'global south', 'developing country'],
    'Insurance Status': ['insurance', 'uninsured', 'underinsured', 'medicaid', 'medicare', 'payer'],
    'Intersectional': ['intersectional', 'intersectionality', 'multiple sensitive', 'multi-attribute', 'compound'],
}

LIFECYCLE_STAGE_MAP = {
    'Data Collection': ['data collection', 'data acquisition', 'dataset creation', 'data gathering', 'sampling', 'cohort selection', 'data source'],
    'Data Preprocessing': ['preprocess', 'pre-process', 'data cleaning', 'imputation', 'feature engineering', 'data augmentation', 'resampling', 'oversampling', 'undersampling', 'reweighting', 'reweighing'],
    'Model Development/Training': ['model training', 'model development', 'training phase', 'in-processing', 'adversarial debiasing', 'fairness constraint', 'regularization', 'fair representation', 'calibration'],
    'Model Evaluation': ['model evaluation', 'evaluation', 'testing', 'validation', 'performance assessment', 'fairness metric', 'bias assessment', 'bias detection', 'audit', 'post-processing', 'subgroup analysis'],
    'Deployment': ['deployment', 'clinical implementation', 'real-world', 'production', 'clinical practice', 'clinical workflow', 'monitoring'],
}

APPROACH_MAP = {
    'Reweighting/Resampling': ['reweighting', 'reweighing', 'resampling', 'oversampling', 'undersampling', 'smote'],
    'Adversarial Debiasing': ['adversarial debiasing', 'adversarial learning', 'adversarial training'],
    'Fairness Constraints': ['fairness constraint', 'fairness regulariz', 'fairness-aware training'],
    'Calibration': ['calibration', 'recalibration', 'platt scaling'],
    'Threshold Adjustment': ['threshold', 'cutoff', 'operating point'],
    'Data Augmentation': ['data augmentation', 'synthetic data', 'augmented data', 'fair mixup'],
    'Subgroup Analysis': ['subgroup analysis', 'stratified analysis', 'subpopulation', 'disaggregated'],
    'Fairness Metrics Evaluation': ['demographic parity', 'equalized odds', 'equal opportunity', 'disparate impact', 'fairness metric', 'predictive parity', 'calibration across'],
    'Counterfactual Fairness': ['counterfactual', 'causal inference', 'causal fairness'],
    'Representation Learning': ['representation learning', 'fair representation', 'disentangl', 'embedding'],
    'Transfer Learning': ['transfer learning', 'domain adaptation', 'fine-tun'],
    'Federated Learning': ['federated learning', 'federated model'],
    'Explainability/Interpretability': ['explainab', 'interpretab', 'shap', 'lime', 'attention'],
    'Post-hoc Correction': ['post-processing', 'post-hoc', 'reject option'],
    'Diverse/Representative Data': ['diverse dataset', 'representative data', 'inclusive data', 'data diversity', 'balanced dataset'],
    'Bias Auditing Framework': ['audit', 'bias audit', 'fairness audit', 'model card', 'datasheet'],
    'Multi-task Learning': ['multi-task', 'multitask'],
    'Ensemble Methods': ['ensemble', 'model ensemble'],
    'Regularization': ['regulariz', 'penalty', 'constraint-based'],
}

CLINICAL_SETTING_MAP = {
    'Hospital/Inpatient': ['hospital', 'inpatient', 'hospitalized', 'admitted'],
    'ICU': ['icu', 'intensive care', 'critical care'],
    'Emergency Department': ['emergency department', 'emergency room', 'ed visit'],
    'Primary Care/Outpatient': ['primary care', 'outpatient', 'ambulatory', 'clinic visit', 'office visit'],
    'Public Health/Population': ['public health', 'population', 'community', 'screening program', 'surveillance'],
    'Telehealth/Remote': ['telehealth', 'telemedicine', 'remote', 'mobile health', 'mhealth', 'wearable'],
    'Clinical Trial': ['clinical trial', 'randomized controlled', 'rct'],
    'Laboratory/Pathology': ['laboratory', 'pathology', 'lab test'],
    'Imaging Center': ['imaging center', 'radiology department'],
    'Safety-Net/Underserved': ['safety net', 'safety-net', 'underserved', 'low-resource'],
    'Long-term Care': ['long-term care', 'nursing home', 'assisted living'],
}


def extract_study_type(paper):
    title = (paper.get('title', '') or '').lower()
    abstract = (paper.get('abstract', '') or '').lower()
    pub_types = (paper.get('pub_types', '') or '').lower()
    combined = title + ' ' + abstract + ' ' + pub_types
    if any(t in combined for t in ['survey', 'questionnaire', 'interview']):
        return 'Survey/Qualitative'
    if any(t in combined for t in ['commentary', 'editorial', 'perspective', 'viewpoint', 'opinion', 'letter']):
        return 'Commentary/Editorial'
    if any(t in combined for t in ['guideline', 'recommendation', 'consensus', 'policy']):
        return 'Guideline/Policy'
    if any(t in combined for t in ['framework', 'toolkit', 'toolbox', 'pipeline']):
        return 'Framework/Toolkit'
    if any(t in combined for t in ['we propose', 'we develop', 'we introduce', 'novel method', 'novel approach', 'we present a method', 'we present a novel']):
        return 'Methodology'
    if any(t in combined for t in ['experiment', 'dataset', 'we evaluated', 'we trained', 'we tested', 'we applied', 'we analyzed', 'we assessed', 'results show', 'our results', 'we found']):
        return 'Empirical Study'
    return 'Empirical Study'


def extract_mapped(paper, mapping, fields=('title', 'abstract')):
    combined = ' '.join((paper.get(f, '') or '').lower() for f in fields)
    matches = [label for label, terms in mapping.items() if any(t in combined for t in terms)]
    return '; '.join(matches) if matches else 'Not specified'


def extract_assessment_or_mitigation(paper):
    combined = ((paper.get('title', '') or '') + ' ' + (paper.get('abstract', '') or '')).lower()
    assess_terms = ['assess', 'detect', 'evaluat', 'audit', 'measur', 'quantif', 'identif', 'characteriz', 'examin', 'analyz', 'investigat']
    mitigate_terms = ['mitigat', 'debias', 'reduc', 'eliminat', 'correct', 'remov', 'address', 'alleviat', 'improv fairness', 'fairness-aware', 'fair machine', 'reweigh', 'adversarial debias']
    has_assess = any(t in combined for t in assess_terms)
    has_mitigate = any(t in combined for t in mitigate_terms)
    if has_assess and has_mitigate:
        return 'Both'
    elif has_mitigate:
        return 'Mitigation'
    elif has_assess:
        return 'Assessment'
    return 'Not specified'


def extract_key_findings(paper):
    abstract = (paper.get('abstract', '') or '')
    if not abstract:
        return 'No abstract available'
    lower = abstract.lower()
    for marker in ['CONCLUSION:', 'CONCLUSIONS:', 'RESULTS:', 'FINDINGS:']:
        idx = lower.find(marker.lower())
        if idx != -1:
            finding = abstract[idx:].strip()
            sentences = re.split(r'(?<=[.!?])\s+', finding)
            result = ' '.join(sentences[:3])
            if len(result) > 500:
                result = result[:500] + '...'
            return result
    sentences = re.split(r'(?<=[.!?])\s+', abstract)
    if len(sentences) >= 2:
        result = ' '.join(sentences[-2:])
    else:
        result = abstract
    if len(result) > 500:
        result = result[:500] + '...'
    return result


def extract_all(paper):
    paper['study_type'] = extract_study_type(paper)
    paper['ai_ml_method'] = extract_mapped(paper, METHOD_MAP)
    paper['health_domain'] = extract_mapped(paper, DOMAIN_MAP, ('title', 'abstract', 'mesh_terms', 'keywords'))
    paper['bias_axes'] = extract_mapped(paper, BIAS_AXIS_MAP)
    paper['lifecycle_stage'] = extract_mapped(paper, LIFECYCLE_STAGE_MAP)
    paper['assessment_or_mitigation'] = extract_assessment_or_mitigation(paper)
    paper['approach_method'] = extract_mapped(paper, APPROACH_MAP, ('abstract',))
    paper['clinical_setting'] = extract_mapped(paper, CLINICAL_SETTING_MAP)
    paper['key_findings'] = extract_key_findings(paper)


# ============================================================
# EXCEL BUILDING
# ============================================================

def build_excel(papers, config, pipeline_data):
    """Build a styled Excel with Summary, Criteria, and Papers sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    h_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    h_fill = PatternFill(start_color=config['header_color'], end_color=config['header_color'], fill_type='solid')
    h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_align = Alignment(vertical='top', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    alt_fill = PatternFill(start_color=config['alt_color'], end_color=config['alt_color'], fill_type='solid')
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color=config['header_color'])
    section_font = Font(bold=True, size=12, color=config['header_color'])

    p = pipeline_data

    # ---- Summary Sheet ----
    ws_s = wb.create_sheet('Summary', index=0)
    rows = [
        [config['title']],
        [''],
        ['PIPELINE OVERVIEW', ''],
        ['Source file:', p['source_file']],
        ['Source sheet:', p['source_sheet']],
        ['Database:', 'PubMed/MEDLINE (PMIDs fetched via NCBI E-utilities API)'],
        ['Date:', '2026-02-18'],
        [''],
        ['SCREENING FLOW', ''],
        ['PMIDs in source sheet:', str(p['total_pmids_in_sheet'])],
        ['Successfully fetched from NCBI:', str(p['total_fetched'])],
        ['Reviews removed:', str(p['reviews_removed'])],
        ['Papers for screening:', str(p['ta_screened'])],
        [''],
        ['Stage 1 — Title+Abstract Screening:', ''],
        ['  Papers screened:', str(p['ta_included_count'] + p['ta_excluded_count'])],
        ['  Included:', str(p['ta_included_count'])],
        ['  Excluded:', str(p['ta_excluded_count'])],
    ]
    for reason, count in sorted(p.get('ta_exclusion_reasons', {}).items(), key=lambda x: -x[1]):
        rows.append([f'    {reason}:', str(count)])

    rows.append([''])
    rows.append(['Stage 2 — Full-Text Screening:', ''])
    ft = p['phase2_fulltext_screening']
    rows.append(['  PMCIDs found:', str(p['phase1_pmcid_lookup']['pmcids_found'])])
    rows.append(['  No PMC full text:', str(p['phase1_pmcid_lookup']['no_pmcid'])])
    rows.append(['  Full texts screened:', str(ft['ft_included'] + ft['ft_excluded'])])
    rows.append(['  Included (2-stage):', str(ft['ft_included'])])
    rows.append(['  Excluded by full text:', str(ft['ft_excluded'])])
    rows.append(['  Fetch failed:', str(ft.get('ft_fetch_failed', 0))])
    for reason, count in sorted(ft.get('exclusion_reasons', {}).items(), key=lambda x: -x[1]):
        rows.append([f'    {reason}:', str(count)])

    rows.append([''])
    rows.append(['FINAL OUTPUT', ''])
    rows.append([f'  This file ({config["file_label"]}):', str(len(papers))])

    rows.append([''])
    rows.append(['RESEARCH QUESTIONS', ''])
    rows.append(['Primary:', 'What approaches are used to assess or mitigate algorithmic bias in health AI?'])
    rows.append(['Q1:', 'Are certain bias axes (race/ethnicity, gender, SES, language, age, disability) more commonly assessed?'])
    rows.append(['Q2:', 'At what point of the AI development lifecycle does assessment vs. mitigation happen?'])
    rows.append(['Q3:', 'How do approaches vary by clinical setting or context?'])

    for ri, rd in enumerate(rows, 1):
        for ci, v in enumerate(rd, 1):
            cell = ws_s.cell(row=ri, column=ci, value=v)
            if ri == 1:
                cell.font = title_font
            elif v in ['PIPELINE OVERVIEW', 'SCREENING FLOW', 'FINAL OUTPUT', 'RESEARCH QUESTIONS']:
                cell.font = section_font
            elif 'Stage 1' in str(v) or 'Stage 2' in str(v):
                cell.font = Font(bold=True, size=11, color='2F5496')
            elif ci == 1 and ':' in str(v) and not str(v).startswith('    '):
                cell.font = bold
    ws_s.column_dimensions['A'].width = 40
    ws_s.column_dimensions['B'].width = 90

    # ---- Criteria Sheet ----
    ws_c = wb.create_sheet('Screening_Criteria')
    criteria_rows = [
        ['SCREENING CRITERIA'],
        [''],
        ['STAGE 1: Title + Abstract Screening'],
        [''],
        ['Criterion 1 — AI/ML component required:', 'Paper must mention at least one AI/ML term'],
        ['  Terms:', ', '.join(p['screening_criteria']['ai_terms'])],
        [''],
        ['Criterion 2 — Health-related required:', 'Paper must mention at least one health term'],
        ['  Terms:', ', '.join(p['screening_criteria']['health_terms'])],
        [''],
        ['Criterion 3 — Bias/fairness must be central:', 'Strong bias term in title OR >= 2 AI bias terms in abstract'],
        ['  Title terms:', ', '.join(p['screening_criteria']['strong_title_terms'])],
        ['  Abstract terms:', ', '.join(p['screening_criteria']['ai_bias_terms'])],
        [''],
        ['Criterion 4 — Exclude human cognitive biases:', 'Exclude if only human bias terms AND no AI-specific bias terms'],
        ['  Human-only terms:', ', '.join(p['screening_criteria']['human_only_terms'])],
        ['  AI-specific override:', ', '.join(p['screening_criteria']['ai_specific_terms'])],
        [''],
        [''],
        ['STAGE 2: Full-Text Screening'],
        [''],
        ['Source:', 'Full text fetched from PubMed Central (PMC) via PMID -> PMCID conversion'],
        [''],
        ['Criteria:', 'Must have AI/ML terms AND health terms in full text, PLUS:'],
        ['  Option A:', 'approach_count >= 2 AND bias term in title'],
        ['  Option B:', 'approach_count >= 3 (regardless of title)'],
        ['  Option C:', 'bias term in title AND approach_count >= 1'],
        [''],
        ['Approach indicators:', ', '.join(p['ft_screening_criteria']['approach_indicators'])],
        [''],
        ['Bias title terms:', ', '.join(p['ft_screening_criteria']['bias_title_terms'])],
    ]
    for ri, rd in enumerate(criteria_rows, 1):
        for ci, v in enumerate(rd, 1):
            cell = ws_c.cell(row=ri, column=ci, value=v)
            if ri == 1:
                cell.font = title_font
            elif v in ['STAGE 1: Title + Abstract Screening', 'STAGE 2: Full-Text Screening']:
                cell.font = section_font
            elif ci == 1 and ':' in str(v) and not str(v).startswith('  '):
                cell.font = bold
    ws_c.column_dimensions['A'].width = 40
    ws_c.column_dimensions['B'].width = 120

    # ---- Papers Sheet ----
    headers = [
        ('No.', 5), ('PMID', 10), ('Title', 65), ('Authors', 20),
        ('Year', 6), ('Journal', 35), ('DOI', 30), ('URL', 35),
        ('Abstract', 80),
        ('Full-Text Status', 25),
        ('Study Type', 20), ('AI/ML Method', 30), ('Health Domain', 25),
        ('Bias Axes (Q1)', 30), ('Lifecycle Stage (Q2)', 25),
        ('Assessment vs Mitigation (Q2)', 22), ('Approach/Method', 35),
        ('Clinical Setting (Q3)', 25), ('Key Findings', 50),
        ('Keywords/MeSH', 35), ('Publication Type', 20),
    ]

    ws = wb.create_sheet('Papers')
    for ci, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    for idx, paper in enumerate(papers, 1):
        r = idx + 1
        vals = [
            idx, paper.get('pmid', ''), paper.get('title', ''), paper.get('authors', ''),
            paper.get('year', ''), paper.get('journal', ''), paper.get('doi', ''),
            paper.get('url', ''),
            (paper.get('abstract', '') or '')[:32000],
            paper.get('ft_status', ''),
            paper.get('study_type', ''), paper.get('ai_ml_method', ''),
            paper.get('health_domain', ''),
            paper.get('bias_axes', ''), paper.get('lifecycle_stage', ''),
            paper.get('assessment_or_mitigation', ''),
            paper.get('approach_method', ''),
            paper.get('clinical_setting', ''),
            paper.get('key_findings', ''),
            '; '.join(filter(None, [paper.get('keywords', ''), paper.get('mesh_terms', '')])),
            paper.get('pub_types', ''),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=str(v) if v else '')
            cell.alignment = c_align
            cell.border = border
        if idx % 2 == 0:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=r, column=ci).fill = alt_fill

    wb.save(config['output_path'])
    print(f"  Saved: {config['output_path']} ({len(papers)} papers, {len(wb.sheetnames)} sheets)")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 70)
    print("PubMed v2 STEP 4: Generate Final Excel Files")
    print("=" * 70)

    with open(INPUT_FILE) as f:
        data = json.load(f)

    ft_included = data['ft_included']
    no_fulltext = data['no_fulltext']

    print(f"\nExtracting structured columns...")
    for p in ft_included + no_fulltext:
        extract_all(p)

    ft_included.sort(key=lambda x: x.get('year', '0'), reverse=True)
    no_fulltext.sort(key=lambda x: x.get('year', '0'), reverse=True)

    print(f"  Extracted for {len(ft_included)} full-text papers")
    print(f"  Extracted for {len(no_fulltext)} no-full-text papers")

    # File 1: Full-text screened
    print(f"\nGenerating File 1: Full-Text Screened...")
    build_excel(ft_included, {
        'title': 'PubMed — Full-Text Screened Papers (2-Stage)',
        'file_label': '2-stage screened',
        'header_color': '2F5496',
        'alt_color': 'F2F7FB',
        'output_path': f'{BASE_DIR}/PubMed_FullText_Screened.xlsx',
    }, data)

    # File 2: No full text
    print(f"\nGenerating File 2: No Full Text (1-Stage Only)...")
    build_excel(no_fulltext, {
        'title': 'PubMed — No Full Text Available (1-Stage Only)',
        'file_label': '1-stage only, no full text',
        'header_color': '7B5B3A',
        'alt_color': 'FDF5EC',
        'output_path': f'{BASE_DIR}/PubMed_NoFullText.xlsx',
    }, data)

    # Print column summaries
    for label, papers in [('Full-Text Screened', ft_included), ('No Full Text', no_fulltext)]:
        print(f"\n{'='*70}")
        print(f"COLUMN SUMMARY: {label} ({len(papers)} papers)")
        print(f"{'='*70}")
        for col in ['study_type', 'ai_ml_method', 'health_domain', 'bias_axes',
                     'lifecycle_stage', 'assessment_or_mitigation', 'approach_method', 'clinical_setting']:
            vals = {}
            for p in papers:
                for v in p.get(col, 'Not specified').split('; '):
                    v = v.strip()
                    if v:
                        vals[v] = vals.get(v, 0) + 1
            print(f"\n  {col}:")
            for v, c in sorted(vals.items(), key=lambda x: -x[1])[:8]:
                print(f"    {c:>4}  {v}")

    print(f"\n{'='*70}")
    print(f"DONE — PubMed v2 (source: PubMed2 sheet, 1899 PMIDs)")
    print(f"{'='*70}")
    print(f"  Pipeline: {data['total_pmids_in_sheet']} PMIDs in sheet")
    print(f"         -> {data['total_fetched']} fetched")
    print(f"         -> {data['reviews_removed']} reviews removed")
    print(f"         -> {data['ta_screened']} screened (T/A)")
    print(f"         -> {data['ta_included_count']} T/A included")
    print(f"         -> {data['phase2_fulltext_screening']['ft_included']} FT included / {data['phase1_pmcid_lookup']['no_pmcid']} no FT")
    print(f"  File 1: PubMed_FullText_Screened.xlsx  ({len(ft_included)} papers)")
    print(f"  File 2: PubMed_NoFullText.xlsx         ({len(no_fulltext)} papers)")


if __name__ == '__main__':
    main()
