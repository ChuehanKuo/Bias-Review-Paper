#!/usr/bin/env python3
"""
Cross-Database Deduplication for Systematic Review:
"Approaches for Assessing and Mitigating Algorithmic Bias in Health AI"

Merges results from PubMed, Scopus, ACM, and IEEE.
Deduplicates by DOI (primary) and normalized title (secondary).
Produces a combined Excel workbook and PRISMA-style summary.
"""

import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = '/home/user/Bias-Review-Paper'

# Priority order: PubMed first (richest metadata), then Scopus, IEEE, ACM
DB_PRIORITY = ['pubmed', 'scopus', 'ieee', 'acm']

DB_FILES = {
    'pubmed': f'{BASE_DIR}/pubmed_screening_data.json',
    'scopus': f'{BASE_DIR}/scopus_screening_data.json',
    'acm': f'{BASE_DIR}/acm_screening_data.json',
    'ieee': f'{BASE_DIR}/ieee_screening_data.json',
}

DB_LABELS = {
    'pubmed': 'PubMed/MEDLINE',
    'scopus': 'Scopus',
    'acm': 'ACM Digital Library',
    'ieee': 'IEEE Xplore',
}


def normalize_doi(doi):
    """Normalize DOI for matching."""
    if not doi:
        return ''
    doi = doi.strip().lower()
    doi = doi.replace('https://doi.org/', '').replace('http://doi.org/', '')
    return doi


def normalize_title(title):
    """Normalize title for fuzzy matching."""
    if not title:
        return ''
    t = title.lower().strip()
    t = re.sub(r'[^a-z0-9\s]', '', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def load_db_results(json_path, db_name):
    """Load included papers from a database JSON file."""
    if not os.path.exists(json_path):
        print(f"  WARNING: {json_path} not found — skipping {db_name}")
        return [], {}

    with open(json_path) as f:
        data = json.load(f)

    included = data.get('included', data.get('ft_included', []))
    meta = data.get('meta', {})

    # Tag each paper with its source database
    for p in included:
        p['source_db'] = db_name
        p['source_db_label'] = DB_LABELS.get(db_name, db_name)

    return included, meta


def deduplicate(all_papers):
    """
    Deduplicate papers across databases.
    Priority: PubMed > Scopus > IEEE > ACM (keeps first seen).
    Returns (unique_papers, duplicate_records).
    """
    doi_index = {}
    title_index = {}
    unique = []
    duplicates = []

    for paper in all_papers:
        doi = normalize_doi(paper.get('doi', ''))
        title = normalize_title(paper.get('title', ''))
        is_dup = False
        dup_of = None

        # Check DOI match
        if doi and doi in doi_index:
            is_dup = True
            dup_of = doi_index[doi]
        # Check title match
        elif title and len(title) > 20 and title in title_index:
            is_dup = True
            dup_of = title_index[title]

        if is_dup:
            paper['duplicate_of'] = dup_of.get('title', '')[:80]
            paper['duplicate_of_db'] = dup_of.get('source_db_label', '')
            duplicates.append(paper)
        else:
            if doi:
                doi_index[doi] = paper
            if title and len(title) > 20:
                title_index[title] = paper
            unique.append(paper)

    return unique, duplicates


def build_combined_excel(unique, duplicates, db_stats):
    """Build combined Excel with all unique papers and dedup summary."""
    wb = Workbook()
    wb.remove(wb.active)

    h_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    h_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_align = Alignment(vertical='top', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    alt_fill = PatternFill(start_color='EDF2F9', end_color='EDF2F9', fill_type='solid')

    # --- PRISMA Summary ---
    ws_s = wb.create_sheet('PRISMA_Summary', index=0)
    summary_rows = [
        ['SYSTEMATIC REVIEW: Cross-Database Deduplication Summary'],
        [''],
        ['Review Title:', 'Approaches for Assessing and Mitigating Algorithmic Bias in Health AI'],
        ['Date:', '2026-02-17'],
        [''],
        ['DATABASE SEARCH RESULTS:', ''],
    ]
    for db in DB_PRIORITY:
        stats = db_stats.get(db, {})
        label = DB_LABELS.get(db, db)
        total = stats.get('total_unique', 0)
        included = stats.get('included', 0)
        summary_rows.append([f'  {label}:', f'{total} searched, {included} included after screening'])

    total_before_dedup = sum(s.get('included', 0) for s in db_stats.values())
    summary_rows.extend([
        [''],
        ['DEDUPLICATION:', ''],
        ['  Total papers before dedup:', str(total_before_dedup)],
        ['  Duplicates removed:', str(len(duplicates))],
        ['  Unique papers after dedup:', str(len(unique))],
        [''],
        ['FINAL UNIQUE PAPERS BY SOURCE:', ''],
    ])
    source_counts = {}
    for p in unique:
        db = p.get('source_db_label', 'Unknown')
        source_counts[db] = source_counts.get(db, 0) + 1
    for db_label in ['PubMed/MEDLINE', 'Scopus', 'IEEE Xplore', 'ACM Digital Library']:
        count = source_counts.get(db_label, 0)
        if count > 0:
            summary_rows.append([f'  {db_label}:', str(count)])

    for ri, rd in enumerate(summary_rows, 1):
        for ci, v in enumerate(rd, 1):
            cell = ws_s.cell(row=ri, column=ci, value=v)
            if ri == 1:
                cell.font = Font(bold=True, size=14, color='1F4E79')
            elif 'DEDUPLICATION' in str(v) or 'DATABASE SEARCH' in str(v) or 'FINAL UNIQUE' in str(v):
                cell.font = Font(bold=True, size=12, color='1F4E79')
            elif ci == 1 and ':' in str(v):
                cell.font = Font(bold=True)
    ws_s.column_dimensions['A'].width = 35
    ws_s.column_dimensions['B'].width = 85

    # --- All Unique Papers ---
    headers = [
        ('No.', 5), ('Source DB', 15), ('DOI', 30), ('Title', 65), ('Authors', 20),
        ('Year', 6), ('Journal/Venue', 35), ('URL', 35),
        ('Study Type', 20), ('AI/ML Method', 30), ('Health Domain', 25),
        ('Bias Axes (Q1)', 30), ('Lifecycle Stage (Q2)', 25),
        ('Assessment vs Mitigation (Q2)', 22), ('Approach/Method', 35),
        ('Clinical Setting (Q3)', 25), ('Key Findings', 50),
    ]
    ws = wb.create_sheet('All_Unique_Papers')
    for ci, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    for idx, p in enumerate(unique, 1):
        r = idx + 1
        vals = [
            idx, p.get('source_db_label', ''), p.get('doi', ''),
            p.get('title', ''), p.get('authors', ''),
            p.get('year', ''), p.get('journal', ''), p.get('url', ''),
            p.get('study_type', ''), p.get('ai_ml_method', ''),
            p.get('health_domain', ''),
            p.get('bias_axes', ''), p.get('lifecycle_stage', ''),
            p.get('assessment_or_mitigation', ''),
            p.get('approach_method', ''),
            p.get('clinical_setting', ''),
            p.get('key_findings', ''),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=str(v) if v else '')
            cell.alignment = c_align
            cell.border = border
        if idx % 2 == 0:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=r, column=ci).fill = alt_fill

    # --- Duplicates ---
    ws_d = wb.create_sheet('Duplicates_Removed')
    dh = [('No.', 5), ('Source DB', 15), ('DOI', 30), ('Title', 65),
          ('Year', 6), ('Duplicate Of', 65), ('Kept From', 15)]
    for ci, (h, w) in enumerate(dh, 1):
        cell = ws_d.cell(row=1, column=ci, value=h)
        cell.font = h_font
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = h_align
        cell.border = border
        ws_d.column_dimensions[get_column_letter(ci)].width = w
    ws_d.freeze_panes = 'A2'

    for idx, p in enumerate(duplicates, 1):
        vals = [idx, p.get('source_db_label', ''), p.get('doi', ''),
                p.get('title', ''), p.get('year', ''),
                p.get('duplicate_of', ''), p.get('duplicate_of_db', '')]
        for ci, v in enumerate(vals, 1):
            cell = ws_d.cell(row=idx + 1, column=ci, value=str(v) if v else '')
            cell.alignment = c_align
            cell.border = border

    # --- Per-database sheets ---
    for db in DB_PRIORITY:
        label = DB_LABELS.get(db, db).replace('/', '_').replace(' ', '_')
        db_papers = [p for p in unique if p.get('source_db') == db]
        if not db_papers:
            continue
        ws_db = wb.create_sheet(label)
        for ci, (h, w) in enumerate(headers, 1):
            cell = ws_db.cell(row=1, column=ci, value=h)
            cell.font = h_font
            cell.fill = h_fill
            cell.alignment = h_align
            cell.border = border
            ws_db.column_dimensions[get_column_letter(ci)].width = w
        ws_db.freeze_panes = 'A2'
        for idx, p in enumerate(db_papers, 1):
            r = idx + 1
            vals = [
                idx, p.get('source_db_label', ''), p.get('doi', ''),
                p.get('title', ''), p.get('authors', ''),
                p.get('year', ''), p.get('journal', ''), p.get('url', ''),
                p.get('study_type', ''), p.get('ai_ml_method', ''),
                p.get('health_domain', ''),
                p.get('bias_axes', ''), p.get('lifecycle_stage', ''),
                p.get('assessment_or_mitigation', ''),
                p.get('approach_method', ''),
                p.get('clinical_setting', ''),
                p.get('key_findings', ''),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws_db.cell(row=r, column=ci, value=str(v) if v else '')
                cell.alignment = c_align
                cell.border = border

    out = f'{BASE_DIR}/Combined_Deduplicated_Results.xlsx'
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Sheets: {wb.sheetnames}")
    return out


def main():
    print("=" * 70)
    print("Cross-Database Deduplication")
    print("=" * 70)

    # Load all database results
    all_papers = []
    db_stats = {}

    for db in DB_PRIORITY:
        path = DB_FILES[db]
        print(f"\nLoading {DB_LABELS[db]} from {path}...")
        included, meta = load_db_results(path, db)
        print(f"  Loaded: {len(included)} included papers")
        all_papers.extend(included)
        db_stats[db] = {
            'total_unique': meta.get('total_unique', 0),
            'included': len(included),
        }

    print(f"\nTotal papers before deduplication: {len(all_papers)}")

    # Deduplicate
    unique, duplicates = deduplicate(all_papers)
    print(f"Duplicates found: {len(duplicates)}")
    print(f"Unique papers: {len(unique)}")

    # Duplicate breakdown
    dup_by_db = {}
    for p in duplicates:
        db = p.get('source_db_label', 'Unknown')
        dup_by_db[db] = dup_by_db.get(db, 0) + 1
    if dup_by_db:
        print(f"\nDuplicates removed by source:")
        for db, c in sorted(dup_by_db.items(), key=lambda x: -x[1]):
            print(f"  {c:>4}  {db}")

    # Unique by source
    unique_by_db = {}
    for p in unique:
        db = p.get('source_db_label', 'Unknown')
        unique_by_db[db] = unique_by_db.get(db, 0) + 1
    print(f"\nUnique papers by source (after dedup):")
    for db_label in ['PubMed/MEDLINE', 'Scopus', 'IEEE Xplore', 'ACM Digital Library']:
        count = unique_by_db.get(db_label, 0)
        if count > 0:
            print(f"  {count:>4}  {db_label}")

    # Build combined Excel
    build_combined_excel(unique, duplicates, db_stats)

    # Save dedup JSON
    dedup_path = f'{BASE_DIR}/deduplicated_results.json'
    with open(dedup_path, 'w') as f:
        json.dump({
            'unique': unique,
            'duplicates': duplicates,
            'db_stats': db_stats,
            'total_before_dedup': len(all_papers),
            'total_after_dedup': len(unique),
            'total_duplicates': len(duplicates),
        }, f, indent=2)
    print(f"Saved JSON: {dedup_path}")

    # PRISMA-style summary
    print(f"\n{'='*70}")
    print("PRISMA FLOW SUMMARY")
    print(f"{'='*70}")
    for db in DB_PRIORITY:
        stats = db_stats.get(db, {})
        label = DB_LABELS.get(db, db)
        print(f"  {label}: {stats.get('total_unique', 0)} identified, {stats.get('included', 0)} included")
    print(f"\n  Total before dedup: {len(all_papers)}")
    print(f"  Duplicates removed: {len(duplicates)}")
    print(f"  FINAL UNIQUE INCLUDED: {len(unique)}")


if __name__ == '__main__':
    main()
