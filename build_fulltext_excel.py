#!/usr/bin/env python3
"""
Build Excel after full-text screening for systematic review.
Reads pubmed_fulltext_screening.json and produces updated Excel.
"""

import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styling
H_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
H_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
C_ALIGN = Alignment(vertical='top', wrap_text=True)
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
ALT_FILL = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

HEADERS = [
    ('No.', 5), ('PMID', 10), ('PMCID', 12), ('Title', 65), ('Authors', 20),
    ('Year', 6), ('Journal', 35), ('DOI', 30), ('URL', 35),
    ('Abstract', 80),
    ('FT Screening Source', 15), ('FT Screening Reason', 40),
    ('Study Type', 20), ('AI/ML Method', 30), ('Health Domain', 25),
    ('Bias Axes Assessed\n(Q1)', 30),
    ('AI Lifecycle Stage\n(Q2)', 25),
    ('Assessment vs Mitigation\n(Q2)', 22),
    ('Approach/Method', 35),
    ('Clinical Setting/Context\n(Q3)', 25),
    ('Key Findings', 50),
    ('Keywords/MeSH', 35), ('Publication Type', 20), ('Notes', 20),
]


def main():
    with open('/home/user/Bias-Review-Paper/pubmed_fulltext_screening.json', 'r') as f:
        data = json.load(f)

    included = data['ft_included']
    excluded = data['ft_excluded']
    query_stats = data.get('query_stats', [])
    meta = data['meta']

    # Sort included by year (newest first)
    included.sort(key=lambda x: x.get('year', '0'), reverse=True)

    print(f"Building Excel: {len(included)} included, {len(excluded)} excluded after full-text screening")

    wb = Workbook()
    wb.remove(wb.active)

    # ============================================================
    # SHEET 1: SUMMARY
    # ============================================================
    ws_s = wb.create_sheet('Summary', index=0)
    rows = [
        ['SYSTEMATIC REVIEW: PubMed/MEDLINE — Full-Text Screening Results'],
        [''],
        ['Review Title:', 'Approaches for Assessing and Mitigating Algorithmic Bias in Health AI'],
        ['Database:', 'PubMed/MEDLINE via NCBI E-utilities API'],
        ['Date:', '2026-02-16'],
        [''],
        ['SCREENING FLOW:', ''],
        ['  Title+Abstract Screening:', f'{meta["title_abstract_included"]} papers included'],
        ['  PMC Full Text Available:', f'{meta["pmc_available"]} papers ({meta["pmc_available"]}/{meta["title_abstract_included"]} = {100*meta["pmc_available"]//meta["title_abstract_included"]}%)'],
        ['  Full Texts Fetched:', f'{meta["fulltext_fetched"]} papers'],
        ['  Full-Text Screened (FT):', f'{meta["fulltext_fetched"]} papers (with full text)'],
        ['  Abstract-Only Screened:', f'{meta["title_abstract_included"] - meta["fulltext_fetched"]} papers (no PMC full text)'],
        [''],
        ['FINAL RESULTS:', ''],
        ['  INCLUDED:', str(meta['ft_included'])],
        ['  EXCLUDED:', str(meta['ft_excluded'])],
        [''],
        ['PRIMARY RESEARCH QUESTION:'],
        ['', 'What approaches are used to assess or mitigate algorithmic bias in health AI?'],
        [''],
        ['SECONDARY RESEARCH QUESTIONS:'],
        ['Q1:', 'Are certain bias axes (race/ethnicity, gender, SES, language, age, disability) more commonly assessed?'],
        ['Q2:', 'At what point of the AI development lifecycle does assessment vs. mitigation happen?'],
        ['Q3:', 'How do approaches vary by clinical setting or context?'],
        [''],
        ['FULL-TEXT INCLUSION CRITERIA:'],
        ['', '1. Specifically describes an approach/method for assessing OR mitigating algorithmic bias'],
        ['', '2. Bias/fairness is a central focus of the paper (not a side mention)'],
        ['', '3. Involves AI/ML algorithms in a health context'],
        [''],
        ['FULL-TEXT EXCLUSION CRITERIA:'],
        ['', 'Bias only mentioned in intro/limitations but paper is about something else'],
        ['', 'Paper about human biases, not algorithmic'],
        ['', 'General AI ethics without specific bias assessment/mitigation content'],
        ['', 'Only discusses bias conceptually without any approach/method'],
    ]
    for ri, rd in enumerate(rows, 1):
        for ci, v in enumerate(rd, 1):
            cell = ws_s.cell(row=ri, column=ci, value=v)
            if ri == 1:
                cell.font = Font(bold=True, size=14, color='2F5496')
            elif v in ['  INCLUDED:', '  EXCLUDED:']:
                cell.font = Font(bold=True, size=12, color='006600' if 'INCL' in v else 'CC0000')
            elif v in ['FINAL RESULTS:', 'SCREENING FLOW:']:
                cell.font = Font(bold=True, size=12, color='2F5496')
            elif ci == 1 and ':' in str(v):
                cell.font = Font(bold=True)
    ws_s.column_dimensions['A'].width = 30
    ws_s.column_dimensions['B'].width = 85

    # ============================================================
    # SHEET 2: INCLUDED PAPERS (after full-text screening)
    # ============================================================
    ws = wb.create_sheet('Included_Papers')
    for ci, (h, w) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = H_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    for idx, p in enumerate(included, 1):
        r = idx + 1
        vals = [
            idx, p.get('pmid', ''), p.get('pmcid', ''),
            p.get('title', ''), p.get('authors', ''),
            p.get('year', ''), p.get('journal', ''),
            p.get('doi', ''), p.get('url', ''),
            (p.get('abstract', '') or '')[:32000],
            p.get('ft_source', ''), p.get('ft_reason', ''),
            p.get('study_type', ''), p.get('ai_ml_method', ''),
            p.get('health_domain', ''),
            p.get('bias_axes', ''), p.get('lifecycle_stage', ''),
            p.get('assessment_or_mitigation', ''),
            p.get('approach_method', ''),
            p.get('clinical_setting', ''),
            p.get('key_findings', ''),
            p.get('keywords', ''), p.get('pub_types', ''), ''
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=str(v) if v else '')
            cell.alignment = C_ALIGN
            cell.border = BORDER
        if idx % 2 == 0:
            for ci in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=ci).fill = ALT_FILL

    # ============================================================
    # SHEET 3: EXCLUDED AT FULL-TEXT STAGE
    # ============================================================
    ws_e = wb.create_sheet('Excluded_FullText')
    eh = [('No.', 5), ('PMID', 10), ('Title', 65), ('Year', 6),
          ('Journal', 35), ('FT Source', 15), ('FT Exclusion Reason', 50)]
    for ci, (h, w) in enumerate(eh, 1):
        cell = ws_e.cell(row=1, column=ci, value=h)
        cell.font = H_FONT
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = H_ALIGN
        cell.border = BORDER
        ws_e.column_dimensions[get_column_letter(ci)].width = w
    ws_e.freeze_panes = 'A2'

    for idx, p in enumerate(excluded, 1):
        vals = [idx, p.get('pmid', ''), p.get('title', ''),
                p.get('year', ''), p.get('journal', ''),
                p.get('ft_source', ''), p.get('ft_reason', '')]
        for ci, v in enumerate(vals, 1):
            cell = ws_e.cell(row=idx + 1, column=ci, value=str(v) if v else '')
            cell.alignment = C_ALIGN
            cell.border = BORDER

    # ============================================================
    # SHEET 4: SEARCH STRATEGY
    # ============================================================
    ws_q = wb.create_sheet('Search_Strategy')
    qh = [('ID', 5), ('Label', 40), ('Query', 80), ('Total in PubMed', 15),
          ('Retrieved', 12), ('New Unique', 12), ('Cumulative', 12)]
    for ci, (h, w) in enumerate(qh, 1):
        cell = ws_q.cell(row=1, column=ci, value=h)
        cell.font = H_FONT
        cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
        cell.alignment = H_ALIGN
        cell.border = BORDER
        ws_q.column_dimensions[get_column_letter(ci)].width = w
    ws_q.freeze_panes = 'A2'

    for idx, qs in enumerate(query_stats, 1):
        vals = [qs.get('id', ''), qs.get('label', ''), qs.get('query', ''),
                qs.get('total_in_pubmed', ''), qs.get('retrieved', ''),
                qs.get('new_unique', ''), qs.get('cumulative', '')]
        for ci, v in enumerate(vals, 1):
            cell = ws_q.cell(row=idx + 1, column=ci, value=v)
            cell.alignment = C_ALIGN
            cell.border = BORDER

    # Save
    out = '/home/user/Bias-Review-Paper/PubMed_FullText_Screening_Results.xlsx'
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Included: {len(included)} papers")
    print(f"Excluded at full-text stage: {len(excluded)} papers")

    # Year distribution
    years = {}
    for p in included:
        y = p.get('year', 'Unknown')
        years[y] = years.get(y, 0) + 1
    print(f"\nIncluded papers by year:")
    for y in sorted(years.keys(), reverse=True):
        print(f"  {y}: {years[y]}")


if __name__ == '__main__':
    main()
