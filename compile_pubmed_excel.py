#!/usr/bin/env python3
"""
Compile PubMed-only Excel for the systematic review.
336 title-filtered papers with all available metadata.
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

HEADERS = [
    'No.', 'PMID', 'Title', 'Year', 'Journal', 'URL', 'DOI',
    'Relevance Filter',
    'Include/Exclude (Abstract Screen)',
    'Study Type',
    'AI/ML Method',
    'Health Domain',
    'Bias Axes Assessed (Q1)',
    'AI Lifecycle Stage (Q2)',
    'Assessment vs Mitigation (Q2)',
    'Approach/Method',
    'Clinical Setting/Context (Q3)',
    'Key Findings',
    'Notes'
]

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = 'PubMed_MEDLINE'

    # Header row
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    ws.freeze_panes = 'A2'

    # Read PubMed filtered CSV
    csv_path = '/home/user/Bias-Review-Paper/pubmed_filtered_results.csv'
    row_num = 2
    with open(csv_path, 'r') as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader, 1):
            data = [
                idx,
                row.get('PMID', ''),
                row.get('Title', ''),
                row.get('Year', ''),
                row.get('Journal', ''),
                row.get('URL', ''),
                row.get('DOI', ''),
                row.get('RelevanceReason', ''),
                '',  # Include/Exclude - to be filled
                '',  # Study Type
                '',  # AI/ML Method
                '',  # Health Domain
                '',  # Bias Axes
                '',  # Lifecycle Stage
                '',  # Assessment vs Mitigation
                '',  # Approach
                '',  # Clinical Setting
                '',  # Key Findings
                '',  # Notes
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=str(value))
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
            row_num += 1

    # Auto-width columns
    col_widths = {
        1: 6, 2: 12, 3: 60, 4: 6, 5: 40, 6: 40, 7: 35,
        8: 15, 9: 20, 10: 20, 11: 25, 12: 25, 13: 30,
        14: 25, 15: 20, 16: 30, 17: 25, 18: 50, 19: 30
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Add summary info
    ws_summary = wb.create_sheet(title='Summary', index=0)
    summary_data = [
        ['Systematic Review Database Screening: PubMed/MEDLINE'],
        [''],
        ['Review Title:', 'Approaches for Assessing and Mitigating Algorithmic Bias in Health AI'],
        ['Database:', 'PubMed/MEDLINE via NCBI E-utilities API'],
        ['Search Date:', '2026-02-16'],
        ['Total Papers Retrieved:', '1,523'],
        ['Title-Filtered Papers:', str(row_num - 2)],
        ['Search Queries Used:', '50 query variations'],
        [''],
        ['Research Questions:'],
        ['Q1:', 'Are certain bias axes (race/ethnicity, gender, SES, language, age, disability) more commonly assessed?'],
        ['Q2:', 'At what point of the AI development lifecycle does assessment vs. mitigation happen?'],
        ['Q3:', 'How do approaches vary by clinical setting or context?'],
        [''],
        ['Inclusion Criteria:'],
        ['', 'Papers about bias/fairness in ML/AI algorithms applied to health/healthcare/medicine'],
        ['', 'All years, English language, all study types'],
        [''],
        ['Exclusion Criteria:'],
        ['', 'Cognitive/human biases NOT about AI'],
        ['', 'Statistical bias without AI/ML involvement'],
        ['', 'Papers not related to health/healthcare/medicine'],
        [''],
        ['Column Descriptions:'],
        ['Include/Exclude:', 'To be filled during abstract screening - mark Include or Exclude with reason'],
        ['Study Type:', 'Empirical / Framework / Review / Commentary / Guideline / Methodology'],
        ['AI/ML Method:', 'e.g., Deep learning, NLP, Clinical prediction, Imaging AI, LLM'],
        ['Health Domain:', 'e.g., Radiology, Dermatology, Cardiology, EHR, Mental health'],
        ['Bias Axes (Q1):', 'Race, Gender, Age, SES, Language, Disability, Insurance, Geographic'],
        ['Lifecycle Stage (Q2):', 'Data collection, Preprocessing, Model development, Evaluation, Deployment, Monitoring'],
        ['Assessment vs Mitigation (Q2):', 'Assessment only / Mitigation only / Both'],
        ['Clinical Setting (Q3):', 'e.g., Hospital, Primary care, Public health, Emergency'],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif col_idx == 1 and ':' in str(value):
                cell.font = Font(bold=True)
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 80

    output_path = '/home/user/Bias-Review-Paper/PubMed_Screening_Results.xlsx'
    wb.save(output_path)
    print(f"PubMed Excel saved to: {output_path}")
    print(f"Total papers: {row_num - 2}")

if __name__ == '__main__':
    main()
