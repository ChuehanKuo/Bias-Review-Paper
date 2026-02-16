#!/usr/bin/env python3
"""
PubMed Final Pipeline for Systematic Review:
"Approaches for Assessing and Mitigating Algorithmic Bias in Health AI"

1. Search with corrected, focused queries (no cap)
2. Deduplicate
3. Fetch title + abstract
4. Screen by title+abstract
5. Extract ALL structured columns from abstract text
6. Build Excel
"""

import json
import urllib.request
import urllib.parse
import time
import xml.etree.ElementTree as ET
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# QUERIES — corrected and focused
# ============================================================

QUERIES = [
    {
        "id": "Q1",
        "label": "Core: algorithmic bias/fairness terms + health",
        "query": (
            '("algorithmic bias" OR "algorithmic fairness" OR "AI bias" OR "AI fairness" '
            'OR "machine learning bias" OR "machine learning fairness" OR "bias mitigation" '
            'OR "debiasing" OR "fairness-aware" OR "bias detection" OR "bias audit" '
            'OR "disparate impact" OR "demographic parity" OR "equalized odds") '
            'AND '
            '("health" OR "healthcare" OR "clinical" OR "medical" OR "biomedical")'
        )
    },
    {
        "id": "Q2",
        "label": "Bias assessment/mitigation approaches in health AI",
        "query": (
            '("bias" OR "fairness") '
            'AND ("assess" OR "mitigat" OR "detect" OR "evaluat" OR "framework" OR "approach" OR "method") '
            'AND ("artificial intelligence" OR "machine learning" OR "deep learning" OR "algorithm") '
            'AND ("health" OR "healthcare" OR "clinical" OR "medical")'
        )
    },
    {
        "id": "Q3",
        "label": "Specific bias axes in clinical AI",
        "query": (
            '("racial bias" OR "gender bias" OR "age bias" OR "socioeconomic bias" OR "ethnic bias") '
            'AND '
            '("artificial intelligence" OR "machine learning" OR "deep learning" '
            'OR "clinical prediction" OR "clinical algorithm" OR "clinical decision support")'
        )
    }
]


# ============================================================
# PUBMED API FUNCTIONS
# ============================================================

def search_pubmed(query, retmax=10000):
    encoded = urllib.parse.quote(query)
    count_url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=pubmed&term={encoded}&retmax=0&retmode=json"
    try:
        with urllib.request.urlopen(urllib.request.Request(count_url), timeout=30) as resp:
            data = json.loads(resp.read().decode('utf-8'))
        total = int(data.get('esearchresult', {}).get('count', '0'))
    except Exception as e:
        print(f"  ERROR: {e}")
        return [], 0

    all_ids = []
    for start in range(0, total, retmax):
        url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=pubmed&term={encoded}&retmax={retmax}&retstart={start}&retmode=json"
        try:
            with urllib.request.urlopen(urllib.request.Request(url), timeout=30) as resp:
                data = json.loads(resp.read().decode('utf-8'))
            all_ids.extend(data.get('esearchresult', {}).get('idlist', []))
        except Exception as e:
            print(f"  ERROR at offset {start}: {e}")
        time.sleep(0.4)

    return all_ids, total


def fetch_papers(pmids, batch_size=200):
    papers = {}
    total_batches = (len(pmids) - 1) // batch_size + 1
    for i in range(0, len(pmids), batch_size):
        batch = pmids[i:i+batch_size]
        ids_str = ','.join(batch)
        url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?db=pubmed&id={ids_str}&rettype=xml&retmode=xml"
        try:
            with urllib.request.urlopen(urllib.request.Request(url), timeout=60) as resp:
                xml_data = resp.read().decode('utf-8')
            root = ET.fromstring(xml_data)
            for article in root.findall('.//PubmedArticle'):
                pmid_elem = article.find('.//PMID')
                if pmid_elem is None:
                    continue
                pmid = pmid_elem.text
                title_elem = article.find('.//ArticleTitle')
                title = ''.join(title_elem.itertext()) if title_elem is not None else ''
                abstract_parts = []
                for at in article.findall('.//AbstractText'):
                    label = at.get('Label', '')
                    text = ''.join(at.itertext())
                    abstract_parts.append(f"{label}: {text}" if label else text)
                abstract = ' '.join(abstract_parts)
                journal_elem = article.find('.//Journal/Title')
                journal = journal_elem.text if journal_elem is not None else ''
                year = ''
                for yp in ['.//PubDate/Year', './/ArticleDate/Year', './/PubDate/MedlineDate']:
                    ye = article.find(yp)
                    if ye is not None and ye.text:
                        year = ye.text[:4]
                        break
                doi = ''
                for aid in article.findall('.//ArticleId'):
                    if aid.get('IdType') == 'doi':
                        doi = aid.text
                        break
                if not doi:
                    for eloc in article.findall('.//ELocationID'):
                        if eloc.get('EIdType') == 'doi':
                            doi = eloc.text
                            break
                authors = []
                for au in article.findall('.//Author'):
                    last = au.find('LastName')
                    fore = au.find('ForeName')
                    if last is not None and last.text:
                        name = last.text + (f" {fore.text}" if fore is not None and fore.text else '')
                        authors.append(name)
                author_str = authors[0] + ' et al.' if len(authors) > 1 else (authors[0] if authors else 'N/A')
                keywords = [kw.text for kw in article.findall('.//Keyword') if kw.text]
                mesh = [m.text for m in article.findall('.//MeshHeading/DescriptorName') if m.text]
                pub_types = [pt.text for pt in article.findall('.//PublicationType') if pt.text]

                papers[pmid] = {
                    'pmid': pmid, 'title': title, 'abstract': abstract,
                    'journal': journal, 'year': year, 'doi': doi,
                    'authors': author_str,
                    'keywords': '; '.join(keywords[:20]),
                    'mesh_terms': '; '.join(mesh[:20]),
                    'pub_types': '; '.join(pub_types),
                    'url': f'https://pubmed.ncbi.nlm.nih.gov/{pmid}/'
                }
            print(f"  Batch {i//batch_size+1}/{total_batches} done ({len(papers)} total)")
        except Exception as e:
            print(f"  ERROR batch {i//batch_size+1}: {e}")
        time.sleep(0.4)
    return papers


# ============================================================
# SCREENING
# ============================================================

def screen_paper(paper):
    title = (paper.get('title', '') or '').lower()
    abstract = (paper.get('abstract', '') or '').lower()
    combined = title + ' ' + abstract
    kw = (paper.get('keywords', '') or '').lower()
    mesh = (paper.get('mesh_terms', '') or '').lower()
    all_text = combined + ' ' + kw + ' ' + mesh

    ai_terms = [
        'machine learning', 'deep learning', 'artificial intelligence', 'neural network',
        'algorithm', 'predictive model', 'prediction model', 'classifier', 'classification',
        'natural language processing', 'nlp', 'computer vision', 'random forest',
        'logistic regression', 'xgboost', 'convolutional', 'transformer',
        'large language model', 'llm', 'decision support', 'risk prediction',
        'federated learning', 'reinforcement learning', 'foundation model',
        'chatgpt', 'gpt-4', 'supervised learning'
    ]
    has_ai = any(t in all_text for t in ai_terms)

    health_terms = [
        'health', 'clinical', 'medical', 'patient', 'hospital', 'disease',
        'diagnosis', 'treatment', 'care', 'pathology', 'radiology', 'dermatology',
        'cardiology', 'oncology', 'ophthalmology', 'psychiatry', 'mental health',
        'ehr', 'electronic health', 'biomedical', 'mortality', 'readmission',
        'sepsis', 'icu', 'emergency', 'chest x-ray', 'mammograph', 'cancer',
        'diabetes', 'cardiovascular', 'public health', 'healthcare', 'medicine'
    ]
    has_health = any(t in all_text for t in health_terms)

    # Bias/fairness must be central
    strong_in_title = any(t in title for t in [
        'bias', 'fairness', 'fair ', 'unfair', 'equitable', 'equity',
        'disparity', 'disparities', 'discrimination', 'debiasing', 'debias',
        'underdiagnos', 'underrepresent', 'inequit'
    ])
    ai_bias_terms = [
        'algorithmic bias', 'algorithmic fairness', 'ai bias', 'ai fairness',
        'machine learning bias', 'machine learning fairness', 'model bias',
        'bias mitigation', 'bias detection', 'bias assessment', 'fairness-aware',
        'fair machine learning', 'debiasing', 'disparate impact', 'demographic parity',
        'equalized odds', 'equal opportunity', 'fairness metric', 'fairness evaluation',
        'bias in ai', 'bias in machine learning', 'bias in algorithm', 'biased model',
        'biased prediction', 'mitigating bias', 'addressing bias', 'reducing bias',
        'assessing bias', 'evaluating bias', 'detecting bias', 'sources of bias',
        'racial bias', 'gender bias', 'age bias', 'socioeconomic bias', 'ethnic bias',
        'underdiagnosis bias', 'representation bias', 'data bias', 'label bias',
        'health disparit', 'health equit', 'health inequit', 'unfair'
    ]
    bias_count = sum(1 for t in ai_bias_terms if t in abstract)
    has_substantial_bias = strong_in_title or bias_count >= 2

    if not has_ai:
        return False, 'No AI/ML component'
    if not has_health:
        return False, 'Not health-related'
    if not has_substantial_bias:
        return False, 'Bias/fairness not central topic'

    # Exclude human cognitive biases only
    human_only = ['implicit bias training', 'implicit bias among', 'clinician bias',
                  'physician bias', 'provider bias', 'unconscious bias training',
                  'implicit association test', 'weight bias', 'anti-fat bias']
    ai_specific = ['algorithmic bias', 'ai bias', 'model bias', 'bias mitigation',
                   'bias in ai', 'bias in machine learning', 'algorithmic fairness',
                   'fairness-aware', 'debiasing algorithm']
    if any(t in all_text for t in human_only) and not any(t in all_text for t in ai_specific):
        return False, 'Human cognitive biases only'

    return True, 'Included'


# ============================================================
# COLUMN EXTRACTION FROM ABSTRACT
# ============================================================

def extract_study_type(paper):
    title = (paper.get('title', '') or '').lower()
    abstract = (paper.get('abstract', '') or '').lower()
    pub_types = (paper.get('pub_types', '') or '').lower()
    combined = title + ' ' + abstract + ' ' + pub_types

    if any(t in combined for t in ['systematic review', 'prisma', 'systematic literature review']):
        return 'Systematic Review'
    if any(t in combined for t in ['scoping review', 'scoping study']):
        return 'Scoping Review'
    if any(t in combined for t in ['narrative review', 'literature review', 'review article']):
        return 'Narrative Review'
    if 'meta-analysis' in combined:
        return 'Meta-Analysis'
    if any(t in combined for t in ['survey', 'questionnaire', 'interview']):
        return 'Survey/Qualitative'
    if any(t in combined for t in ['commentary', 'editorial', 'perspective', 'viewpoint', 'opinion', 'letter']):
        return 'Commentary/Editorial'
    if any(t in combined for t in ['guideline', 'recommendation', 'consensus', 'policy']):
        return 'Guideline/Policy'
    if any(t in combined for t in ['framework', 'toolkit', 'toolbox', 'pipeline']):
        return 'Framework/Toolkit'
    if any(t in combined for t in ['we propose', 'we develop', 'we introduce', 'novel method', 'novel approach', 'we present a method', 'we present a novel']):
        return 'Methodology'
    if any(t in combined for t in ['experiment', 'dataset', 'we evaluated', 'we trained', 'we tested', 'we applied', 'we analyzed', 'we assessed', 'results show', 'our results', 'we found']):
        return 'Empirical Study'
    if 'review' in combined:
        return 'Review'
    return 'Empirical Study'


def extract_ai_ml_method(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    combined = title + ' ' + abstract

    methods = []
    method_map = {
        'Deep Learning': ['deep learning', 'deep neural', 'cnn', 'convolutional neural', 'resnet', 'densenet', 'vgg', 'inception'],
        'NLP/LLM': ['natural language processing', 'nlp', 'large language model', 'llm', 'chatgpt', 'gpt-4', 'gpt-3', 'bert', 'transformer', 'text mining', 'text classification'],
        'Random Forest': ['random forest'],
        'Logistic Regression': ['logistic regression'],
        'XGBoost/Gradient Boosting': ['xgboost', 'gradient boosting', 'lightgbm', 'catboost'],
        'Support Vector Machine': ['support vector', 'svm'],
        'Neural Network': ['neural network', 'mlp', 'perceptron'],
        'Decision Tree': ['decision tree', 'cart'],
        'Ensemble Methods': ['ensemble', 'bagging', 'stacking'],
        'Federated Learning': ['federated learning'],
        'Reinforcement Learning': ['reinforcement learning'],
        'Computer Vision/Imaging AI': ['computer vision', 'image classification', 'image segmentation', 'object detection', 'medical imaging'],
        'Clinical Prediction Model': ['clinical prediction', 'risk prediction', 'risk score', 'prediction model', 'predictive model', 'prognostic model'],
        'Clinical Decision Support': ['clinical decision support', 'decision support system', 'cdss'],
        'Foundation Model': ['foundation model'],
        'Generative AI': ['generative ai', 'generative model', 'gan', 'generative adversarial'],
        'Clustering': ['clustering', 'k-means', 'unsupervised'],
        'Regression': ['linear regression', 'regression model'],
        'Survival Analysis': ['survival analysis', 'cox regression', 'hazard model'],
    }
    for label, terms in method_map.items():
        if any(t in combined for t in terms):
            methods.append(label)

    return '; '.join(methods) if methods else 'Not specified'


def extract_health_domain(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    mesh = (paper.get('mesh_terms', '') or '').lower()
    combined = title + ' ' + abstract + ' ' + mesh

    domains = []
    domain_map = {
        'Radiology/Medical Imaging': ['radiology', 'radiograph', 'x-ray', 'ct scan', 'mri', 'mammograph', 'ultrasound', 'medical imaging', 'chest radiograph'],
        'Dermatology': ['dermatology', 'skin lesion', 'skin cancer', 'melanoma', 'skin disease', 'dermatoscop'],
        'Cardiology': ['cardiology', 'cardiovascular', 'heart', 'cardiac', 'ecg', 'ekg', 'echocardiogram', 'atrial fibrillation', 'coronary'],
        'Oncology': ['oncology', 'cancer', 'tumor', 'tumour', 'neoplasm', 'malignant', 'chemotherapy', 'radiation therapy'],
        'Ophthalmology': ['ophthalmology', 'retinal', 'retinopathy', 'diabetic retinopathy', 'glaucoma', 'fundus', 'optic'],
        'Mental Health/Psychiatry': ['mental health', 'psychiatry', 'depression', 'anxiety', 'suicide', 'psychiatric', 'behavioral health', 'substance abuse', 'opioid'],
        'Emergency Medicine': ['emergency department', 'emergency medicine', 'trauma', 'triage', 'acute care'],
        'ICU/Critical Care': ['icu', 'intensive care', 'critical care', 'sepsis', 'mechanical ventilation'],
        'EHR/Health Informatics': ['electronic health record', 'ehr', 'emr', 'health information', 'clinical informatics', 'health informatics'],
        'Primary Care': ['primary care', 'family medicine', 'general practice', 'ambulatory'],
        'Surgery': ['surgery', 'surgical', 'operative', 'postoperative'],
        'Pathology': ['pathology', 'histopathology', 'cytology', 'biopsy'],
        'Pediatrics': ['pediatric', 'paediatric', 'child', 'neonatal', 'newborn'],
        'Obstetrics/Maternal Health': ['obstetric', 'maternal', 'pregnancy', 'prenatal', 'perinatal', 'childbirth'],
        'Nephrology': ['kidney', 'renal', 'nephrology', 'dialysis'],
        'Pulmonology': ['pulmonary', 'respiratory', 'lung', 'pneumonia', 'covid-19', 'asthma', 'copd'],
        'Neurology': ['neurology', 'neurological', 'stroke', 'alzheimer', 'dementia', 'brain', 'epilepsy'],
        'Genomics/Genetics': ['genomic', 'genetic', 'genome', 'gene expression', 'variant', 'sequencing'],
        'Public Health': ['public health', 'population health', 'epidemiology', 'surveillance', 'community health'],
        'Drug Discovery/Pharmacology': ['drug discovery', 'pharmacology', 'pharmaceutical', 'drug development'],
        'Pain Management': ['pain', 'analgesic', 'opioid prescribing'],
        'Infectious Disease': ['infectious disease', 'infection', 'hiv', 'tuberculosis', 'malaria', 'hepatitis'],
        'Endocrinology/Diabetes': ['diabetes', 'diabetic', 'endocrin', 'insulin', 'glucose', 'hba1c'],
        'Wearables/Remote Monitoring': ['wearable', 'pulse oximetry', 'spo2', 'remote monitoring', 'mobile health', 'mhealth'],
    }
    for label, terms in domain_map.items():
        if any(t in combined for t in terms):
            domains.append(label)

    return '; '.join(domains) if domains else 'General Healthcare'


def extract_bias_axes(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    combined = title + ' ' + abstract

    axes = []
    axis_map = {
        'Race/Ethnicity': ['race', 'racial', 'ethnicity', 'ethnic', 'african american', 'black', 'white', 'hispanic', 'latino', 'asian', 'indigenous', 'minority', 'skin color', 'skin tone', 'skin pigment'],
        'Gender/Sex': ['gender', 'sex', 'male', 'female', 'women', 'men', 'transgender', 'sex-based', 'gender-based'],
        'Age': ['age', 'aging', 'elderly', 'older adult', 'pediatric', 'geriatric', 'age-related', 'age group'],
        'Socioeconomic Status': ['socioeconomic', 'income', 'poverty', 'insurance', 'uninsured', 'underinsured', 'medicaid', 'sociodemographic', 'education level', 'employment'],
        'Language': ['language', 'non-english', 'english proficiency', 'limited english', 'linguistic'],
        'Disability': ['disability', 'disabled', 'impairment', 'handicap'],
        'Geographic': ['geographic', 'rural', 'urban', 'region', 'low-resource', 'low-income country', 'global south', 'developing country'],
        'Insurance Status': ['insurance', 'uninsured', 'underinsured', 'medicaid', 'medicare', 'payer'],
        'Intersectional': ['intersectional', 'intersectionality', 'multiple sensitive', 'multi-attribute', 'compound'],
    }
    for label, terms in axis_map.items():
        if any(t in combined for t in terms):
            axes.append(label)

    return '; '.join(axes) if axes else 'Not specified'


def extract_lifecycle_stage(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    combined = title + ' ' + abstract

    stages = []
    stage_map = {
        'Data Collection': ['data collection', 'data acquisition', 'dataset creation', 'data gathering', 'sampling', 'cohort selection', 'data source'],
        'Data Preprocessing': ['preprocess', 'pre-process', 'data cleaning', 'imputation', 'feature engineering', 'data augmentation', 'resampling', 'oversampling', 'undersampling', 'reweighting', 'reweighing'],
        'Model Development/Training': ['model training', 'model development', 'training phase', 'in-processing', 'adversarial debiasing', 'fairness constraint', 'regularization', 'fair representation', 'calibration'],
        'Model Evaluation': ['model evaluation', 'evaluation', 'testing', 'validation', 'performance assessment', 'fairness metric', 'bias assessment', 'bias detection', 'audit', 'post-processing', 'subgroup analysis'],
        'Deployment': ['deployment', 'clinical implementation', 'real-world', 'production', 'clinical practice', 'clinical workflow', 'monitoring'],
    }
    for label, terms in stage_map.items():
        if any(t in combined for t in terms):
            stages.append(label)

    return '; '.join(stages) if stages else 'Not specified'


def extract_assessment_or_mitigation(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    combined = title + ' ' + abstract

    assess_terms = ['assess', 'detect', 'evaluat', 'audit', 'measur', 'quantif', 'identif', 'characteriz', 'examin', 'analyz', 'investigat']
    mitigate_terms = ['mitigat', 'debias', 'reduc', 'eliminat', 'correct', 'remov', 'address', 'alleviat', 'improv fairness', 'fairness-aware', 'fair machine', 'reweigh', 'adversarial debias']

    has_assess = any(t in combined for t in assess_terms)
    has_mitigate = any(t in combined for t in mitigate_terms)

    if has_assess and has_mitigate:
        return 'Both'
    elif has_mitigate:
        return 'Mitigation'
    elif has_assess:
        return 'Assessment'
    return 'Not specified'


def extract_approach_method(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    combined = abstract

    approaches = []
    approach_map = {
        'Reweighting/Resampling': ['reweighting', 'reweighing', 'resampling', 'oversampling', 'undersampling', 'smote'],
        'Adversarial Debiasing': ['adversarial debiasing', 'adversarial learning', 'adversarial training'],
        'Fairness Constraints': ['fairness constraint', 'fairness regulariz', 'fairness-aware training'],
        'Calibration': ['calibration', 'recalibration', 'platt scaling'],
        'Threshold Adjustment': ['threshold', 'cutoff', 'operating point'],
        'Data Augmentation': ['data augmentation', 'synthetic data', 'augmented data', 'fair mixup'],
        'Subgroup Analysis': ['subgroup analysis', 'stratified analysis', 'subpopulation', 'disaggregated'],
        'Fairness Metrics Evaluation': ['demographic parity', 'equalized odds', 'equal opportunity', 'disparate impact', 'fairness metric', 'predictive parity', 'calibration across'],
        'Counterfactual Fairness': ['counterfactual', 'causal inference', 'causal fairness'],
        'Representation Learning': ['representation learning', 'fair representation', 'disentangl', 'embedding'],
        'Transfer Learning': ['transfer learning', 'domain adaptation', 'fine-tun'],
        'Federated Learning': ['federated learning', 'federated model'],
        'Explainability/Interpretability': ['explainab', 'interpretab', 'shap', 'lime', 'attention'],
        'Post-hoc Correction': ['post-processing', 'post-hoc', 'reject option'],
        'Diverse/Representative Data': ['diverse dataset', 'representative data', 'inclusive data', 'data diversity', 'balanced dataset'],
        'Bias Auditing Framework': ['audit', 'bias audit', 'fairness audit', 'model card', 'datasheet'],
        'Multi-task Learning': ['multi-task', 'multitask'],
        'Ensemble Methods': ['ensemble', 'model ensemble'],
        'Regularization': ['regulariz', 'penalty', 'constraint-based'],
    }
    for label, terms in approach_map.items():
        if any(t in combined for t in terms):
            approaches.append(label)

    return '; '.join(approaches) if approaches else 'Not specified'


def extract_clinical_setting(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    combined = title + ' ' + abstract

    settings = []
    setting_map = {
        'Hospital/Inpatient': ['hospital', 'inpatient', 'hospitalized', 'admitted'],
        'ICU': ['icu', 'intensive care', 'critical care'],
        'Emergency Department': ['emergency department', 'emergency room', 'ed visit'],
        'Primary Care/Outpatient': ['primary care', 'outpatient', 'ambulatory', 'clinic visit', 'office visit'],
        'Public Health/Population': ['public health', 'population', 'community', 'screening program', 'surveillance'],
        'Telehealth/Remote': ['telehealth', 'telemedicine', 'remote', 'mobile health', 'mhealth', 'wearable'],
        'Clinical Trial': ['clinical trial', 'randomized controlled', 'rct'],
        'Laboratory/Pathology': ['laboratory', 'pathology', 'lab test'],
        'Imaging Center': ['imaging center', 'radiology department'],
        'Safety-Net/Underserved': ['safety net', 'safety-net', 'underserved', 'low-resource'],
        'Long-term Care': ['long-term care', 'nursing home', 'assisted living'],
    }
    for label, terms in setting_map.items():
        if any(t in combined for t in terms):
            settings.append(label)

    return '; '.join(settings) if settings else 'Not specified'


def extract_key_findings(paper):
    abstract = (paper.get('abstract', '') or '')
    if not abstract:
        return 'No abstract available'

    # Try to find conclusion/results section
    lower = abstract.lower()
    for marker in ['CONCLUSION:', 'CONCLUSIONS:', 'RESULTS:', 'FINDINGS:', 'DISCUSSION:']:
        idx = lower.find(marker.lower())
        if idx != -1:
            finding = abstract[idx:].strip()
            # Take first 2 sentences
            sentences = re.split(r'(?<=[.!?])\s+', finding)
            result = ' '.join(sentences[:3])
            if len(result) > 500:
                result = result[:500] + '...'
            return result

    # Fallback: last 2 sentences
    sentences = re.split(r'(?<=[.!?])\s+', abstract)
    if len(sentences) >= 2:
        result = ' '.join(sentences[-2:])
    else:
        result = abstract
    if len(result) > 500:
        result = result[:500] + '...'
    return result


# ============================================================
# EXCEL BUILDING
# ============================================================

HEADERS = [
    ('No.', 5), ('PMID', 10), ('Title', 65), ('Authors', 20),
    ('Year', 6), ('Journal', 35), ('DOI', 30), ('URL', 35),
    ('Abstract', 80),
    ('Study Type', 20), ('AI/ML Method', 30), ('Health Domain', 25),
    ('Bias Axes Assessed\n(Q1)', 30),
    ('AI Lifecycle Stage\n(Q2)', 25),
    ('Assessment vs Mitigation\n(Q2)', 22),
    ('Approach/Method', 35),
    ('Clinical Setting/Context\n(Q3)', 25),
    ('Key Findings', 50),
    ('Keywords/MeSH', 35), ('Publication Type', 20), ('Notes', 20),
]

H_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
H_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
C_ALIGN = Alignment(vertical='top', wrap_text=True)
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
ALT_FILL = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')


def build_excel(included, excluded, query_stats, meta):
    wb = Workbook()
    wb.remove(wb.active)

    # --- Summary ---
    ws_s = wb.create_sheet('Summary', index=0)
    rows = [
        ['SYSTEMATIC REVIEW: PubMed/MEDLINE Database Screening'],
        [''],
        ['Review Title:', 'Approaches for Assessing and Mitigating Algorithmic Bias in Health AI'],
        ['Database:', 'PubMed/MEDLINE via NCBI E-utilities API'],
        ['Date:', '2026-02-16'],
        ['Queries:', f'{len(query_stats)} focused Boolean queries (no retrieval cap)'],
        ['Total Unique PMIDs:', str(meta['total_unique'])],
        ['Fetched:', str(meta['total_fetched'])],
        ['Screened:', str(len(included) + len(excluded))],
        ['INCLUDED:', str(len(included))],
        ['EXCLUDED:', str(len(excluded))],
        [''],
        ['PRIMARY RESEARCH QUESTION:'],
        ['', 'What approaches are used to assess or mitigate algorithmic bias in health AI?'],
        [''],
        ['SECONDARY RESEARCH QUESTIONS:'],
        ['Q1:', 'Are certain bias axes (race/ethnicity, gender, SES, language, age, disability) more commonly assessed?'],
        ['Q2:', 'At what point of the AI development lifecycle does assessment vs. mitigation happen?'],
        ['Q3:', 'How do approaches vary by clinical setting or context?'],
        [''],
        ['INCLUSION:', 'Papers about bias/fairness in ML/AI algorithms applied to health (central topic)'],
        ['EXCLUSION:', 'Human cognitive biases, statistical bias without AI/ML, bias mentioned in passing only'],
    ]
    for ri, rd in enumerate(rows, 1):
        for ci, v in enumerate(rd, 1):
            cell = ws_s.cell(row=ri, column=ci, value=v)
            if ri == 1: cell.font = Font(bold=True, size=14, color='2F5496')
            elif v in ['INCLUDED:', 'EXCLUDED:']: cell.font = Font(bold=True, size=12, color='006600' if 'INCL' in v else 'CC0000')
            elif ci == 1 and ':' in str(v): cell.font = Font(bold=True)
    ws_s.column_dimensions['A'].width = 22
    ws_s.column_dimensions['B'].width = 85

    # --- Included Papers ---
    ws = wb.create_sheet('Included_Papers')
    for ci, (h, w) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = H_ALIGN; cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    for idx, p in enumerate(included, 1):
        r = idx + 1
        vals = [
            idx, p.get('pmid',''), p.get('title',''), p.get('authors',''),
            p.get('year',''), p.get('journal',''), p.get('doi',''), p.get('url',''),
            (p.get('abstract','') or '')[:32000],
            p.get('study_type',''), p.get('ai_ml_method',''), p.get('health_domain',''),
            p.get('bias_axes',''), p.get('lifecycle_stage',''),
            p.get('assessment_or_mitigation',''), p.get('approach_method',''),
            p.get('clinical_setting',''), p.get('key_findings',''),
            p.get('keywords',''), p.get('pub_types',''), ''
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=str(v) if v else '')
            cell.alignment = C_ALIGN; cell.border = BORDER
        if idx % 2 == 0:
            for ci in range(1, len(HEADERS)+1):
                ws.cell(row=r, column=ci).fill = ALT_FILL

    # --- Excluded ---
    ws_e = wb.create_sheet('Excluded_Papers')
    eh = [('No.', 5), ('PMID', 10), ('Title', 65), ('Year', 6), ('Journal', 35), ('Reason', 50)]
    for ci, (h, w) in enumerate(eh, 1):
        cell = ws_e.cell(row=1, column=ci, value=h)
        cell.font = H_FONT
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = H_ALIGN; cell.border = BORDER
        ws_e.column_dimensions[get_column_letter(ci)].width = w
    ws_e.freeze_panes = 'A2'
    for idx, p in enumerate(excluded, 1):
        vals = [idx, p.get('pmid',''), p.get('title',''), p.get('year',''), p.get('journal',''), p.get('screen_reason','')]
        for ci, v in enumerate(vals, 1):
            cell = ws_e.cell(row=idx+1, column=ci, value=str(v) if v else '')
            cell.alignment = C_ALIGN; cell.border = BORDER

    # --- Search Strategy ---
    ws_q = wb.create_sheet('Search_Strategy')
    qh = [('ID', 5), ('Label', 40), ('Query', 80), ('Total in PubMed', 15), ('Retrieved', 12), ('New Unique', 12), ('Cumulative', 12)]
    for ci, (h, w) in enumerate(qh, 1):
        cell = ws_q.cell(row=1, column=ci, value=h)
        cell.font = H_FONT
        cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
        cell.alignment = H_ALIGN; cell.border = BORDER
        ws_q.column_dimensions[get_column_letter(ci)].width = w
    ws_q.freeze_panes = 'A2'
    for idx, qs in enumerate(query_stats, 1):
        vals = [qs['id'], qs['label'], qs['query'], qs['total_in_pubmed'], qs['retrieved'], qs['new_unique'], qs['cumulative']]
        for ci, v in enumerate(vals, 1):
            cell = ws_q.cell(row=idx+1, column=ci, value=v)
            cell.alignment = C_ALIGN; cell.border = BORDER

    out = '/home/user/Bias-Review-Paper/PubMed_Screening_Results.xlsx'
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Sheets: {wb.sheetnames}")
    return out


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 70)
    print("PubMed Final Pipeline — Systematic Review")
    print("=" * 70)

    # Step 1: Search
    print(f"\nSTEP 1: Searching PubMed ({len(QUERIES)} queries, no cap)...")
    all_pmids = set()
    query_stats = []
    for q in QUERIES:
        print(f"\n  {q['id']}: {q['label']}")
        ids, total = search_pubmed(q['query'])
        new = set(ids) - all_pmids
        all_pmids.update(ids)
        query_stats.append({
            'id': q['id'], 'label': q['label'], 'query': q['query'],
            'total_in_pubmed': total, 'retrieved': len(ids),
            'new_unique': len(new), 'cumulative': len(all_pmids)
        })
        print(f"  Total: {total} | Retrieved: {len(ids)} | New: {len(new)} | Cumulative: {len(all_pmids)}")
        time.sleep(0.5)
    print(f"\n  Total unique PMIDs (deduplicated): {len(all_pmids)}")

    # Step 2: Fetch
    print(f"\nSTEP 2: Fetching titles + abstracts...")
    papers = fetch_papers(sorted(all_pmids))
    print(f"  Fetched: {len(papers)}")

    # Step 3: Screen
    print(f"\nSTEP 3: Screening by title+abstract...")
    included = []
    excluded = []
    for pmid, p in papers.items():
        inc, reason = screen_paper(p)
        p['include'] = inc
        p['screen_reason'] = reason
        if inc:
            included.append(p)
        else:
            excluded.append(p)
    print(f"  INCLUDED: {len(included)} | EXCLUDED: {len(excluded)}")

    # Step 4: Extract columns
    print(f"\nSTEP 4: Extracting structured data from abstracts...")
    for p in included:
        p['study_type'] = extract_study_type(p)
        p['ai_ml_method'] = extract_ai_ml_method(p)
        p['health_domain'] = extract_health_domain(p)
        p['bias_axes'] = extract_bias_axes(p)
        p['lifecycle_stage'] = extract_lifecycle_stage(p)
        p['assessment_or_mitigation'] = extract_assessment_or_mitigation(p)
        p['approach_method'] = extract_approach_method(p)
        p['clinical_setting'] = extract_clinical_setting(p)
        p['key_findings'] = extract_key_findings(p)
    print(f"  Extracted all columns for {len(included)} papers")

    # Sort by year (newest first)
    included.sort(key=lambda x: x.get('year', '0'), reverse=True)

    # Step 5: Build Excel
    print(f"\nSTEP 5: Building Excel...")
    meta = {'total_unique': len(all_pmids), 'total_fetched': len(papers)}
    build_excel(included, excluded, query_stats, meta)

    # Save JSON backup
    with open('/home/user/Bias-Review-Paper/pubmed_final_data.json', 'w') as f:
        json.dump({
            'included': included, 'excluded': excluded,
            'query_stats': query_stats, 'meta': meta
        }, f, indent=2)
    print("Saved JSON backup: pubmed_final_data.json")

    # Quick stats
    print(f"\n{'='*70}")
    print("COLUMN EXTRACTION SUMMARY")
    print(f"{'='*70}")
    for col in ['study_type', 'ai_ml_method', 'health_domain', 'bias_axes',
                'lifecycle_stage', 'assessment_or_mitigation', 'approach_method', 'clinical_setting']:
        vals = {}
        for p in included:
            for v in p.get(col, 'Not specified').split('; '):
                v = v.strip()
                if v:
                    vals[v] = vals.get(v, 0) + 1
        print(f"\n{col}:")
        for v, c in sorted(vals.items(), key=lambda x: -x[1])[:10]:
            print(f"  {c:>4}  {v}")


if __name__ == '__main__':
    main()
