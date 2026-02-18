#!/usr/bin/env python3
"""
Filter Combined Deduplicated Results:
- Remove review papers (Systematic Review, Scoping Review, Narrative Review,
  Meta-Analysis, generic Review)
- Remove papers that passed through without full-text screening
- Save removed papers to a separate Excel file
- Save filtered papers to an updated Combined file
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = '/home/user/Bias-Review-Paper'

REVIEW_TYPES = {
    'Systematic Review', 'Scoping Review', 'Narrative Review',
    'Meta-Analysis', 'Review',
}


def is_review(paper):
    st = paper.get('study_type', '')
    return st in REVIEW_TYPES


def is_passthrough(paper):
    """Paper included without actual full-text screening."""
    ft = (paper.get('ft_status', '') or '').lower()
    return 'passed through' in ft


def build_excel(papers, title, out_path, sheet_name='Papers'):
    """Build a styled Excel workbook for a list of papers."""
    wb = Workbook()
    wb.remove(wb.active)

    h_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    h_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_align = Alignment(vertical='top', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    alt_fill = PatternFill(start_color='EDF2F9', end_color='EDF2F9', fill_type='solid')

    # --- Summary ---
    ws_s = wb.create_sheet('Summary', index=0)
    review_count = sum(1 for p in papers if is_review(p))
    passthrough_count = sum(1 for p in papers if is_passthrough(p) and not is_review(p))
    both_count = sum(1 for p in papers if is_review(p) and is_passthrough(p))

    summary_rows = [
        [title],
        [''],
        ['Total papers in this file:', str(len(papers))],
        ['  Review papers:', str(review_count + both_count)],
        ['  Pass-through only (no full-text screening):', str(passthrough_count)],
        ['  Both review AND pass-through:', str(both_count)],
        [''],
        ['Removal reasons:'],
        ['  Review types removed:', ', '.join(sorted(REVIEW_TYPES))],
        ['  Pass-through:', 'Papers with ft_status containing "passed through"'],
    ]
    for ri, rd in enumerate(summary_rows, 1):
        for ci, v in enumerate(rd, 1):
            cell = ws_s.cell(row=ri, column=ci, value=v)
            if ri == 1:
                cell.font = Font(bold=True, size=14, color='1F4E79')
            elif ci == 1 and ':' in str(v):
                cell.font = Font(bold=True)
    ws_s.column_dimensions['A'].width = 45
    ws_s.column_dimensions['B'].width = 85

    # --- Papers sheet ---
    headers = [
        ('No.', 5), ('Source DB', 15), ('DOI', 30), ('Title', 65), ('Authors', 20),
        ('Year', 6), ('Journal/Venue', 35), ('URL', 35),
        ('Full-Text Status', 25), ('Removal Reason', 20),
        ('Study Type', 20), ('AI/ML Method', 30), ('Health Domain', 25),
        ('Bias Axes (Q1)', 30), ('Lifecycle Stage (Q2)', 25),
        ('Assessment vs Mitigation (Q2)', 22), ('Approach/Method', 35),
        ('Clinical Setting (Q3)', 25), ('Key Findings', 50),
    ]
    ws = wb.create_sheet(sheet_name)
    for ci, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    for idx, p in enumerate(papers, 1):
        r = idx + 1
        # Determine removal reason
        reasons = []
        if is_review(p):
            reasons.append(f'Review ({p.get("study_type", "")})')
        if is_passthrough(p):
            reasons.append('No full-text screening')
        removal = '; '.join(reasons)

        vals = [
            idx, p.get('source_db_label', ''), p.get('doi', ''),
            p.get('title', ''), p.get('authors', ''),
            p.get('year', ''), p.get('journal', ''), p.get('url', ''),
            p.get('ft_status', ''), removal,
            p.get('study_type', ''), p.get('ai_ml_method', ''),
            p.get('health_domain', ''),
            p.get('bias_axes', ''), p.get('lifecycle_stage', ''),
            p.get('assessment_or_mitigation', ''),
            p.get('approach_method', ''),
            p.get('clinical_setting', ''),
            p.get('key_findings', ''),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=str(v) if v else '')
            cell.alignment = c_align
            cell.border = border
        if idx % 2 == 0:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=r, column=ci).fill = alt_fill

    wb.save(out_path)
    print(f"Saved: {out_path}  ({len(papers)} papers)")


def rebuild_combined_excel(unique, db_stats):
    """Rebuild Combined_Deduplicated_Results.xlsx with filtered papers only."""
    from deduplicate import build_combined_excel
    # We don't have duplicates anymore, pass empty list
    # Actually let's load the original duplicates
    dedup_path = f'{BASE_DIR}/deduplicated_results.json'
    with open(dedup_path) as f:
        orig = json.load(f)
    duplicates = orig.get('duplicates', [])
    build_combined_excel(unique, duplicates, db_stats)


def main():
    print("=" * 70)
    print("Filtering: Remove Reviews & Pass-Through Papers")
    print("=" * 70)

    # Load deduplicated results
    dedup_path = f'{BASE_DIR}/deduplicated_results.json'
    with open(dedup_path) as f:
        data = json.load(f)

    unique = data['unique']
    db_stats = data.get('db_stats', {})

    print(f"\nTotal papers before filtering: {len(unique)}")

    # Classify each paper
    removed = []
    kept = []

    review_count = 0
    passthrough_count = 0
    both_count = 0

    for p in unique:
        rev = is_review(p)
        pt = is_passthrough(p)

        if rev or pt:
            removed.append(p)
            if rev and pt:
                both_count += 1
            elif rev:
                review_count += 1
            else:
                passthrough_count += 1
        else:
            kept.append(p)

    print(f"\nRemoved: {len(removed)}")
    print(f"  Review papers: {review_count}")
    print(f"  Pass-through (no full-text screening): {passthrough_count}")
    print(f"  Both review AND pass-through: {both_count}")
    print(f"\nKept: {len(kept)}")

    # Breakdown by study type for removed
    type_counts = {}
    for p in removed:
        st = p.get('study_type', 'Unknown')
        type_counts[st] = type_counts.get(st, 0) + 1
    print(f"\nRemoved papers by study type:")
    for st, c in sorted(type_counts.items(), key=lambda x: -x[1]):
        print(f"  {c:>4}  {st}")

    # Save removed papers to separate Excel
    removed_path = f'{BASE_DIR}/Removed_Papers.xlsx'
    build_excel(removed, 'Removed Papers: Reviews & No Full-Text Screening', removed_path,
                sheet_name='Removed_Papers')

    # Update deduplicated JSON with filtered results
    filtered_path = f'{BASE_DIR}/deduplicated_results.json'
    data['unique'] = kept
    data['removed'] = removed
    data['total_after_filtering'] = len(kept)
    data['total_removed'] = len(removed)
    data['removal_breakdown'] = {
        'reviews': review_count,
        'passthrough': passthrough_count,
        'both': both_count,
    }
    with open(filtered_path, 'w') as f:
        json.dump(data, f, indent=2)
    print(f"\nUpdated: {filtered_path}")

    # Rebuild combined Excel with filtered data
    from deduplicate import build_combined_excel
    duplicates = data.get('duplicates', [])
    build_combined_excel(kept, duplicates, db_stats)

    # Final summary
    kept_by_db = {}
    for p in kept:
        db = p.get('source_db_label', 'Unknown')
        kept_by_db[db] = kept_by_db.get(db, 0) + 1
    print(f"\n{'='*70}")
    print(f"FINAL RESULTS")
    print(f"{'='*70}")
    print(f"Original unique papers: {len(unique)}")
    print(f"Removed (reviews + pass-through): {len(removed)}")
    print(f"Final included papers: {len(kept)}")
    print(f"\nBy source:")
    for db_label in ['PubMed/MEDLINE', 'Scopus', 'IEEE Xplore', 'ACM Digital Library']:
        count = kept_by_db.get(db_label, 0)
        if count > 0:
            print(f"  {count:>4}  {db_label}")

    # Study type distribution of kept papers
    kept_types = {}
    for p in kept:
        st = p.get('study_type', 'Unknown')
        kept_types[st] = kept_types.get(st, 0) + 1
    print(f"\nKept papers by study type:")
    for st, c in sorted(kept_types.items(), key=lambda x: -x[1]):
        print(f"  {c:>4}  {st}")


if __name__ == '__main__':
    main()
