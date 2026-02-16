#!/usr/bin/env python3
"""
Compile all database screening results into a single Excel workbook
with separate sheets per database for the systematic review:
"Approaches for Assessing and Mitigating Algorithmic Bias in Health AI"
"""

import csv
import sys
import os

# Add project dir to path
sys.path.insert(0, '/home/user/Bias-Review-Paper')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# STYLING
# ============================================================
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column headers for deeply-screened sheets
DEEP_HEADERS = [
    'No.', 'Title', 'Authors', 'Year', 'Journal/Conference', 'DOI',
    'Study Type', 'AI/ML Method', 'Health Domain',
    'Bias Axes Assessed (Q1)', 'AI Lifecycle Stage (Q2)',
    'Assessment vs Mitigation (Q2)', 'Approach/Method',
    'Clinical Setting/Context (Q3)', 'Key Findings'
]

# Column headers for title-screened sheets (from expanded searches)
TITLE_HEADERS = [
    'No.', 'Title', 'Year', 'Journal/Conference/Venue', 'URL/DOI',
    'Category/Topic Cluster', 'Needs Abstract Screening'
]

# Column headers for PubMed (from CSV with metadata)
PUBMED_HEADERS = [
    'No.', 'PMID', 'Title', 'Year', 'Journal', 'URL', 'DOI',
    'Relevance Filter', 'Needs Abstract Screening'
]

def style_header(ws, headers):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

def style_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = CELL_ALIGNMENT
    cell.border = THIN_BORDER
    return cell

def auto_width(ws, headers, max_width=60):
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx-1]))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def add_deep_screened_sheet(wb, sheet_name, results):
    """Add a sheet with deeply screened papers (full structured data)."""
    ws = wb.create_sheet(title=sheet_name)
    style_header(ws, DEEP_HEADERS)
    ws.freeze_panes = 'A2'

    for idx, paper in enumerate(results, 1):
        bias_axes = paper.get('bias_axes', 'N/A')
        if isinstance(bias_axes, list):
            bias_axes = ', '.join(bias_axes)

        row_data = [
            idx,
            paper.get('title', 'N/A'),
            paper.get('authors', 'N/A'),
            paper.get('year', 'N/A'),
            paper.get('journal_conference', paper.get('journal', 'N/A')),
            paper.get('doi', paper.get('arxiv_id', 'N/A')),
            paper.get('study_type', 'N/A'),
            paper.get('ai_ml_method', 'N/A'),
            paper.get('health_domain', 'N/A'),
            bias_axes,
            paper.get('lifecycle_stage', 'N/A'),
            paper.get('assessment_or_mitigation', 'N/A'),
            paper.get('approach_method', 'N/A'),
            paper.get('clinical_setting', 'N/A'),
            paper.get('key_findings', 'N/A')
        ]
        for col_idx, value in enumerate(row_data, 1):
            style_cell(ws, idx + 1, col_idx, str(value) if value is not None else 'N/A')

    auto_width(ws, DEEP_HEADERS)
    return ws


def add_pubmed_sheet(wb):
    """Add PubMed sheet from CSV file."""
    ws = wb.create_sheet(title='PubMed_MEDLINE')
    style_header(ws, PUBMED_HEADERS)
    ws.freeze_panes = 'A2'

    csv_path = '/home/user/Bias-Review-Paper/pubmed_filtered_results.csv'
    row_num = 2
    with open(csv_path, 'r') as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader, 1):
            row_data = [
                idx,
                row.get('PMID', ''),
                row.get('Title', ''),
                row.get('Year', ''),
                row.get('Journal', ''),
                row.get('URL', ''),
                row.get('DOI', ''),
                row.get('RelevanceReason', ''),
                'Yes'
            ]
            for col_idx, value in enumerate(row_data, 1):
                style_cell(ws, row_num, col_idx, str(value))
            row_num += 1

    auto_width(ws, PUBMED_HEADERS)
    return ws, row_num - 2  # return count


def add_expanded_ieee_acm_sheet(wb):
    """Add expanded IEEE+ACM results from the Wave 2 agent."""
    # IEEE expanded
    ieee_expanded = [
        # Bias Assessment & Fairness Frameworks
        {"title": "Algorithmic Bias in Clinical Populations—Evaluating and Improving Facial Analysis Technology in Older Adults With Dementia", "url": "https://ieeexplore.ieee.org/document/8643365/", "year": 2019, "venue": "IEEE Access", "cluster": "Bias Assessment"},
        {"title": "Framework for Algorithmic Bias Quantification and its Application to Automated Sleep Scoring", "url": "https://ieeexplore.ieee.org/document/10675993/", "year": 2024, "venue": "IEEE SDS Conference", "cluster": "Bias Assessment"},
        {"title": "Systems Analysis of Bias and Risk in AI-Enabled Medical Diagnosis", "url": "https://ieeexplore.ieee.org/document/10371919/", "year": 2023, "venue": "IEEE Conference", "cluster": "Bias Assessment"},
        {"title": "Mitigating Racial Algorithmic Bias in Healthcare AI Systems: A Fairness-Aware Machine Learning Approach", "url": "https://ieeexplore.ieee.org/document/10788666/", "year": 2024, "venue": "IEEE ICSSA Conference", "cluster": "Bias Mitigation"},
        {"title": "Fairness in Healthcare: Assessing Data Bias and Algorithmic Fairness", "url": "https://ieeexplore.ieee.org/document/10783630/", "year": 2024, "venue": "IEEE SIPAIM", "cluster": "Bias Assessment"},
        {"title": "Debias-CLR: A Contrastive Learning Based Debiasing Method for Algorithmic Fairness in Healthcare Applications", "url": "https://ieeexplore.ieee.org/document/10825827/", "year": 2024, "venue": "IEEE Conference", "cluster": "Bias Mitigation"},
        {"title": "Integrated Framework for Equitable Healthcare AI: Bias Mitigation, Community Participation, and Regulatory Governance", "url": "https://ieeexplore.ieee.org/document/10968102/", "year": 2024, "venue": "IEEE CSNT Conference", "cluster": "Bias Mitigation Framework"},
        {"title": "Fair Machine Learning in Healthcare: A Survey", "url": "https://ieeexplore.ieee.org/document/10700762/", "year": 2024, "venue": "IEEE Journal", "cluster": "Survey"},
        {"title": "Towards a Single Goodness Metric of Clinically Relevant, Accurate, Fair and Unbiased ML Predictions", "url": "https://ieeexplore.ieee.org/document/10297674/", "year": 2023, "venue": "IEEE Conference", "cluster": "Fairness Metrics"},
        {"title": "Connecting Fairness in Machine Learning with Public Health Equity", "url": "https://ieeexplore.ieee.org/document/10337160/", "year": 2023, "venue": "IEEE Conference", "cluster": "Health Equity"},
        {"title": "Fairness Artificial Intelligence in Clinical Decision Support: Mitigating Effect of Health Disparity", "url": "https://ieeexplore.ieee.org/document/10786197/", "year": 2024, "venue": "IEEE JBHI", "cluster": "Clinical Decision Support"},
        {"title": "Towards Fairness and Interpretability: Clinical Decision Support for Acute Coronary Syndrome", "url": "https://ieeexplore.ieee.org/document/10069672/", "year": 2023, "venue": "IEEE Conference", "cluster": "Clinical Decision Support"},
        {"title": "Exploring Bias and Prediction Metrics to Characterise the Fairness of ML for Equity-Centered Public Health Decision-Making", "url": "https://ieeexplore.ieee.org/document/10771762/", "year": 2024, "venue": "IEEE Access", "cluster": "Fairness Metrics"},
        {"title": "Framework for Fairness in Machine Learning Using Detecting and Mitigating Bias in AI Algorithms", "url": "https://ieeexplore.ieee.org/document/11258168/", "year": 2025, "venue": "IEEE Conference", "cluster": "Bias Mitigation Framework"},
        {"title": "A Survey of Bias and Fairness in Healthcare AI", "url": "https://ieeexplore.ieee.org/document/10628826/", "year": 2024, "venue": "IEEE ICHI", "cluster": "Survey"},
        {"title": "Fairness in Healthcare AI", "url": "https://ieeexplore.ieee.org/document/9565788/", "year": 2021, "venue": "IEEE ICHI", "cluster": "Survey"},
        {"title": "Ethical AI with Balancing Bias Mitigation and Fairness in Machine Learning Models", "url": "https://ieeexplore.ieee.org/document/10749873/", "year": 2024, "venue": "IEEE FRUCT", "cluster": "Bias Mitigation"},
        {"title": "IEEE Standard 7003-2024: Algorithmic Bias Considerations", "url": "https://ieeexplore.ieee.org/document/10851955/", "year": 2024, "venue": "IEEE Standard", "cluster": "Standards/Regulation"},
        {"title": "Assessing Algorithmic Bias in Machine Learning Classifiers: A Fairness Evaluation", "url": "https://ieeexplore.ieee.org/document/10842230/", "year": 2024, "venue": "IEEE Conference", "cluster": "Bias Assessment"},
        # Bias Mitigation Methods
        {"title": "Hierarchical Bias Mitigation for Semi-Supervised Medical Image Classification (HABIT)", "url": "https://ieeexplore.ieee.org/document/10049601/", "year": 2023, "venue": "IEEE Journal", "cluster": "Medical Imaging"},
        {"title": "Towards Equitable Diagnosis: Bias Evaluation and Mitigation in Skin Cancer Classification", "url": "https://ieeexplore.ieee.org/document/10702173/", "year": 2024, "venue": "IEEE Conference", "cluster": "Dermatology"},
        {"title": "Underdiagnosis Bias Mitigation With Expert Foundation Model's Representation", "url": "https://ieeexplore.ieee.org/document/11048878/", "year": 2025, "venue": "IEEE Journal", "cluster": "Medical Imaging"},
        {"title": "Ethical AI Auditor for Bias Detecting in AI Models Using Adversarial Debiasing", "url": "https://ieeexplore.ieee.org/document/11081330/", "year": 2025, "venue": "IEEE ICHI", "cluster": "Bias Mitigation"},
        {"title": "Comprehensive Bias Mitigation in AI: Evaluating Pre-Processing, In-Processing, and Post-Processing", "url": "https://ieeexplore.ieee.org/document/11141086/", "year": 2025, "venue": "IEEE Conference", "cluster": "Bias Mitigation"},
        {"title": "Comparative Analysis of Pre-Processing, In-processing and Post-Processing Methods for Bias Mitigation", "url": "https://ieeexplore.ieee.org/document/11115514/", "year": 2025, "venue": "IEEE Conference", "cluster": "Bias Mitigation"},
        {"title": "Post-Processing Fairness Evaluation of Federated Models: An Unsupervised Approach in Healthcare", "url": "https://ieeexplore.ieee.org/document/10107702/", "year": 2023, "venue": "IEEE/ACM TCBB", "cluster": "Federated Learning"},
        # Adversarial Debiasing
        {"title": "Adversarial Debiasing for Equitable and Fair Detection of Acute Coronary Syndrome Using 12-Lead ECG", "url": "https://ieeexplore.ieee.org/document/11122264/", "year": 2025, "venue": "IEEE Journal", "cluster": "Cardiology"},
        {"title": "Adversarial Debiasing Techniques Towards 'Fair' Skin Lesion Classification", "url": "https://ieeexplore.ieee.org/document/10123788/", "year": 2023, "venue": "IEEE Conference", "cluster": "Dermatology"},
        # Racial Bias in Clinical AI
        {"title": "Addressing Racial Bias in Cardiovascular Disease Risk Prediction with Fair Data Augmentation", "url": "https://ieeexplore.ieee.org/document/10450015/", "year": 2024, "venue": "IEEE ICCINS", "cluster": "Cardiovascular"},
        {"title": "Mitigating Bias in Opportunistic Screening for MACE With Causal Reasoning", "url": "https://ieeexplore.ieee.org/document/10993302/", "year": 2025, "venue": "IEEE Journal", "cluster": "Cardiovascular"},
        {"title": "Racial Bias in Multimodal Classification of Healthcare Data—Evidence from MIMIC-IV", "url": "https://ieeexplore.ieee.org/document/11144427/", "year": 2025, "venue": "IEEE Conference", "cluster": "EHR/ICU"},
        {"title": "Mitigating Racial Bias in Chest X-Ray Disease Diagnosis Via Disentanglement Learning", "url": "https://ieeexplore.ieee.org/document/10635819/", "year": 2024, "venue": "IEEE Conference", "cluster": "Radiology"},
        {"title": "Reducing Racial Bias in SpO2 Estimation: The Effects of Skin Pigmentation", "url": "https://ieeexplore.ieee.org/document/10341069/", "year": 2023, "venue": "IEEE Conference", "cluster": "Wearables/Pulse Oximetry"},
        # Skin Lesion / Dermatology
        {"title": "Assessing Bias in Skin Lesion Classifiers With Contemporary Deep Learning and Post-Hoc Explainability", "url": "https://ieeexplore.ieee.org/document/10162202/", "year": 2023, "venue": "IEEE Journal", "cluster": "Dermatology"},
        {"title": "(De)Constructing Bias on Skin Lesion Datasets", "url": "https://ieeexplore.ieee.org/document/9025695/", "year": 2020, "venue": "IEEE Conference", "cluster": "Dermatology"},
        {"title": "Mitigating Racial Bias in Skin Lesion Classification with a Novel Deep Learning-Driven Dataset", "url": "https://ieeexplore.ieee.org/document/11113692/", "year": 2024, "venue": "IEEE Conference", "cluster": "Dermatology"},
        {"title": "Evaluating the Impact of Skin Tone Representation on OOD Detection in Dermatology", "url": "https://ieeexplore.ieee.org/document/10635847/", "year": 2024, "venue": "IEEE Conference", "cluster": "Dermatology"},
        {"title": "Bias Inheritance and its Amplification in GAN-Based Synthetic Data Augmentation for Skin Lesion Classification", "url": "https://ieeexplore.ieee.org/document/10574714/", "year": 2024, "venue": "IEEE Conference", "cluster": "Dermatology"},
        {"title": "Debiasing Skin Lesion Datasets and Models? Not So Fast", "url": "https://ieeexplore.ieee.org/document/9150714/", "year": 2020, "venue": "IEEE Conference", "cluster": "Dermatology"},
        # EHR / Clinical Prediction
        {"title": "A Counterfactual Fair Model for Longitudinal EHR via Deconfounder (FLMD)", "url": "https://ieeexplore.ieee.org/document/10415843/", "year": 2024, "venue": "IEEE Conference", "cluster": "EHR"},
        # Mortality / Readmission / ICU
        {"title": "Unbiased Mortality Prediction for Unbalanced Data Using Machine Learning", "url": "https://ieeexplore.ieee.org/document/8980003/", "year": 2020, "venue": "IEEE Conference", "cluster": "ICU/Mortality"},
        {"title": "Auditing ICU Readmission Rates in a Clinical Database", "url": "https://ieeexplore.ieee.org/document/10337287/", "year": 2023, "venue": "IEEE Conference", "cluster": "ICU/Readmission"},
        # Sepsis
        {"title": "A Reinforcement Learning Approach for Predicting Onset of Septic Shock with Unfair Bias", "url": "https://ieeexplore.ieee.org/document/10731539/", "year": 2024, "venue": "IEEE Conference", "cluster": "Sepsis"},
        # Mental Health
        {"title": "Bias Reducing Multitask Learning on Mental Health Prediction", "url": "https://ieeexplore.ieee.org/document/9953850/", "year": 2022, "venue": "IEEE ACII", "cluster": "Mental Health"},
        {"title": "Multimodal Gender Fairness in Depression Prediction", "url": "https://ieeexplore.ieee.org/document/10970366/", "year": 2025, "venue": "IEEE Conference", "cluster": "Mental Health"},
        {"title": "Identifying Gender Bias in Generative Models for Mental Health Synthetic Data", "url": "https://ieeexplore.ieee.org/document/10337173/", "year": 2023, "venue": "IEEE Conference", "cluster": "Mental Health"},
        # Radiology / Medical Imaging
        {"title": "Bias Analysis on Public X-Ray Image Datasets of Pneumonia and COVID-19 Patients", "url": "https://ieeexplore.ieee.org/document/9374968/", "year": 2021, "venue": "IEEE Journal", "cluster": "Radiology"},
        {"title": "A Comparative Study of Fairness in Medical Machine Learning", "url": "https://ieeexplore.ieee.org/document/10230368/", "year": 2023, "venue": "IEEE Conference", "cluster": "Medical Imaging"},
        # Federated Learning
        {"title": "A Survey on Bias Mitigation in Federated Learning", "url": "https://ieeexplore.ieee.org/document/10372031/", "year": 2023, "venue": "IEEE Conference", "cluster": "Federated Learning"},
        {"title": "Towards Fairness-Aware Federated Learning", "url": "https://ieeexplore.ieee.org/document/10097767/", "year": 2023, "venue": "IEEE Journal", "cluster": "Federated Learning"},
        # Fairness-Aware Models
        {"title": "Fairness-Aware Graph Neural Networks for ICU Length of Stay Prediction in IoT-Enabled Environments", "url": "https://ieeexplore.ieee.org/document/10963712/", "year": 2025, "venue": "IEEE Journal", "cluster": "ICU"},
        # Gradient-Based
        {"title": "Gradient-Based Reconciliation of Fairness and Performance in Healthcare AI", "url": "https://ieeexplore.ieee.org/document/11081643/", "year": 2025, "venue": "IEEE ICHI", "cluster": "Bias Mitigation"},
        {"title": "Enhancing Multi-Attribute Fairness in Healthcare Predictive Modeling", "url": "https://ieeexplore.ieee.org/document/11081631/", "year": 2025, "venue": "IEEE ICHI", "cluster": "Bias Mitigation"},
        # Fairness Metrics
        {"title": "Fairness Metrics in AI Healthcare Applications: A Review", "url": "https://ieeexplore.ieee.org/document/10703970/", "year": 2024, "venue": "IEEE IRI", "cluster": "Fairness Metrics"},
    ]

    # ACM expanded
    acm_expanded = [
        {"title": "Dissecting Racial Bias in an Algorithm that Guides Health Decisions for 70 Million People", "url": "https://dl.acm.org/doi/10.1145/3287560.3287593", "year": 2019, "venue": "ACM FAccT", "cluster": "Landmark Study"},
        {"title": "A Survey on Bias and Fairness in Machine Learning", "url": "https://dl.acm.org/doi/10.1145/3457607", "year": 2021, "venue": "ACM Computing Surveys", "cluster": "Survey"},
        {"title": "Bias in Reinforcement Learning: A Review in Healthcare Applications", "url": "https://dl.acm.org/doi/10.1145/3609502", "year": 2023, "venue": "ACM Computing Surveys", "cluster": "Survey"},
        {"title": "Target Specification Bias, Counterfactual Prediction, and Algorithmic Fairness in Healthcare", "url": "https://dl.acm.org/doi/10.1145/3600211.3604678", "year": 2023, "venue": "AAAI/ACM AIES", "cluster": "Bias Assessment"},
        {"title": "Fairness in Machine Learning for Healthcare (Tutorial)", "url": "https://dl.acm.org/doi/10.1145/3394486.3406461", "year": 2020, "venue": "ACM SIGKDD", "cluster": "Tutorial"},
        {"title": "Net Benefit, Calibration, Threshold Selection, and Training Objectives for Algorithmic Fairness in Healthcare", "url": "https://dl.acm.org/doi/10.1145/3531146.3533166", "year": 2022, "venue": "ACM FAccT", "cluster": "Fairness Metrics"},
        {"title": "Multi-disciplinary Fairness Considerations in Machine Learning for Clinical Trials", "url": "https://dl.acm.org/doi/10.1145/3531146.3533154", "year": 2022, "venue": "ACM FAccT", "cluster": "Clinical Trials"},
        {"title": "In the Name of Fairness: Assessing the Bias in Clinical Record De-identification", "url": "https://dl.acm.org/doi/10.1145/3593013.3593982", "year": 2023, "venue": "ACM FAccT", "cluster": "NLP/Clinical Text"},
        {"title": "Improving Fairness in AI Models on EHR: The Case for Federated Learning Methods", "url": "https://dl.acm.org/doi/10.1145/3593013.3594102", "year": 2023, "venue": "ACM FAccT", "cluster": "Federated Learning"},
        {"title": "Advancing Health Equity with Machine Learning", "url": "https://dl.acm.org/doi/10.1145/3600211.3604753", "year": 2023, "venue": "AAAI/ACM AIES", "cluster": "Health Equity"},
        {"title": "Evaluating the Impact of Social Determinants on Health Prediction in the ICU", "url": "https://dl.acm.org/doi/10.1145/3600211.3604719", "year": 2023, "venue": "AAAI/ACM AIES", "cluster": "ICU/SDoH"},
        {"title": "Inherent Bias in Electronic Health Records: A Scoping Review of Sources of Bias", "url": "https://dl.acm.org/doi/10.1145/3757924", "year": 2025, "venue": "ACM TIST", "cluster": "EHR"},
        {"title": "Sample Selection Bias in Machine Learning for Healthcare", "url": "https://dl.acm.org/doi/10.1145/3761822", "year": 2025, "venue": "ACM Trans. Computing Healthcare", "cluster": "EHR"},
        {"title": "Evaluating Accuracy and Fairness of Clinical Decision Support Algorithms", "url": "https://dl.acm.org/doi/10.1016/j.jbi.2024.104664", "year": 2024, "venue": "J. Biomedical Informatics", "cluster": "Clinical Decision Support"},
        {"title": "Assessing Racial Bias in Healthcare Predictive Models: 30-Day Hospital Readmission", "url": "https://dl.acm.org/doi/10.1016/j.jbi.2024.104683", "year": 2024, "venue": "J. Biomedical Informatics", "cluster": "Readmission Prediction"},
        {"title": "AI Fairness in Medical Imaging: Controlling for Disease Severity", "url": "https://dl.acm.org/doi/10.1007/978-3-031-72787-0_3", "year": 2024, "venue": "FAIMI/MICCAI", "cluster": "Medical Imaging"},
        {"title": "On Fairness of Medical Image Classification with Multiple Sensitive Attributes", "url": "https://dl.acm.org/doi/10.1007/978-3-031-34048-2_13", "year": 2023, "venue": "IPMI", "cluster": "Medical Imaging"},
        {"title": "An Investigation into Race Bias in Random Forest Models Based on Breast DCE-MRI Derived Radiomics", "url": "https://dl.acm.org/doi/10.1007/978-3-031-45249-9_22", "year": 2023, "venue": "FAIMI/MICCAI", "cluster": "Radiology/Breast"},
        {"title": "MEDebiaser: A Human-AI Feedback System for Mitigating Bias in Medical Image Classification", "url": "https://dl.acm.org/doi/10.1145/3746059.3747725", "year": 2025, "venue": "ACM UIST", "cluster": "Medical Imaging"},
        {"title": "Mitigating Racial Biases for ML-Based Skin Cancer Detection", "url": "https://dl.acm.org/doi/10.1145/3565287.3617639", "year": 2023, "venue": "ACM MobiHoc", "cluster": "Dermatology"},
        {"title": "The Case for Globalizing Fairness: A Mixed Methods Study on Colonialism, AI, and Health in Africa", "url": "https://dl.acm.org/doi/10.1145/3689904.3694708", "year": 2024, "venue": "ACM EAAMO", "cluster": "Global Health Equity"},
        {"title": "Understanding Fairness in Recommender Systems: A Healthcare Perspective", "url": "https://dl.acm.org/doi/10.1145/3640457.3691711", "year": 2024, "venue": "ACM RecSys", "cluster": "Recommender Systems"},
        {"title": "Fairness and Inclusion Methods for Biomedical Informatics Research", "url": "https://dl.acm.org/doi/10.1016/j.jbi.2024.104713", "year": 2024, "venue": "J. Biomedical Informatics", "cluster": "Biomedical Informatics"},
        {"title": "Enhancing Data Diversity and Traceability to Mitigate Bias in Healthcare AI: A Blockchain-Driven Approach", "url": "https://dl.acm.org/doi/10.1145/3632634.3655881", "year": 2024, "venue": "ACM SIGMIS-CPR", "cluster": "Data Diversity"},
        {"title": "Explanatory Debiasing: Involving Domain Experts in Data Generation to Mitigate Representation Bias", "url": "https://dl.acm.org/doi/10.1145/3706598.3713497", "year": 2025, "venue": "ACM CHI", "cluster": "Debiasing Methods"},
        {"title": "Who Is Included in Human Perceptions of AI? Trust and Perceived Fairness in Healthcare AI", "url": "https://dl.acm.org/doi/10.1145/3411764.3445570", "year": 2021, "venue": "ACM CHI", "cluster": "Trust/Perception"},
        {"title": "Interpretable Bias Mitigation for Textual Data: Reducing Genderization in Patient Notes", "url": "https://dl.acm.org/doi/10.1145/3524887", "year": 2022, "venue": "ACM Trans. Computing Healthcare", "cluster": "NLP/Clinical Text"},
        {"title": "Fairness Challenges in the Design of ML Applications for Healthcare", "url": "https://dl.acm.org/doi/10.1145/3728368", "year": 2025, "venue": "ACM Trans. Computing Healthcare", "cluster": "Framework"},
        {"title": "Fair Foundation Models for Medical Image Analysis", "url": "https://dl.acm.org/doi/10.1145/3793542", "year": 2025, "venue": "ACM Trans. Computing Healthcare", "cluster": "Medical Imaging"},
        {"title": "Stress-Testing Bias Mitigation Algorithms to Understand Fairness Vulnerabilities", "url": "https://dl.acm.org/doi/10.1145/3600211.3604713", "year": 2023, "venue": "AAAI/ACM AIES", "cluster": "Bias Mitigation Testing"},
        {"title": "Leveraging Simulation Data to Understand Bias in Predictive Models of Infectious Disease Spread", "url": "https://dl.acm.org/doi/10.1145/3660631", "year": 2024, "venue": "ACM Trans. Spatial Algorithms", "cluster": "Infectious Disease"},
        {"title": "Reward Systems for Trustworthy Medical Federated Learning", "url": "https://dl.acm.org/doi/10.1145/3761821", "year": 2025, "venue": "ACM Trans. Computing Healthcare", "cluster": "Federated Learning"},
        {"title": "Chasing Your Long Tails: Differentially Private Prediction in Health Care Settings", "url": "https://dl.acm.org/doi/10.1145/3442188.3445934", "year": 2021, "venue": "ACM FAccT", "cluster": "Privacy/Fairness"},
        {"title": "Counterfactual Risk Assessments, Evaluation, and Fairness", "url": "https://dl.acm.org/doi/10.1145/3351095.3372851", "year": 2020, "venue": "ACM FAccT", "cluster": "Counterfactual Fairness"},
        {"title": "Evaluating and Reducing Subgroup Disparity in AI Models Predicting Pediatric COVID-19 Test Outcomes", "url": "https://ieeexplore.ieee.org/document/10825572/", "year": 2024, "venue": "IEEE Conference", "cluster": "Pediatrics/COVID"},
    ]

    # IEEE Expanded sheet
    ws_ieee = wb.create_sheet(title='IEEE_Expanded')
    style_header(ws_ieee, TITLE_HEADERS)
    ws_ieee.freeze_panes = 'A2'
    for idx, paper in enumerate(ieee_expanded, 1):
        row_data = [idx, paper['title'], paper['year'], paper['venue'], paper['url'], paper['cluster'], 'Yes']
        for col_idx, value in enumerate(row_data, 1):
            style_cell(ws_ieee, idx + 1, col_idx, str(value))
    auto_width(ws_ieee, TITLE_HEADERS)

    # ACM Expanded sheet
    ws_acm = wb.create_sheet(title='ACM_Expanded')
    style_header(ws_acm, TITLE_HEADERS)
    ws_acm.freeze_panes = 'A2'
    for idx, paper in enumerate(acm_expanded, 1):
        row_data = [idx, paper['title'], paper['year'], paper['venue'], paper['url'], paper['cluster'], 'Yes']
        for col_idx, value in enumerate(row_data, 1):
            style_cell(ws_acm, idx + 1, col_idx, str(value))
    auto_width(ws_acm, TITLE_HEADERS)

    return len(ieee_expanded), len(acm_expanded)


# ============================================================
# MAIN
# ============================================================
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ---- Summary Sheet ----
    ws_summary = wb.create_sheet(title='Summary', index=0)
    summary_headers = ['Database', 'Papers Found', 'Title-Screened', 'Deep-Screened (Full Data)', 'Notes']
    style_header(ws_summary, summary_headers)
    ws_summary.freeze_panes = 'A2'

    # ---- 1. PubMed/MEDLINE ----
    print("Adding PubMed/MEDLINE sheet...")
    ws_pubmed, pubmed_count = add_pubmed_sheet(wb)

    # ---- 2. IEEE Xplore (deeply screened) ----
    print("Adding IEEE Xplore (deep screened) sheet...")
    from ieee_screening_results import IEEE_RESULTS
    add_deep_screened_sheet(wb, 'IEEE_DeepScreened', IEEE_RESULTS)
    ieee_deep_count = len(IEEE_RESULTS)

    # ---- 3. arXiv (deeply screened) ----
    print("Adding arXiv (deep screened) sheet...")
    from arxiv_screening_results import ARXIV_RESULTS
    add_deep_screened_sheet(wb, 'arXiv_DeepScreened', ARXIV_RESULTS)
    arxiv_deep_count = len(ARXIV_RESULTS)

    # ---- 4. IEEE + ACM expanded ----
    print("Adding IEEE+ACM expanded sheets...")
    ieee_exp_count, acm_exp_count = add_expanded_ieee_acm_sheet(wb)

    # ---- 5. Google Scholar ----
    # Add Google Scholar papers from original search
    scholar_papers = [
        {"title": "Bias recognition and mitigation strategies in AI healthcare applications", "url": "https://www.nature.com/articles/s41746-025-01503-7", "year": 2025, "venue": "npj Digital Medicine", "cluster": "Review"},
        {"title": "Fairness in AI for healthcare", "url": "https://www.sciencedirect.com/science/article/pii/S2514664524015674", "year": 2024, "venue": "ScienceDirect", "cluster": "Review"},
        {"title": "AI-driven healthcare: A review on ensuring fairness and mitigating bias", "url": "https://journals.plos.org/digitalhealth/article?id=10.1371/journal.pdig.0000864", "year": 2024, "venue": "PLOS Digital Health", "cluster": "Review"},
        {"title": "A survey of recent methods for addressing AI fairness and bias in biomedicine", "url": "https://www.sciencedirect.com/science/article/pii/S1532046424000649", "year": 2024, "venue": "J Biomedical Informatics", "cluster": "Survey"},
        {"title": "Mitigating bias in machine learning for medicine", "url": "https://www.nature.com/articles/s43856-021-00028-w", "year": 2021, "venue": "Communications Medicine", "cluster": "Review"},
        {"title": "Identifying and mitigating algorithmic bias in the safety net", "url": "https://www.nature.com/articles/s41746-025-01732-w", "year": 2025, "venue": "npj Digital Medicine", "cluster": "Empirical"},
        {"title": "Evaluation and Mitigation of Racial Bias in Clinical ML Models: Scoping Review", "url": "https://pmc.ncbi.nlm.nih.gov/articles/PMC9198828/", "year": 2022, "venue": "JMIR Medical Informatics", "cluster": "Scoping Review"},
        {"title": "Algorithmic fairness and bias mitigation for clinical ML with deep reinforcement learning", "url": "https://pmc.ncbi.nlm.nih.gov/articles/PMC10442224/", "year": 2023, "venue": "Nature Scientific Reports", "cluster": "Empirical"},
    ]
    ws_scholar = wb.create_sheet(title='Google_Scholar')
    style_header(ws_scholar, TITLE_HEADERS)
    ws_scholar.freeze_panes = 'A2'
    for idx, paper in enumerate(scholar_papers, 1):
        row_data = [idx, paper['title'], paper['year'], paper['venue'], paper['url'], paper['cluster'], 'Yes']
        for col_idx, value in enumerate(row_data, 1):
            style_cell(ws_scholar, idx + 1, col_idx, str(value))
    auto_width(ws_scholar, TITLE_HEADERS)

    # ---- Fill Summary ----
    summary_data = [
        ['PubMed/MEDLINE', '1,523 total', str(pubmed_count), '—', 'Title-filtered from E-utilities API; needs abstract screening'],
        ['IEEE Xplore', str(ieee_exp_count + ieee_deep_count), str(ieee_exp_count), str(ieee_deep_count), 'Deep screening: full data extraction for 17 papers'],
        ['ACM Digital Library', str(acm_exp_count), str(acm_exp_count), '—', 'Title-screened; needs abstract screening'],
        ['arXiv', str(arxiv_deep_count), str(arxiv_deep_count), str(arxiv_deep_count), 'Deep screening: full data extraction for all 26 papers'],
        ['Google Scholar', str(len(scholar_papers)), str(len(scholar_papers)), '—', 'Supplementary; many overlap with PubMed'],
        ['Scopus/Web of Science', 'Pending', '—', '—', 'Search was interrupted; results pending'],
        ['TOTAL (before dedup)', str(pubmed_count + ieee_exp_count + ieee_deep_count + acm_exp_count + arxiv_deep_count + len(scholar_papers)), '', '', '']
    ]
    for row_idx, row_data in enumerate(summary_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = style_cell(ws_summary, row_idx, col_idx, value)
            if row_idx == len(summary_data) + 1:
                cell.font = Font(bold=True)
    auto_width(ws_summary, summary_headers)

    # Save
    output_path = '/home/user/Bias-Review-Paper/Systematic_Review_Database_Screening.xlsx'
    wb.save(output_path)
    print(f"\nExcel file saved to: {output_path}")
    print(f"\nSheets created:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

    total = pubmed_count + ieee_exp_count + ieee_deep_count + acm_exp_count + arxiv_deep_count + len(scholar_papers)
    print(f"\nTotal papers across all databases (before deduplication): {total}")

if __name__ == '__main__':
    main()
