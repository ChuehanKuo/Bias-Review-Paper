#!/usr/bin/env python3
"""
Embase Screening Pipeline for Systematic Review:
"Approaches for Assessing and Mitigating Algorithmic Bias in Health AI"

Uses OpenAlex API to search Embase papers.
Same screening criteria and column extraction as PubMed pipeline.

1. Search OpenAlex with Embase filter
2. Deduplicate
3. Screen by title+abstract
4. Extract structured columns
5. Full-text screen (abstract-based, Embase papers mostly paywalled)
6. Build Excel
"""

import json
import urllib.request
import urllib.parse
import time
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# EMBASE FILTER ID in OpenAlex
# ============================================================
FILTER_ID = 'p4310311648'

# ============================================================
# SEARCH QUERIES — adapted for OpenAlex text search
# ============================================================
# OpenAlex search uses relevance-based text matching.
# We use multiple focused queries to maximize recall.
QUERIES = [
    {
        "id": "Q1",
        "label": "Core: algorithmic bias/fairness + health",
        "search": "algorithmic bias fairness health healthcare clinical medical"
    },
    {
        "id": "Q2",
        "label": "AI bias mitigation health",
        "search": "AI bias mitigation debiasing health healthcare clinical"
    },
    {
        "id": "Q3",
        "label": "Machine learning fairness health",
        "search": "machine learning fairness bias health clinical medical"
    },
    {
        "id": "Q4",
        "label": "Bias assessment AI healthcare",
        "search": "bias assessment evaluation audit artificial intelligence healthcare"
    },
    {
        "id": "Q5",
        "label": "Racial gender bias clinical AI",
        "search": "racial gender bias clinical prediction algorithm machine learning"
    },
    {
        "id": "Q6",
        "label": "Health disparities AI algorithm",
        "search": "health disparities algorithmic bias machine learning deep learning"
    },
    {
        "id": "Q7",
        "label": "Fairness-aware ML health",
        "search": "fairness-aware machine learning equalized odds demographic parity health"
    },
    {
        "id": "Q8",
        "label": "Bias EHR clinical decision support",
        "search": "bias electronic health record clinical decision support algorithm"
    },
    {
        "id": "Q9",
        "label": "Bias medical imaging AI",
        "search": "bias medical imaging radiology dermatology deep learning AI"
    },
    {
        "id": "Q10",
        "label": "Equitable AI health",
        "search": "equitable AI health equity fairness algorithm clinical"
    },
    {
        "id": "Q11",
        "label": "NLP bias clinical health",
        "search": "natural language processing bias clinical health medical NLP"
    },
    {
        "id": "Q12",
        "label": "Bias federated learning health",
        "search": "bias fairness federated learning health clinical"
    },
    {
        "id": "Q13",
        "label": "LLM bias health",
        "search": "large language model bias health clinical medical"
    },
    {
        "id": "Q14",
        "label": "Disparate impact health prediction",
        "search": "disparate impact health prediction model algorithm bias"
    },
    {
        "id": "Q15",
        "label": "Bias risk prediction clinical",
        "search": "bias risk prediction mortality readmission clinical algorithm"
    },
]


# ============================================================
# OpenAlex API FUNCTIONS
# ============================================================

def reconstruct_abstract(inverted_index):
    """Reconstruct abstract text from OpenAlex inverted index format."""
    if not inverted_index:
        return ''
    words = []
    for word, positions in inverted_index.items():
        for pos in positions:
            words.append((pos, word))
    words.sort()
    return ' '.join(w for _, w in words)


def search_openalex_embase(query, per_page=200, max_results=2000):
    """Search OpenAlex for Embase papers matching query."""
    results = {}
    cursor = '*'
    fetched = 0

    while cursor and fetched < max_results:
        params = urllib.parse.urlencode({
            'search': query,
            'filter': f'primary_location.source.publisher_lineage:{FILTER_ID}',
            'per_page': per_page,
            'cursor': cursor,
            'mailto': 'review@example.com',
            'select': 'id,doi,title,publication_year,primary_location,authorships,'
                      'abstract_inverted_index,keywords,type,cited_by_count'
        })
        url = f'https://api.openalex.org/works?{params}'

        try:
            with urllib.request.urlopen(urllib.request.Request(url), timeout=60) as resp:
                data = json.loads(resp.read().decode('utf-8'))

            total = data.get('meta', {}).get('count', 0)
            for work in data.get('results', []):
                doi = work.get('doi', '') or ''
                openalex_id = work.get('id', '')
                # Use DOI as primary key, fall back to OpenAlex ID
                key = doi if doi else openalex_id
                if not key or key in results:
                    continue

                # Reconstruct abstract
                abstract = reconstruct_abstract(work.get('abstract_inverted_index'))

                # Authors
                authors = []
                for auth in work.get('authorships', []):
                    name = auth.get('author', {}).get('display_name', '')
                    if name:
                        authors.append(name)
                author_str = authors[0] + ' et al.' if len(authors) > 1 else (authors[0] if authors else 'N/A')

                # Source/venue
                source = work.get('primary_location', {}).get('source', {}) or {}
                venue = source.get('display_name', '')

                # Keywords
                kw_list = work.get('keywords', []) or []
                keywords = '; '.join([k.get('display_name', '') for k in kw_list if k.get('display_name')])

                results[key] = {
                    'openalex_id': openalex_id,
                    'doi': doi.replace('https://doi.org/', '') if doi else '',
                    'title': work.get('title', '') or '',
                    'abstract': abstract,
                    'year': str(work.get('publication_year', '')) or '',
                    'journal': venue,
                    'authors': author_str,
                    'keywords': keywords,
                    'mesh_terms': '',
                    'pub_types': work.get('type', ''),
                    'url': doi if doi else openalex_id,
                    'cited_by_count': work.get('cited_by_count', 0),
                }
                fetched += 1

            # Get next cursor
            cursor = data.get('meta', {}).get('next_cursor')
            if not data.get('results'):
                break

        except Exception as e:
            print(f"    ERROR: {e}")
            break

        time.sleep(0.2)  # Be polite to the API

    return results, fetched


# ============================================================
# SCREENING (same as PubMed pipeline)
# ============================================================

def screen_paper(paper):
    title = (paper.get('title', '') or '').lower()
    abstract = (paper.get('abstract', '') or '').lower()
    combined = title + ' ' + abstract
    kw = (paper.get('keywords', '') or '').lower()
    all_text = combined + ' ' + kw

    ai_terms = [
        'machine learning', 'deep learning', 'artificial intelligence', 'neural network',
        'algorithm', 'predictive model', 'prediction model', 'classifier', 'classification',
        'natural language processing', 'nlp', 'computer vision', 'random forest',
        'logistic regression', 'xgboost', 'convolutional', 'transformer',
        'large language model', 'llm', 'decision support', 'risk prediction',
        'federated learning', 'reinforcement learning', 'foundation model',
        'chatgpt', 'gpt-4', 'supervised learning'
    ]
    has_ai = any(t in all_text for t in ai_terms)

    health_terms = [
        'health', 'clinical', 'medical', 'patient', 'hospital', 'disease',
        'diagnosis', 'treatment', 'care', 'pathology', 'radiology', 'dermatology',
        'cardiology', 'oncology', 'ophthalmology', 'psychiatry', 'mental health',
        'ehr', 'electronic health', 'biomedical', 'mortality', 'readmission',
        'sepsis', 'icu', 'emergency', 'chest x-ray', 'mammograph', 'cancer',
        'diabetes', 'cardiovascular', 'public health', 'healthcare', 'medicine'
    ]
    has_health = any(t in all_text for t in health_terms)

    strong_in_title = any(t in title for t in [
        'bias', 'fairness', 'fair ', 'unfair', 'equitable', 'equity',
        'disparity', 'disparities', 'discrimination', 'debiasing', 'debias',
        'underdiagnos', 'underrepresent', 'inequit'
    ])
    ai_bias_terms = [
        'algorithmic bias', 'algorithmic fairness', 'ai bias', 'ai fairness',
        'machine learning bias', 'machine learning fairness', 'model bias',
        'bias mitigation', 'bias detection', 'bias assessment', 'fairness-aware',
        'fair machine learning', 'debiasing', 'disparate impact', 'demographic parity',
        'equalized odds', 'equal opportunity', 'fairness metric', 'fairness evaluation',
        'bias in ai', 'bias in machine learning', 'bias in algorithm', 'biased model',
        'biased prediction', 'mitigating bias', 'addressing bias', 'reducing bias',
        'assessing bias', 'evaluating bias', 'detecting bias', 'sources of bias',
        'racial bias', 'gender bias', 'age bias', 'socioeconomic bias', 'ethnic bias',
        'underdiagnosis bias', 'representation bias', 'data bias', 'label bias',
        'health disparit', 'health equit', 'health inequit', 'unfair'
    ]
    bias_count = sum(1 for t in ai_bias_terms if t in abstract)
    has_substantial_bias = strong_in_title or bias_count >= 2

    if not has_ai:
        return False, 'No AI/ML component'
    if not has_health:
        return False, 'Not health-related'
    if not has_substantial_bias:
        return False, 'Bias/fairness not central topic'

    human_only = ['implicit bias training', 'implicit bias among', 'clinician bias',
                  'physician bias', 'provider bias', 'unconscious bias training',
                  'implicit association test', 'weight bias', 'anti-fat bias']
    ai_specific = ['algorithmic bias', 'ai bias', 'model bias', 'bias mitigation',
                   'bias in ai', 'bias in machine learning', 'algorithmic fairness',
                   'fairness-aware', 'debiasing algorithm']
    if any(t in all_text for t in human_only) and not any(t in all_text for t in ai_specific):
        return False, 'Human cognitive biases only'

    return True, 'Included'


# ============================================================
# FULL-TEXT SCREENING (abstract-based for Embase)
# ============================================================

def fulltext_screen(paper):
    """
    Full-text screening with strict criteria.
    For Embase papers, we use abstract only since most are paywalled.
    """
    title = (paper.get('title', '') or '').lower()
    abstract = (paper.get('abstract', '') or '').lower()
    combined = title + ' ' + abstract

    approach_indicators = [
        'bias assessment', 'bias detection', 'bias evaluation', 'bias audit',
        'fairness assessment', 'fairness evaluation', 'fairness audit',
        'bias measurement', 'bias quantification', 'bias analysis',
        'measuring bias', 'detecting bias', 'evaluating bias',
        'fairness metric', 'fairness measure', 'bias metric',
        'demographic parity', 'equalized odds', 'equal opportunity',
        'disparate impact', 'predictive parity', 'calibration across',
        'subgroup analysis', 'disaggregated', 'stratified performance',
        'bias mitigation', 'bias reduction', 'bias correction',
        'debiasing', 'debias', 'fairness-aware',
        'fair machine learning', 'fair classification',
        'adversarial debiasing', 'reweighting', 'reweighing',
        'resampling', 'data augmentation for fairness',
        'fairness constraint', 'fairness regularization',
        'threshold adjustment', 'calibration', 'post-processing',
        'pre-processing', 'in-processing',
        'counterfactual fairness', 'causal fairness',
        'fairness framework', 'bias framework', 'equity framework',
        'ai fairness 360', 'aequitas', 'fairlearn',
        'bias toolkit', 'fairness toolkit',
        'model card', 'datasheet',
        'review of bias', 'survey of fairness', 'review of fairness',
        'bias mitigation strategies', 'approaches to fairness',
        'methods for bias', 'techniques for fairness',
    ]
    approach_count = sum(1 for t in approach_indicators if t in combined)

    ai_terms = [
        'machine learning', 'deep learning', 'artificial intelligence',
        'neural network', 'algorithm', 'predictive model', 'classifier',
        'natural language processing', 'computer vision',
        'random forest', 'logistic regression', 'xgboost',
        'convolutional', 'transformer', 'large language model', 'llm',
        'decision support', 'risk prediction', 'federated learning',
        'reinforcement learning', 'foundation model', 'chatgpt', 'gpt'
    ]
    has_ai = any(t in combined for t in ai_terms)

    health_terms = [
        'health', 'clinical', 'medical', 'patient', 'hospital', 'disease',
        'diagnosis', 'treatment', 'care', 'radiology', 'dermatology',
        'cardiology', 'oncology', 'ophthalmology', 'psychiatry',
        'ehr', 'electronic health', 'biomedical', 'mortality',
        'readmission', 'sepsis', 'icu', 'emergency', 'chest x-ray',
        'cancer', 'diabetes', 'cardiovascular', 'public health',
        'healthcare', 'medicine'
    ]
    has_health = any(t in combined for t in health_terms)

    bias_in_title = any(t in title for t in [
        'bias', 'fairness', 'fair ', 'equitable', 'equity', 'disparity',
        'disparities', 'debiasing', 'debias', 'underdiagnos', 'inequit',
        'discrimination'
    ])

    if not has_ai:
        return False, 'No AI/ML component'
    if not has_health:
        return False, 'Not health-related'

    # Abstract-only screening (slightly more lenient than full-text)
    if approach_count >= 2 and bias_in_title:
        return True, 'Included: bias central + approach content'
    elif approach_count >= 3:
        return True, 'Included: substantial approach content in abstract'
    elif bias_in_title and approach_count >= 1:
        return True, 'Included: bias is central topic (abstract only)'
    else:
        return False, f'Excluded: insufficient approach content ({approach_count} indicators)'


# ============================================================
# COLUMN EXTRACTION (same as PubMed pipeline)
# ============================================================

def extract_study_type(paper):
    title = (paper.get('title', '') or '').lower()
    abstract = (paper.get('abstract', '') or '').lower()
    pub_types = (paper.get('pub_types', '') or '').lower()
    combined = title + ' ' + abstract + ' ' + pub_types

    if any(t in combined for t in ['systematic review', 'prisma', 'systematic literature review']):
        return 'Systematic Review'
    if any(t in combined for t in ['scoping review', 'scoping study']):
        return 'Scoping Review'
    if any(t in combined for t in ['narrative review', 'literature review', 'review article']):
        return 'Narrative Review'
    if 'meta-analysis' in combined:
        return 'Meta-Analysis'
    if any(t in combined for t in ['survey', 'questionnaire', 'interview']):
        return 'Survey/Qualitative'
    if any(t in combined for t in ['commentary', 'editorial', 'perspective', 'viewpoint', 'opinion', 'letter']):
        return 'Commentary/Editorial'
    if any(t in combined for t in ['guideline', 'recommendation', 'consensus', 'policy']):
        return 'Guideline/Policy'
    if any(t in combined for t in ['framework', 'toolkit', 'toolbox', 'pipeline']):
        return 'Framework/Toolkit'
    if any(t in combined for t in ['we propose', 'we develop', 'we introduce', 'novel method', 'novel approach']):
        return 'Methodology'
    if any(t in combined for t in ['experiment', 'dataset', 'we evaluated', 'we trained', 'we tested', 'we applied', 'we analyzed', 'we assessed', 'results show', 'our results', 'we found']):
        return 'Empirical Study'
    if 'review' in combined:
        return 'Review'
    return 'Empirical Study'


def extract_ai_ml_method(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    combined = title + ' ' + abstract
    methods = []
    method_map = {
        'Deep Learning': ['deep learning', 'deep neural', 'cnn', 'convolutional neural', 'resnet', 'densenet', 'vgg', 'inception'],
        'NLP/LLM': ['natural language processing', 'nlp', 'large language model', 'llm', 'chatgpt', 'gpt-4', 'gpt-3', 'bert', 'transformer', 'text mining', 'text classification'],
        'Random Forest': ['random forest'],
        'Logistic Regression': ['logistic regression'],
        'XGBoost/Gradient Boosting': ['xgboost', 'gradient boosting', 'lightgbm', 'catboost'],
        'Support Vector Machine': ['support vector', 'svm'],
        'Neural Network': ['neural network', 'mlp', 'perceptron'],
        'Decision Tree': ['decision tree', 'cart'],
        'Ensemble Methods': ['ensemble', 'bagging', 'stacking'],
        'Federated Learning': ['federated learning'],
        'Reinforcement Learning': ['reinforcement learning'],
        'Computer Vision/Imaging AI': ['computer vision', 'image classification', 'image segmentation', 'object detection', 'medical imaging'],
        'Clinical Prediction Model': ['clinical prediction', 'risk prediction', 'risk score', 'prediction model', 'predictive model', 'prognostic model'],
        'Clinical Decision Support': ['clinical decision support', 'decision support system', 'cdss'],
        'Foundation Model': ['foundation model'],
        'Generative AI': ['generative ai', 'generative model', 'gan', 'generative adversarial'],
        'Clustering': ['clustering', 'k-means', 'unsupervised'],
        'Regression': ['linear regression', 'regression model'],
        'Survival Analysis': ['survival analysis', 'cox regression', 'hazard model'],
    }
    for label, terms in method_map.items():
        if any(t in combined for t in terms):
            methods.append(label)
    return '; '.join(methods) if methods else 'Not specified'


def extract_health_domain(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    kw = (paper.get('keywords', '') or '').lower()
    combined = title + ' ' + abstract + ' ' + kw
    domains = []
    domain_map = {
        'Radiology/Medical Imaging': ['radiology', 'radiograph', 'x-ray', 'ct scan', 'mri', 'mammograph', 'ultrasound', 'medical imaging', 'chest radiograph'],
        'Dermatology': ['dermatology', 'skin lesion', 'skin cancer', 'melanoma', 'skin disease', 'dermatoscop'],
        'Cardiology': ['cardiology', 'cardiovascular', 'heart', 'cardiac', 'ecg', 'ekg', 'echocardiogram', 'atrial fibrillation', 'coronary'],
        'Oncology': ['oncology', 'cancer', 'tumor', 'tumour', 'neoplasm', 'malignant', 'chemotherapy', 'radiation therapy'],
        'Ophthalmology': ['ophthalmology', 'retinal', 'retinopathy', 'diabetic retinopathy', 'glaucoma', 'fundus', 'optic'],
        'Mental Health/Psychiatry': ['mental health', 'psychiatry', 'depression', 'anxiety', 'suicide', 'psychiatric', 'behavioral health', 'substance abuse', 'opioid'],
        'Emergency Medicine': ['emergency department', 'emergency medicine', 'trauma', 'triage', 'acute care'],
        'ICU/Critical Care': ['icu', 'intensive care', 'critical care', 'sepsis', 'mechanical ventilation'],
        'EHR/Health Informatics': ['electronic health record', 'ehr', 'emr', 'health information', 'clinical informatics', 'health informatics'],
        'Primary Care': ['primary care', 'family medicine', 'general practice', 'ambulatory'],
        'Surgery': ['surgery', 'surgical', 'operative', 'postoperative'],
        'Pathology': ['pathology', 'histopathology', 'cytology', 'biopsy'],
        'Pediatrics': ['pediatric', 'paediatric', 'child', 'neonatal', 'newborn'],
        'Obstetrics/Maternal Health': ['obstetric', 'maternal', 'pregnancy', 'prenatal', 'perinatal', 'childbirth'],
        'Nephrology': ['kidney', 'renal', 'nephrology', 'dialysis'],
        'Pulmonology': ['pulmonary', 'respiratory', 'lung', 'pneumonia', 'covid-19', 'asthma', 'copd'],
        'Neurology': ['neurology', 'neurological', 'stroke', 'alzheimer', 'dementia', 'brain', 'epilepsy'],
        'Genomics/Genetics': ['genomic', 'genetic', 'genome', 'gene expression', 'variant', 'sequencing'],
        'Public Health': ['public health', 'population health', 'epidemiology', 'surveillance', 'community health'],
        'Drug Discovery/Pharmacology': ['drug discovery', 'pharmacology', 'pharmaceutical', 'drug development'],
        'Infectious Disease': ['infectious disease', 'infection', 'hiv', 'tuberculosis', 'malaria', 'hepatitis'],
        'Endocrinology/Diabetes': ['diabetes', 'diabetic', 'endocrin', 'insulin', 'glucose', 'hba1c'],
        'Wearables/Remote Monitoring': ['wearable', 'pulse oximetry', 'spo2', 'remote monitoring', 'mobile health', 'mhealth'],
    }
    for label, terms in domain_map.items():
        if any(t in combined for t in terms):
            domains.append(label)
    return '; '.join(domains) if domains else 'General Healthcare'


def extract_bias_axes(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    combined = title + ' ' + abstract
    axes = []
    axis_map = {
        'Race/Ethnicity': ['race', 'racial', 'ethnicity', 'ethnic', 'african american', 'black', 'white', 'hispanic', 'latino', 'asian', 'indigenous', 'minority', 'skin color', 'skin tone', 'skin pigment'],
        'Gender/Sex': ['gender', 'sex', 'male', 'female', 'women', 'men', 'transgender', 'sex-based', 'gender-based'],
        'Age': ['age', 'aging', 'elderly', 'older adult', 'pediatric', 'geriatric', 'age-related', 'age group'],
        'Socioeconomic Status': ['socioeconomic', 'income', 'poverty', 'insurance', 'uninsured', 'underinsured', 'medicaid', 'sociodemographic', 'education level', 'employment'],
        'Language': ['language', 'non-english', 'english proficiency', 'limited english', 'linguistic'],
        'Disability': ['disability', 'disabled', 'impairment', 'handicap'],
        'Geographic': ['geographic', 'rural', 'urban', 'region', 'low-resource', 'low-income country', 'global south', 'developing country'],
        'Insurance Status': ['insurance', 'uninsured', 'underinsured', 'medicaid', 'medicare', 'payer'],
        'Intersectional': ['intersectional', 'intersectionality', 'multiple sensitive', 'multi-attribute', 'compound'],
    }
    for label, terms in axis_map.items():
        if any(t in combined for t in terms):
            axes.append(label)
    return '; '.join(axes) if axes else 'Not specified'


def extract_lifecycle_stage(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    combined = title + ' ' + abstract
    stages = []
    stage_map = {
        'Data Collection': ['data collection', 'data acquisition', 'dataset creation', 'data gathering', 'sampling', 'cohort selection', 'data source'],
        'Data Preprocessing': ['preprocess', 'pre-process', 'data cleaning', 'imputation', 'feature engineering', 'data augmentation', 'resampling', 'oversampling', 'undersampling', 'reweighting', 'reweighing'],
        'Model Development/Training': ['model training', 'model development', 'training phase', 'in-processing', 'adversarial debiasing', 'fairness constraint', 'regularization', 'fair representation', 'calibration'],
        'Model Evaluation': ['model evaluation', 'evaluation', 'testing', 'validation', 'performance assessment', 'fairness metric', 'bias assessment', 'bias detection', 'audit', 'post-processing', 'subgroup analysis'],
        'Deployment': ['deployment', 'clinical implementation', 'real-world', 'production', 'clinical practice', 'clinical workflow', 'monitoring'],
    }
    for label, terms in stage_map.items():
        if any(t in combined for t in terms):
            stages.append(label)
    return '; '.join(stages) if stages else 'Not specified'


def extract_assessment_or_mitigation(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    combined = title + ' ' + abstract
    assess_terms = ['assess', 'detect', 'evaluat', 'audit', 'measur', 'quantif', 'identif', 'characteriz', 'examin', 'analyz', 'investigat']
    mitigate_terms = ['mitigat', 'debias', 'reduc', 'eliminat', 'correct', 'remov', 'address', 'alleviat', 'improv fairness', 'fairness-aware', 'fair machine', 'reweigh', 'adversarial debias']
    has_assess = any(t in combined for t in assess_terms)
    has_mitigate = any(t in combined for t in mitigate_terms)
    if has_assess and has_mitigate:
        return 'Both'
    elif has_mitigate:
        return 'Mitigation'
    elif has_assess:
        return 'Assessment'
    return 'Not specified'


def extract_approach_method(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    approaches = []
    approach_map = {
        'Reweighting/Resampling': ['reweighting', 'reweighing', 'resampling', 'oversampling', 'undersampling', 'smote'],
        'Adversarial Debiasing': ['adversarial debiasing', 'adversarial learning', 'adversarial training'],
        'Fairness Constraints': ['fairness constraint', 'fairness regulariz', 'fairness-aware training'],
        'Calibration': ['calibration', 'recalibration', 'platt scaling'],
        'Threshold Adjustment': ['threshold', 'cutoff', 'operating point'],
        'Data Augmentation': ['data augmentation', 'synthetic data', 'augmented data', 'fair mixup'],
        'Subgroup Analysis': ['subgroup analysis', 'stratified analysis', 'subpopulation', 'disaggregated'],
        'Fairness Metrics Evaluation': ['demographic parity', 'equalized odds', 'equal opportunity', 'disparate impact', 'fairness metric', 'predictive parity', 'calibration across'],
        'Counterfactual Fairness': ['counterfactual', 'causal inference', 'causal fairness'],
        'Representation Learning': ['representation learning', 'fair representation', 'disentangl', 'embedding'],
        'Transfer Learning': ['transfer learning', 'domain adaptation', 'fine-tun'],
        'Federated Learning': ['federated learning', 'federated model'],
        'Explainability/Interpretability': ['explainab', 'interpretab', 'shap', 'lime', 'attention'],
        'Post-hoc Correction': ['post-processing', 'post-hoc', 'reject option'],
        'Diverse/Representative Data': ['diverse dataset', 'representative data', 'inclusive data', 'data diversity', 'balanced dataset'],
        'Bias Auditing Framework': ['audit', 'bias audit', 'fairness audit', 'model card', 'datasheet'],
        'Multi-task Learning': ['multi-task', 'multitask'],
        'Ensemble Methods': ['ensemble', 'model ensemble'],
        'Regularization': ['regulariz', 'penalty', 'constraint-based'],
    }
    for label, terms in approach_map.items():
        if any(t in abstract for t in terms):
            approaches.append(label)
    return '; '.join(approaches) if approaches else 'Not specified'


def extract_clinical_setting(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    title = (paper.get('title', '') or '').lower()
    combined = title + ' ' + abstract
    settings = []
    setting_map = {
        'Hospital/Inpatient': ['hospital', 'inpatient', 'hospitalized', 'admitted'],
        'ICU': ['icu', 'intensive care', 'critical care'],
        'Emergency Department': ['emergency department', 'emergency room', 'ed visit'],
        'Primary Care/Outpatient': ['primary care', 'outpatient', 'ambulatory', 'clinic visit', 'office visit'],
        'Public Health/Population': ['public health', 'population', 'community', 'screening program', 'surveillance'],
        'Telehealth/Remote': ['telehealth', 'telemedicine', 'remote', 'mobile health', 'mhealth', 'wearable'],
        'Clinical Trial': ['clinical trial', 'randomized controlled', 'rct'],
        'Laboratory/Pathology': ['laboratory', 'pathology', 'lab test'],
        'Safety-Net/Underserved': ['safety net', 'safety-net', 'underserved', 'low-resource'],
        'Long-term Care': ['long-term care', 'nursing home', 'assisted living'],
    }
    for label, terms in setting_map.items():
        if any(t in combined for t in terms):
            settings.append(label)
    return '; '.join(settings) if settings else 'Not specified'


def extract_key_findings(paper):
    abstract = (paper.get('abstract', '') or '')
    if not abstract:
        return 'No abstract available'
    lower = abstract.lower()
    for marker in ['CONCLUSION:', 'CONCLUSIONS:', 'RESULTS:', 'FINDINGS:', 'DISCUSSION:']:
        idx = lower.find(marker.lower())
        if idx != -1:
            finding = abstract[idx:].strip()
            sentences = re.split(r'(?<=[.!?])\s+', finding)
            result = ' '.join(sentences[:3])
            if len(result) > 500:
                result = result[:500] + '...'
            return result
    sentences = re.split(r'(?<=[.!?])\s+', abstract)
    if len(sentences) >= 2:
        result = ' '.join(sentences[-2:])
    else:
        result = abstract
    if len(result) > 500:
        result = result[:500] + '...'
    return result


# ============================================================
# EXCEL BUILDING
# ============================================================

HEADERS = [
    ('No.', 5), ('DOI', 30), ('Title', 65), ('Authors', 20),
    ('Year', 6), ('Venue/Journal', 35), ('URL', 35),
    ('Abstract', 80),
    ('FT Screening Reason', 40),
    ('Study Type', 20), ('AI/ML Method', 30), ('Health Domain', 25),
    ('Bias Axes Assessed\n(Q1)', 30),
    ('AI Lifecycle Stage\n(Q2)', 25),
    ('Assessment vs Mitigation\n(Q2)', 22),
    ('Approach/Method', 35),
    ('Clinical Setting/Context\n(Q3)', 25),
    ('Key Findings', 50),
    ('Keywords', 35), ('Publication Type', 20), ('Cited By', 8), ('Notes', 20),
]

H_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
H_FILL = PatternFill(start_color='002147', end_color='002147', fill_type='solid')  # Embase (OUP)
H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
C_ALIGN = Alignment(vertical='top', wrap_text=True)
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
ALT_FILL = PatternFill(start_color='E8EDF5', end_color='E8EDF5', fill_type='solid')


def build_excel(included, excluded, query_stats, meta):
    wb = Workbook()
    wb.remove(wb.active)

    # --- Summary ---
    ws_s = wb.create_sheet('Summary', index=0)
    rows = [
        ['SYSTEMATIC REVIEW: Embase Screening Results'],
        [''],
        ['Review Title:', 'Approaches for Assessing and Mitigating Algorithmic Bias in Health AI'],
        ['Database:', 'Embase (via Oxford University Press publisher filter in OpenAlex)'],
        ['Date:', '2026-02-16'],
        ['Queries:', f'{len(query_stats)} focused text searches filtered to Oxford University Press publisher (strong biomedical coverage)'],
        [''],
        ['SCREENING FLOW:', ''],
        ['  Total Unique Papers Found:', str(meta['total_unique'])],
        ['  Title+Abstract Screened:', str(meta['total_unique'])],
        ['  Passed Title+Abstract:', str(meta['ta_included'])],
        ['  Passed Full-Text Screen:', str(meta['ft_included'])],
        [''],
        ['FINAL RESULTS:', ''],
        ['  INCLUDED:', str(meta['ft_included'])],
        ['  EXCLUDED (title+abstract):', str(meta['ta_excluded'])],
        ['  EXCLUDED (full-text):', str(meta['ft_excluded'])],
        [''],
        ['PRIMARY RESEARCH QUESTION:'],
        ['', 'What approaches are used to assess or mitigate algorithmic bias in health AI?'],
        [''],
        ['SECONDARY RESEARCH QUESTIONS:'],
        ['Q1:', 'Are certain bias axes (race/ethnicity, gender, SES, language, age, disability) more commonly assessed?'],
        ['Q2:', 'At what point of the AI development lifecycle does assessment vs. mitigation happen?'],
        ['Q3:', 'How do approaches vary by clinical setting or context?'],
        [''],
        ['INCLUSION:', 'Papers about bias/fairness in ML/AI algorithms applied to health (central topic)'],
        ['EXCLUSION:', 'Human cognitive biases, statistical bias without AI/ML, bias mentioned in passing only'],
    ]
    for ri, rd in enumerate(rows, 1):
        for ci, v in enumerate(rd, 1):
            cell = ws_s.cell(row=ri, column=ci, value=v)
            if ri == 1:
                cell.font = Font(bold=True, size=14, color='002147')
            elif v in ['  INCLUDED:']:
                cell.font = Font(bold=True, size=12, color='006600')
            elif 'EXCLUDED' in str(v) and ci == 1:
                cell.font = Font(bold=True, size=12, color='CC0000')
            elif v in ['FINAL RESULTS:', 'SCREENING FLOW:']:
                cell.font = Font(bold=True, size=12, color='002147')
            elif ci == 1 and ':' in str(v):
                cell.font = Font(bold=True)
    ws_s.column_dimensions['A'].width = 30
    ws_s.column_dimensions['B'].width = 85

    # --- Included Papers ---
    ws = wb.create_sheet('Included_Papers')
    for ci, (h, w) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = H_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    for idx, p in enumerate(included, 1):
        r = idx + 1
        vals = [
            idx, p.get('doi', ''), p.get('title', ''), p.get('authors', ''),
            p.get('year', ''), p.get('journal', ''), p.get('url', ''),
            (p.get('abstract', '') or '')[:32000],
            p.get('ft_reason', ''),
            p.get('study_type', ''), p.get('ai_ml_method', ''),
            p.get('health_domain', ''),
            p.get('bias_axes', ''), p.get('lifecycle_stage', ''),
            p.get('assessment_or_mitigation', ''),
            p.get('approach_method', ''),
            p.get('clinical_setting', ''),
            p.get('key_findings', ''),
            p.get('keywords', ''), p.get('pub_types', ''),
            p.get('cited_by_count', ''), ''
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=str(v) if v else '')
            cell.alignment = C_ALIGN
            cell.border = BORDER
        if idx % 2 == 0:
            for ci in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=ci).fill = ALT_FILL

    # --- Excluded Papers ---
    ws_e = wb.create_sheet('Excluded_Papers')
    eh = [('No.', 5), ('DOI', 30), ('Title', 65), ('Year', 6),
          ('Venue', 35), ('Exclusion Stage', 15), ('Reason', 50)]
    for ci, (h, w) in enumerate(eh, 1):
        cell = ws_e.cell(row=1, column=ci, value=h)
        cell.font = H_FONT
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = H_ALIGN
        cell.border = BORDER
        ws_e.column_dimensions[get_column_letter(ci)].width = w
    ws_e.freeze_panes = 'A2'

    for idx, p in enumerate(excluded, 1):
        vals = [idx, p.get('doi', ''), p.get('title', ''),
                p.get('year', ''), p.get('journal', ''),
                p.get('exclusion_stage', ''), p.get('exclusion_reason', '')]
        for ci, v in enumerate(vals, 1):
            cell = ws_e.cell(row=idx + 1, column=ci, value=str(v) if v else '')
            cell.alignment = C_ALIGN
            cell.border = BORDER

    # --- Search Strategy ---
    ws_q = wb.create_sheet('Search_Strategy')
    qh = [('ID', 5), ('Label', 40), ('Search Terms', 80), ('Results Found', 15), ('New Unique', 12), ('Cumulative', 12)]
    for ci, (h, w) in enumerate(qh, 1):
        cell = ws_q.cell(row=1, column=ci, value=h)
        cell.font = H_FONT
        cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
        cell.alignment = H_ALIGN
        cell.border = BORDER
        ws_q.column_dimensions[get_column_letter(ci)].width = w
    ws_q.freeze_panes = 'A2'

    for idx, qs in enumerate(query_stats, 1):
        vals = [qs['id'], qs['label'], qs['search'], qs['results_found'], qs['new_unique'], qs['cumulative']]
        for ci, v in enumerate(vals, 1):
            cell = ws_q.cell(row=idx + 1, column=ci, value=v)
            cell.alignment = C_ALIGN
            cell.border = BORDER

    out = '/home/user/Bias-Review-Paper/Embase_Screening_Results.xlsx'
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Sheets: {wb.sheetnames}")
    return out


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 70)
    print("Embase Screening Pipeline")
    print("Systematic Review: Algorithmic Bias in Health AI")
    print("=" * 70)

    # Step 1: Search
    print(f"\nSTEP 1: Searching Embase via OpenAlex ({len(QUERIES)} queries)...")
    all_papers = {}
    query_stats = []

    for q in QUERIES:
        print(f"\n  {q['id']}: {q['label']}")
        papers, count = search_openalex_embase(q['search'])
        new_count = 0
        for key, paper in papers.items():
            if key not in all_papers:
                all_papers[key] = paper
                new_count += 1
        query_stats.append({
            'id': q['id'], 'label': q['label'], 'search': q['search'],
            'results_found': count, 'new_unique': new_count,
            'cumulative': len(all_papers)
        })
        print(f"    Found: {count} | New unique: {new_count} | Cumulative: {len(all_papers)}")
        time.sleep(0.3)

    print(f"\n  Total unique Embase papers: {len(all_papers)}")
    with_abstract = sum(1 for p in all_papers.values() if p.get('abstract'))
    print(f"  Papers with abstracts: {with_abstract}")

    # Step 2: Title+Abstract Screening
    print(f"\nSTEP 2: Screening {len(all_papers)} papers by title+abstract...")
    ta_included = []
    ta_excluded = []

    for key, paper in all_papers.items():
        inc, reason = screen_paper(paper)
        paper['include'] = inc
        paper['screen_reason'] = reason
        if inc:
            ta_included.append(paper)
        else:
            ta_excluded.append(paper)

    print(f"  Title+Abstract INCLUDED: {len(ta_included)}")
    print(f"  Title+Abstract EXCLUDED: {len(ta_excluded)}")

    # Step 3: Full-text screening (abstract-based for Embase)
    print(f"\nSTEP 3: Full-text screening (abstract-based)...")
    ft_included = []
    ft_excluded_papers = []

    for paper in ta_included:
        inc, reason = fulltext_screen(paper)
        paper['ft_include'] = inc
        paper['ft_reason'] = reason
        if inc:
            ft_included.append(paper)
        else:
            ft_excluded_papers.append(paper)

    print(f"  Full-text INCLUDED: {len(ft_included)}")
    print(f"  Full-text EXCLUDED: {len(ft_excluded_papers)}")

    # Step 4: Extract columns
    print(f"\nSTEP 4: Extracting structured data...")
    for p in ft_included:
        p['study_type'] = extract_study_type(p)
        p['ai_ml_method'] = extract_ai_ml_method(p)
        p['health_domain'] = extract_health_domain(p)
        p['bias_axes'] = extract_bias_axes(p)
        p['lifecycle_stage'] = extract_lifecycle_stage(p)
        p['assessment_or_mitigation'] = extract_assessment_or_mitigation(p)
        p['approach_method'] = extract_approach_method(p)
        p['clinical_setting'] = extract_clinical_setting(p)
        p['key_findings'] = extract_key_findings(p)
    print(f"  Extracted all columns for {len(ft_included)} papers")

    # Sort by year (newest first)
    ft_included.sort(key=lambda x: x.get('year', '0'), reverse=True)

    # Combine all excluded papers for the Excel
    all_excluded = []
    for p in ta_excluded:
        all_excluded.append({
            **p,
            'exclusion_stage': 'Title+Abstract',
            'exclusion_reason': p.get('screen_reason', '')
        })
    for p in ft_excluded_papers:
        all_excluded.append({
            **p,
            'exclusion_stage': 'Full-Text',
            'exclusion_reason': p.get('ft_reason', '')
        })

    # Step 5: Build Excel
    print(f"\nSTEP 5: Building Excel...")
    meta = {
        'total_unique': len(all_papers),
        'ta_included': len(ta_included),
        'ta_excluded': len(ta_excluded),
        'ft_included': len(ft_included),
        'ft_excluded': len(ft_excluded_papers),
    }
    build_excel(ft_included, all_excluded, query_stats, meta)

    # Save JSON backup
    with open('/home/user/Bias-Review-Paper/embase_screening_data.json', 'w') as f:
        json.dump({
            'ft_included': ft_included,
            'ft_excluded': ft_excluded_papers,
            'ta_excluded': ta_excluded,
            'query_stats': query_stats,
            'meta': meta
        }, f, indent=2)
    print("Saved JSON backup: embase_screening_data.json")

    # Results
    print(f"\n{'='*70}")
    print("EMBASE SCREENING RESULTS")
    print(f"{'='*70}")
    print(f"  Total unique papers: {len(all_papers)}")
    print(f"  Title+Abstract included: {len(ta_included)}")
    print(f"  Full-text included: {len(ft_included)}")
    print(f"  Total excluded: {len(all_excluded)}")

    # Year distribution
    years = {}
    for p in ft_included:
        y = p.get('year', 'Unknown')
        years[y] = years.get(y, 0) + 1
    print(f"\nIncluded papers by year:")
    for y in sorted(years.keys(), reverse=True):
        print(f"  {y}: {years[y]}")

    # FT exclusion breakdown
    ft_reasons = {}
    for p in ft_excluded_papers:
        r = p.get('ft_reason', '')
        ft_reasons[r] = ft_reasons.get(r, 0) + 1
    print(f"\nFull-text exclusion breakdown:")
    for r, c in sorted(ft_reasons.items(), key=lambda x: -x[1]):
        print(f"  {c:>4}  {r}")

    # Column summary
    print(f"\nCOLUMN EXTRACTION SUMMARY")
    for col in ['study_type', 'ai_ml_method', 'health_domain', 'bias_axes',
                'lifecycle_stage', 'assessment_or_mitigation']:
        vals = {}
        for p in ft_included:
            for v in p.get(col, 'Not specified').split('; '):
                v = v.strip()
                if v:
                    vals[v] = vals.get(v, 0) + 1
        print(f"\n{col}:")
        for v, c in sorted(vals.items(), key=lambda x: -x[1])[:8]:
            print(f"  {c:>4}  {v}")


if __name__ == '__main__':
    main()
