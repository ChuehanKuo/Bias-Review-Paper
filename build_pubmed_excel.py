#!/usr/bin/env python3
"""
Build PubMed Excel file for systematic review screening.
Includes all papers that passed title+abstract screening.
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styling
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Column definitions
HEADERS = [
    ('No.', 5),
    ('PMID', 10),
    ('Title', 65),
    ('Authors', 20),
    ('Year', 6),
    ('Journal', 35),
    ('DOI', 30),
    ('URL', 35),
    ('Abstract', 80),
    ('Include/Exclude\n(Your Decision)', 18),
    ('Exclusion Reason\n(if excluded)', 25),
    ('Study Type', 20),
    ('AI/ML Method', 30),
    ('Health Domain', 25),
    ('Bias Axes Assessed\n(Q1: Which bias axes?)', 30),
    ('AI Lifecycle Stage\n(Q2: When in lifecycle?)', 25),
    ('Assessment vs Mitigation\n(Q2: Assess, mitigate, or both?)', 22),
    ('Approach/Method\nfor Bias Assessment/Mitigation', 35),
    ('Clinical Setting/Context\n(Q3: How vary by setting?)', 25),
    ('Key Findings', 50),
    ('Keywords/MeSH', 35),
    ('Publication Type', 20),
    ('Notes', 25),
]


def main():
    # Load data
    with open('/home/user/Bias-Review-Paper/pubmed_screening_data.json', 'r') as f:
        data = json.load(f)

    included = data['included']
    excluded = data['excluded']

    # Sort included by year (newest first)
    included.sort(key=lambda x: x.get('year', '0'), reverse=True)

    print(f"Building Excel with {len(included)} included papers...")

    wb = Workbook()
    wb.remove(wb.active)

    # ============================================================
    # SHEET 1: SUMMARY
    # ============================================================
    ws_sum = wb.create_sheet('Summary', index=0)
    summary_rows = [
        ['SYSTEMATIC REVIEW DATABASE SCREENING: PubMed/MEDLINE', '', ''],
        ['', '', ''],
        ['Review Title:', 'Approaches for Assessing and Mitigating Algorithmic Bias in Health AI', ''],
        ['Database:', 'PubMed/MEDLINE via NCBI E-utilities API', ''],
        ['Screening Date:', '2026-02-16', ''],
        ['Search Queries Used:', f'{len(data["query_stats"])} queries', ''],
        ['Total Unique PMIDs:', str(data['total_pmids']), ''],
        ['Papers Fetched:', str(data['total_fetched']), ''],
        ['Title+Abstract Screened:', str(len(included) + len(excluded)), ''],
        ['INCLUDED:', str(len(included)), ''],
        ['EXCLUDED:', str(len(excluded)), ''],
        ['', '', ''],
        ['RESEARCH QUESTIONS:', '', ''],
        ['Q1:', 'Are certain bias axes (race/ethnicity, gender, SES, language, age, disability) more commonly assessed in health AI literature?', ''],
        ['Q2:', 'At what point of the AI development lifecycle does assessment vs. mitigation happen?', ''],
        ['Q3:', 'How do approaches vary by clinical setting or context?', ''],
        ['', '', ''],
        ['INCLUSION CRITERIA:', '', ''],
        ['', 'Papers about bias/fairness in ML/AI algorithms applied to health/healthcare/medicine', ''],
        ['', 'All years, English language, all study types (empirical, framework, review, commentary, guideline)', ''],
        ['', '', ''],
        ['EXCLUSION CRITERIA:', '', ''],
        ['', 'Cognitive/human biases in clinical decision-making (NOT about AI)', ''],
        ['', 'Statistical bias (e.g., selection bias in epidemiology) without AI/ML', ''],
        ['', 'Papers not related to health/healthcare/medicine', ''],
        ['', '', ''],
        ['COLUMN GUIDE:', '', ''],
        ['Include/Exclude:', 'Your final decision after reading abstract (Include / Exclude)', ''],
        ['Study Type:', 'Empirical / Framework / Review / Scoping Review / Commentary / Guideline / Methodology', ''],
        ['AI/ML Method:', 'e.g., Deep learning, NLP, Logistic regression, Random forest, LLM, Imaging AI, Clinical prediction', ''],
        ['Health Domain:', 'e.g., Radiology, Dermatology, Cardiology, EHR, Mental health, Oncology, Emergency medicine', ''],
        ['Bias Axes (Q1):', 'Race/Ethnicity, Gender/Sex, Age, SES, Language, Disability, Insurance, Geographic, Intersectional', ''],
        ['Lifecycle Stage (Q2):', 'Data collection, Preprocessing, Model development/training, Evaluation, Deployment, Monitoring, Multiple', ''],
        ['Assessment vs Mitigation:', 'Assessment only / Mitigation only / Both / Framework / N/A', ''],
        ['Clinical Setting (Q3):', 'Hospital/Inpatient, Primary care, Emergency, Public health, ICU, Outpatient, Population-level', ''],
    ]

    for row_idx, row_data in enumerate(summary_rows, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_sum.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color='2F5496')
            elif value and col_idx == 1 and ':' in str(value):
                cell.font = Font(bold=True, size=11)
            elif value in ['INCLUDED:', 'EXCLUDED:']:
                cell.font = Font(bold=True, size=12, color='006600' if 'INCL' in value else 'CC0000')
    ws_sum.column_dimensions['A'].width = 25
    ws_sum.column_dimensions['B'].width = 85
    ws_sum.column_dimensions['C'].width = 15

    # ============================================================
    # SHEET 2: INCLUDED PAPERS (main screening sheet)
    # ============================================================
    ws = wb.create_sheet('Included_Papers')

    # Headers
    for col_idx, (header, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'

    # Data rows
    for idx, paper in enumerate(included, 1):
        row = idx + 1
        values = [
            idx,
            paper.get('pmid', ''),
            paper.get('title', ''),
            paper.get('authors', ''),
            paper.get('year', ''),
            paper.get('journal', ''),
            paper.get('doi', ''),
            paper.get('url', ''),
            paper.get('abstract', '')[:32000],  # Excel cell limit
            '',  # Include/Exclude - user fills
            '',  # Exclusion reason - user fills
            '',  # Study Type - user fills
            '',  # AI/ML Method - user fills
            '',  # Health Domain - user fills
            '',  # Bias Axes (Q1) - user fills
            '',  # Lifecycle Stage (Q2) - user fills
            '',  # Assessment vs Mitigation (Q2) - user fills
            '',  # Approach/Method - user fills
            '',  # Clinical Setting (Q3) - user fills
            '',  # Key Findings - user fills
            paper.get('keywords', ''),
            paper.get('pub_types', ''),
            '',  # Notes - user fills
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=str(value) if value else '')
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

        # Alternate row coloring
        if idx % 2 == 0:
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col_idx).fill = PatternFill(
                    start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

    # ============================================================
    # SHEET 3: EXCLUDED PAPERS
    # ============================================================
    ws_exc = wb.create_sheet('Excluded_Papers')
    exc_headers = [('No.', 5), ('PMID', 10), ('Title', 65), ('Year', 6),
                   ('Journal', 35), ('Exclusion Reason', 50)]

    for col_idx, (header, width) in enumerate(exc_headers, 1):
        cell = ws_exc.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws_exc.column_dimensions[get_column_letter(col_idx)].width = width

    ws_exc.freeze_panes = 'A2'

    for idx, paper in enumerate(excluded, 1):
        row = idx + 1
        values = [
            idx,
            paper.get('pmid', ''),
            paper.get('title', ''),
            paper.get('year', ''),
            paper.get('journal', ''),
            paper.get('screen_reason', '')
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws_exc.cell(row=row, column=col_idx, value=str(value) if value else '')
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

    # ============================================================
    # SHEET 4: SEARCH STRATEGY
    # ============================================================
    ws_search = wb.create_sheet('Search_Strategy')
    search_headers = [('Query No.', 8), ('Query', 80), ('Total in PubMed', 15),
                      ('Retrieved', 12), ('New Unique', 12)]

    for col_idx, (header, width) in enumerate(search_headers, 1):
        cell = ws_search.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws_search.column_dimensions[get_column_letter(col_idx)].width = width

    ws_search.freeze_panes = 'A2'

    for idx, qs in enumerate(data['query_stats'], 1):
        row = idx + 1
        values = [idx, qs['query'], qs['total_in_pubmed'], qs['retrieved'], qs['new_unique']]
        for col_idx, value in enumerate(values, 1):
            cell = ws_search.cell(row=row, column=col_idx, value=value)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

    # Save
    output = '/home/user/Bias-Review-Paper/PubMed_Screening_Results.xlsx'
    wb.save(output)
    print(f"\nExcel saved: {output}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Included papers: {len(included)}")
    print(f"Excluded papers: {len(excluded)}")


if __name__ == '__main__':
    main()
