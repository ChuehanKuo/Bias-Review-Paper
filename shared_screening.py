#!/usr/bin/env python3
"""
Shared screening criteria and column extraction for systematic review:
"Approaches for Assessing and Mitigating Algorithmic Bias in Health AI"

Single source of truth — used by PubMed, Scopus, ACM, and IEEE pipelines.
"""

import re
import json
import time
import urllib.request
import urllib.parse
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# TERM LISTS (consistent across all databases)
# ============================================================

AI_TERMS = [
    'machine learning', 'deep learning', 'artificial intelligence', 'neural network',
    'algorithm', 'predictive model', 'prediction model', 'classifier', 'classification',
    'natural language processing', 'nlp', 'computer vision', 'random forest',
    'logistic regression', 'xgboost', 'convolutional', 'transformer',
    'large language model', 'llm', 'decision support', 'risk prediction',
    'federated learning', 'reinforcement learning', 'foundation model',
    'chatgpt', 'gpt-4', 'supervised learning'
]

HEALTH_TERMS = [
    'health', 'clinical', 'medical', 'patient', 'hospital', 'disease',
    'diagnosis', 'treatment', 'care', 'pathology', 'radiology', 'dermatology',
    'cardiology', 'oncology', 'ophthalmology', 'psychiatry', 'mental health',
    'ehr', 'electronic health', 'biomedical', 'mortality', 'readmission',
    'sepsis', 'icu', 'emergency', 'chest x-ray', 'mammograph', 'cancer',
    'diabetes', 'cardiovascular', 'public health', 'healthcare', 'medicine'
]

STRONG_TITLE_TERMS = [
    'bias', 'fairness', 'fair ', 'unfair', 'equitable', 'equity',
    'disparity', 'disparities', 'discrimination', 'debiasing', 'debias',
    'underdiagnos', 'underrepresent', 'inequit'
]

AI_BIAS_TERMS = [
    'algorithmic bias', 'algorithmic fairness', 'ai bias', 'ai fairness',
    'machine learning bias', 'machine learning fairness', 'model bias',
    'bias mitigation', 'bias detection', 'bias assessment', 'fairness-aware',
    'fair machine learning', 'debiasing', 'disparate impact', 'demographic parity',
    'equalized odds', 'equal opportunity', 'fairness metric', 'fairness evaluation',
    'bias in ai', 'bias in machine learning', 'bias in algorithm', 'biased model',
    'biased prediction', 'mitigating bias', 'addressing bias', 'reducing bias',
    'assessing bias', 'evaluating bias', 'detecting bias', 'sources of bias',
    'racial bias', 'gender bias', 'age bias', 'socioeconomic bias', 'ethnic bias',
    'underdiagnosis bias', 'representation bias', 'data bias', 'label bias',
    'health disparit', 'health equit', 'health inequit', 'unfair'
]

HUMAN_ONLY_TERMS = [
    'implicit bias training', 'implicit bias among', 'clinician bias',
    'physician bias', 'provider bias', 'unconscious bias training',
    'implicit association test', 'weight bias', 'anti-fat bias'
]

AI_SPECIFIC_TERMS = [
    'algorithmic bias', 'ai bias', 'model bias', 'bias mitigation',
    'bias in ai', 'bias in machine learning', 'algorithmic fairness',
    'fairness-aware', 'debiasing algorithm'
]

APPROACH_INDICATORS = [
    'bias assessment', 'bias detection', 'bias evaluation', 'bias audit',
    'fairness assessment', 'fairness evaluation', 'fairness audit',
    'bias measurement', 'bias quantification', 'bias analysis',
    'measuring bias', 'detecting bias', 'evaluating bias',
    'fairness metric', 'fairness measure', 'bias metric',
    'demographic parity', 'equalized odds', 'equal opportunity',
    'disparate impact', 'predictive parity', 'calibration across',
    'subgroup analysis', 'disaggregated', 'stratified performance',
    'bias mitigation', 'bias reduction', 'bias correction',
    'debiasing', 'debias', 'fairness-aware',
    'fair machine learning', 'fair classification',
    'adversarial debiasing', 'reweighting', 'reweighing',
    'resampling', 'data augmentation for fairness',
    'fairness constraint', 'fairness regularization',
    'threshold adjustment', 'calibration', 'post-processing',
    'pre-processing', 'in-processing',
    'counterfactual fairness', 'causal fairness',
    'fairness framework', 'bias framework', 'equity framework',
    'ai fairness 360', 'aequitas', 'fairlearn',
    'bias toolkit', 'fairness toolkit',
    'model card', 'datasheet',
    'review of bias', 'survey of fairness', 'review of fairness',
    'bias mitigation strategies', 'approaches to fairness',
    'methods for bias', 'techniques for fairness',
]

METHOD_MAP = {
    'Deep Learning': ['deep learning', 'deep neural', 'cnn', 'convolutional neural', 'resnet', 'densenet', 'vgg', 'inception'],
    'NLP/LLM': ['natural language processing', 'nlp', 'large language model', 'llm', 'chatgpt', 'gpt-4', 'gpt-3', 'bert', 'transformer', 'text mining', 'text classification'],
    'Random Forest': ['random forest'],
    'Logistic Regression': ['logistic regression'],
    'XGBoost/Gradient Boosting': ['xgboost', 'gradient boosting', 'lightgbm', 'catboost'],
    'Support Vector Machine': ['support vector', 'svm'],
    'Neural Network': ['neural network', 'mlp', 'perceptron'],
    'Decision Tree': ['decision tree', 'cart'],
    'Ensemble Methods': ['ensemble', 'bagging', 'stacking'],
    'Federated Learning': ['federated learning'],
    'Reinforcement Learning': ['reinforcement learning'],
    'Computer Vision/Imaging AI': ['computer vision', 'image classification', 'image segmentation', 'object detection', 'medical imaging'],
    'Clinical Prediction Model': ['clinical prediction', 'risk prediction', 'risk score', 'prediction model', 'predictive model', 'prognostic model'],
    'Clinical Decision Support': ['clinical decision support', 'decision support system', 'cdss'],
    'Foundation Model': ['foundation model'],
    'Generative AI': ['generative ai', 'generative model', 'gan', 'generative adversarial'],
    'Clustering': ['clustering', 'k-means', 'unsupervised'],
    'Regression': ['linear regression', 'regression model'],
    'Survival Analysis': ['survival analysis', 'cox regression', 'hazard model'],
}

DOMAIN_MAP = {
    'Radiology/Medical Imaging': ['radiology', 'radiograph', 'x-ray', 'ct scan', 'mri', 'mammograph', 'ultrasound', 'medical imaging', 'chest radiograph'],
    'Dermatology': ['dermatology', 'skin lesion', 'skin cancer', 'melanoma', 'skin disease', 'dermatoscop'],
    'Cardiology': ['cardiology', 'cardiovascular', 'heart', 'cardiac', 'ecg', 'ekg', 'echocardiogram', 'atrial fibrillation', 'coronary'],
    'Oncology': ['oncology', 'cancer', 'tumor', 'tumour', 'neoplasm', 'malignant', 'chemotherapy', 'radiation therapy'],
    'Ophthalmology': ['ophthalmology', 'retinal', 'retinopathy', 'diabetic retinopathy', 'glaucoma', 'fundus', 'optic'],
    'Mental Health/Psychiatry': ['mental health', 'psychiatry', 'depression', 'anxiety', 'suicide', 'psychiatric', 'behavioral health', 'substance abuse', 'opioid'],
    'Emergency Medicine': ['emergency department', 'emergency medicine', 'trauma', 'triage', 'acute care'],
    'ICU/Critical Care': ['icu', 'intensive care', 'critical care', 'sepsis', 'mechanical ventilation'],
    'EHR/Health Informatics': ['electronic health record', 'ehr', 'emr', 'health information', 'clinical informatics', 'health informatics'],
    'Primary Care': ['primary care', 'family medicine', 'general practice', 'ambulatory'],
    'Surgery': ['surgery', 'surgical', 'operative', 'postoperative'],
    'Pathology': ['pathology', 'histopathology', 'cytology', 'biopsy'],
    'Pediatrics': ['pediatric', 'paediatric', 'child', 'neonatal', 'newborn'],
    'Obstetrics/Maternal Health': ['obstetric', 'maternal', 'pregnancy', 'prenatal', 'perinatal', 'childbirth'],
    'Nephrology': ['kidney', 'renal', 'nephrology', 'dialysis'],
    'Pulmonology': ['pulmonary', 'respiratory', 'lung', 'pneumonia', 'covid-19', 'asthma', 'copd'],
    'Neurology': ['neurology', 'neurological', 'stroke', 'alzheimer', 'dementia', 'brain', 'epilepsy'],
    'Genomics/Genetics': ['genomic', 'genetic', 'genome', 'gene expression', 'variant', 'sequencing'],
    'Public Health': ['public health', 'population health', 'epidemiology', 'surveillance', 'community health'],
    'Drug Discovery/Pharmacology': ['drug discovery', 'pharmacology', 'pharmaceutical', 'drug development'],
    'Pain Management': ['pain', 'analgesic', 'opioid prescribing'],
    'Infectious Disease': ['infectious disease', 'infection', 'hiv', 'tuberculosis', 'malaria', 'hepatitis'],
    'Endocrinology/Diabetes': ['diabetes', 'diabetic', 'endocrin', 'insulin', 'glucose', 'hba1c'],
    'Wearables/Remote Monitoring': ['wearable', 'pulse oximetry', 'spo2', 'remote monitoring', 'mobile health', 'mhealth'],
}

BIAS_AXIS_MAP = {
    'Race/Ethnicity': ['race', 'racial', 'ethnicity', 'ethnic', 'african american', 'black', 'white', 'hispanic', 'latino', 'asian', 'indigenous', 'minority', 'skin color', 'skin tone', 'skin pigment'],
    'Gender/Sex': ['gender', 'sex', 'male', 'female', 'women', 'men', 'transgender', 'sex-based', 'gender-based'],
    'Age': ['age', 'aging', 'elderly', 'older adult', 'pediatric', 'geriatric', 'age-related', 'age group'],
    'Socioeconomic Status': ['socioeconomic', 'income', 'poverty', 'insurance', 'uninsured', 'underinsured', 'medicaid', 'sociodemographic', 'education level', 'employment'],
    'Language': ['language', 'non-english', 'english proficiency', 'limited english', 'linguistic'],
    'Disability': ['disability', 'disabled', 'impairment', 'handicap'],
    'Geographic': ['geographic', 'rural', 'urban', 'region', 'low-resource', 'low-income country', 'global south', 'developing country'],
    'Insurance Status': ['insurance', 'uninsured', 'underinsured', 'medicaid', 'medicare', 'payer'],
    'Intersectional': ['intersectional', 'intersectionality', 'multiple sensitive', 'multi-attribute', 'compound'],
}

LIFECYCLE_STAGE_MAP = {
    'Data Collection': ['data collection', 'data acquisition', 'dataset creation', 'data gathering', 'sampling', 'cohort selection', 'data source'],
    'Data Preprocessing': ['preprocess', 'pre-process', 'data cleaning', 'imputation', 'feature engineering', 'data augmentation', 'resampling', 'oversampling', 'undersampling', 'reweighting', 'reweighing'],
    'Model Development/Training': ['model training', 'model development', 'training phase', 'in-processing', 'adversarial debiasing', 'fairness constraint', 'regularization', 'fair representation', 'calibration'],
    'Model Evaluation': ['model evaluation', 'evaluation', 'testing', 'validation', 'performance assessment', 'fairness metric', 'bias assessment', 'bias detection', 'audit', 'post-processing', 'subgroup analysis'],
    'Deployment': ['deployment', 'clinical implementation', 'real-world', 'production', 'clinical practice', 'clinical workflow', 'monitoring'],
}

APPROACH_MAP = {
    'Reweighting/Resampling': ['reweighting', 'reweighing', 'resampling', 'oversampling', 'undersampling', 'smote'],
    'Adversarial Debiasing': ['adversarial debiasing', 'adversarial learning', 'adversarial training'],
    'Fairness Constraints': ['fairness constraint', 'fairness regulariz', 'fairness-aware training'],
    'Calibration': ['calibration', 'recalibration', 'platt scaling'],
    'Threshold Adjustment': ['threshold', 'cutoff', 'operating point'],
    'Data Augmentation': ['data augmentation', 'synthetic data', 'augmented data', 'fair mixup'],
    'Subgroup Analysis': ['subgroup analysis', 'stratified analysis', 'subpopulation', 'disaggregated'],
    'Fairness Metrics Evaluation': ['demographic parity', 'equalized odds', 'equal opportunity', 'disparate impact', 'fairness metric', 'predictive parity', 'calibration across'],
    'Counterfactual Fairness': ['counterfactual', 'causal inference', 'causal fairness'],
    'Representation Learning': ['representation learning', 'fair representation', 'disentangl', 'embedding'],
    'Transfer Learning': ['transfer learning', 'domain adaptation', 'fine-tun'],
    'Federated Learning': ['federated learning', 'federated model'],
    'Explainability/Interpretability': ['explainab', 'interpretab', 'shap', 'lime', 'attention'],
    'Post-hoc Correction': ['post-processing', 'post-hoc', 'reject option'],
    'Diverse/Representative Data': ['diverse dataset', 'representative data', 'inclusive data', 'data diversity', 'balanced dataset'],
    'Bias Auditing Framework': ['audit', 'bias audit', 'fairness audit', 'model card', 'datasheet'],
    'Multi-task Learning': ['multi-task', 'multitask'],
    'Ensemble Methods': ['ensemble', 'model ensemble'],
    'Regularization': ['regulariz', 'penalty', 'constraint-based'],
}

CLINICAL_SETTING_MAP = {
    'Hospital/Inpatient': ['hospital', 'inpatient', 'hospitalized', 'admitted'],
    'ICU': ['icu', 'intensive care', 'critical care'],
    'Emergency Department': ['emergency department', 'emergency room', 'ed visit'],
    'Primary Care/Outpatient': ['primary care', 'outpatient', 'ambulatory', 'clinic visit', 'office visit'],
    'Public Health/Population': ['public health', 'population', 'community', 'screening program', 'surveillance'],
    'Telehealth/Remote': ['telehealth', 'telemedicine', 'remote', 'mobile health', 'mhealth', 'wearable'],
    'Clinical Trial': ['clinical trial', 'randomized controlled', 'rct'],
    'Laboratory/Pathology': ['laboratory', 'pathology', 'lab test'],
    'Imaging Center': ['imaging center', 'radiology department'],
    'Safety-Net/Underserved': ['safety net', 'safety-net', 'underserved', 'low-resource'],
    'Long-term Care': ['long-term care', 'nursing home', 'assisted living'],
}


# ============================================================
# TITLE+ABSTRACT SCREENING
# ============================================================

def screen_paper(paper):
    """
    Title+Abstract screening. Identical logic across all databases.
    Returns (included: bool, reason: str).
    """
    title = (paper.get('title', '') or '').lower()
    abstract = (paper.get('abstract', '') or '').lower()
    combined = title + ' ' + abstract
    kw = (paper.get('keywords', '') or '').lower()
    mesh = (paper.get('mesh_terms', '') or '').lower()
    all_text = combined + ' ' + kw + ' ' + mesh

    has_ai = any(t in all_text for t in AI_TERMS)
    has_health = any(t in all_text for t in HEALTH_TERMS)

    strong_in_title = any(t in title for t in STRONG_TITLE_TERMS)
    bias_count = sum(1 for t in AI_BIAS_TERMS if t in abstract)
    has_substantial_bias = strong_in_title or bias_count >= 2

    if not has_ai:
        return False, 'No AI/ML component'
    if not has_health:
        return False, 'Not health-related'
    if not has_substantial_bias:
        return False, 'Bias/fairness not central topic'

    # Exclude human cognitive biases only
    if any(t in all_text for t in HUMAN_ONLY_TERMS) and not any(t in all_text for t in AI_SPECIFIC_TERMS):
        return False, 'Human cognitive biases only'

    return True, 'Included'


# ============================================================
# FULL-TEXT SCREENING (for actual full text only)
# ============================================================

def fulltext_screen(paper, full_text):
    """
    Screen using actual full-text content.
    Only called when real full text is available (e.g., from PMC).
    Returns (included: bool, reason: str).
    """
    title = (paper.get('title', '') or '').lower()
    text = full_text.lower()
    combined = title + ' ' + text

    approach_count = sum(1 for t in APPROACH_INDICATORS if t in combined)
    has_ai = any(t in combined for t in AI_TERMS)
    has_health = any(t in combined for t in HEALTH_TERMS)

    bias_in_title = any(t in title for t in [
        'bias', 'fairness', 'fair ', 'equitable', 'equity', 'disparity',
        'disparities', 'debiasing', 'debias', 'underdiagnos', 'inequit',
        'discrimination'
    ])

    if not has_ai:
        return False, 'No AI/ML component in full text'
    if not has_health:
        return False, 'Not health-related in full text'

    if approach_count >= 2 and bias_in_title:
        return True, 'Included: bias central + approach content (full text)'
    elif approach_count >= 3:
        return True, 'Included: substantial approach content (full text)'
    elif bias_in_title and approach_count >= 1:
        return True, 'Included: bias is central topic (full text)'
    else:
        return False, f'Excluded: insufficient approach content in full text ({approach_count} indicators)'


# ============================================================
# PMC FULL-TEXT RETRIEVAL (shared across all pipelines)
# ============================================================

def batch_get_pmcids(ids, batch_size=200):
    """
    Convert PMIDs or DOIs to PMCIDs using NCBI ID converter.
    Accepts a list of PMIDs (strings) or DOIs (strings).
    Returns dict mapping input_id -> pmcid.
    """
    id_to_pmcid = {}
    for i in range(0, len(ids), batch_size):
        batch = ids[i:i+batch_size]
        ids_str = ','.join(batch)
        url = f"https://www.ncbi.nlm.nih.gov/pmc/utils/idconv/v1.0/?ids={ids_str}&format=json"
        try:
            with urllib.request.urlopen(urllib.request.Request(url), timeout=30) as resp:
                data = json.loads(resp.read().decode('utf-8'))
            for rec in data.get('records', []):
                pmcid = rec.get('pmcid', '')
                if not pmcid:
                    continue
                # Map back by pmid or doi (API returns pmid as int)
                pmid = str(rec.get('pmid', ''))
                doi = str(rec.get('doi', ''))
                if pmid and pmid in batch:
                    id_to_pmcid[pmid] = pmcid
                if doi and doi in batch:
                    id_to_pmcid[doi] = pmcid
                # Also try lowercase DOI match
                if doi:
                    for b in batch:
                        if b.lower() == doi.lower():
                            id_to_pmcid[b] = pmcid
        except Exception as e:
            print(f"    PMCID batch error: {e}")
        time.sleep(0.4)
    return id_to_pmcid


def fetch_pmc_fulltext(pmcid, retries=2):
    """Fetch full text from PMC as plain text."""
    url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?db=pmc&id={pmcid}&rettype=xml&retmode=xml"
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'SystematicReview/1.0'})
            with urllib.request.urlopen(req, timeout=30) as resp:
                xml_data = resp.read().decode('utf-8')
            root = ET.fromstring(xml_data)
            body = root.find('.//body')
            if body is not None:
                text = ' '.join(body.itertext())
                if len(text) > 100:
                    return text
            abstract = root.find('.//abstract')
            if abstract is not None:
                text = ' '.join(abstract.itertext())
                if len(text) > 100:
                    return text
            return None
        except Exception:
            if attempt < retries - 1:
                time.sleep(1)
    return None


def run_fulltext_screening(ta_included, id_field='doi', db_label=''):
    """
    Full-text screening for any database pipeline.

    Looks up PMCIDs via DOI or PMID, fetches PMC full text,
    and applies fulltext_screen().

    Args:
        ta_included: list of papers that passed title+abstract screening
        id_field: 'pmid' for PubMed, 'doi' for others
        db_label: label for logging (e.g. 'Scopus')

    Returns:
        (included, ft_excluded, stats_dict)
    """
    label = db_label or 'DB'

    # Build lookup IDs
    lookup_ids = []
    paper_by_id = {}
    for p in ta_included:
        lid = p.get(id_field, '')
        if lid:
            lookup_ids.append(lid)
            paper_by_id[lid] = p

    print(f"  Looking up PMCIDs for {len(lookup_ids)} papers (via {id_field})...")
    pmcid_map = batch_get_pmcids(lookup_ids)
    print(f"  PMCIDs found: {len(pmcid_map)} / {len(lookup_ids)}")

    included = []
    ft_excluded = []
    ft_screened = 0
    ft_unavailable = 0
    ft_passed = 0

    for p in ta_included:
        lid = p.get(id_field, '')
        pmcid = pmcid_map.get(lid) if lid else None

        if pmcid:
            full_text = fetch_pmc_fulltext(pmcid)
            if full_text and len(full_text) > 200:
                ft_screened += 1
                inc, reason = fulltext_screen(p, full_text)
                if inc:
                    p['ft_status'] = f'Full text screened ({pmcid})'
                    p['ft_reason'] = reason
                    included.append(p)
                    ft_passed += 1
                else:
                    p['ft_status'] = f'Full text excluded ({pmcid})'
                    p['ft_reason'] = reason
                    p['exclusion_stage'] = 'Full-Text'
                    p['exclusion_reason'] = reason
                    ft_excluded.append(p)
            else:
                ft_unavailable += 1
                p['ft_status'] = f'PMC fetch failed ({pmcid}) — passed through'
                p['ft_reason'] = 'Full text unavailable, passed through'
                included.append(p)
            time.sleep(0.15)
        else:
            ft_unavailable += 1
            p['ft_status'] = f'No PMC full text — passed through'
            p['ft_reason'] = f'{label}: no PMC ID found, passed through'
            included.append(p)

    stats = {
        'ft_screened': ft_screened,
        'ft_passed': ft_passed,
        'ft_excluded': len(ft_excluded),
        'ft_unavailable': ft_unavailable,
    }

    print(f"  Full-text screened: {ft_screened}")
    print(f"  Full-text passed: {ft_passed}")
    print(f"  Full-text excluded: {len(ft_excluded)}")
    print(f"  No full text (passed through): {ft_unavailable}")
    print(f"  TOTAL INCLUDED: {len(included)}")

    return included, ft_excluded, stats


# ============================================================
# COLUMN EXTRACTION
# ============================================================

def extract_study_type(paper):
    title = (paper.get('title', '') or '').lower()
    abstract = (paper.get('abstract', '') or '').lower()
    pub_types = (paper.get('pub_types', '') or '').lower()
    combined = title + ' ' + abstract + ' ' + pub_types

    if any(t in combined for t in ['systematic review', 'prisma', 'systematic literature review']):
        return 'Systematic Review'
    if any(t in combined for t in ['scoping review', 'scoping study']):
        return 'Scoping Review'
    if any(t in combined for t in ['narrative review', 'literature review', 'review article']):
        return 'Narrative Review'
    if 'meta-analysis' in combined:
        return 'Meta-Analysis'
    if any(t in combined for t in ['survey', 'questionnaire', 'interview']):
        return 'Survey/Qualitative'
    if any(t in combined for t in ['commentary', 'editorial', 'perspective', 'viewpoint', 'opinion', 'letter']):
        return 'Commentary/Editorial'
    if any(t in combined for t in ['guideline', 'recommendation', 'consensus', 'policy']):
        return 'Guideline/Policy'
    if any(t in combined for t in ['framework', 'toolkit', 'toolbox', 'pipeline']):
        return 'Framework/Toolkit'
    if any(t in combined for t in ['we propose', 'we develop', 'we introduce', 'novel method', 'novel approach', 'we present a method', 'we present a novel']):
        return 'Methodology'
    if any(t in combined for t in ['experiment', 'dataset', 'we evaluated', 'we trained', 'we tested', 'we applied', 'we analyzed', 'we assessed', 'results show', 'our results', 'we found']):
        return 'Empirical Study'
    if 'review' in combined:
        return 'Review'
    return 'Empirical Study'


def extract_ai_ml_method(paper):
    combined = ((paper.get('title', '') or '') + ' ' + (paper.get('abstract', '') or '')).lower()
    methods = [label for label, terms in METHOD_MAP.items() if any(t in combined for t in terms)]
    return '; '.join(methods) if methods else 'Not specified'


def extract_health_domain(paper):
    combined = ((paper.get('title', '') or '') + ' ' + (paper.get('abstract', '') or '') + ' ' + (paper.get('mesh_terms', '') or '') + ' ' + (paper.get('keywords', '') or '')).lower()
    domains = [label for label, terms in DOMAIN_MAP.items() if any(t in combined for t in terms)]
    return '; '.join(domains) if domains else 'General Healthcare'


def extract_bias_axes(paper):
    combined = ((paper.get('title', '') or '') + ' ' + (paper.get('abstract', '') or '')).lower()
    axes = [label for label, terms in BIAS_AXIS_MAP.items() if any(t in combined for t in terms)]
    return '; '.join(axes) if axes else 'Not specified'


def extract_lifecycle_stage(paper):
    combined = ((paper.get('title', '') or '') + ' ' + (paper.get('abstract', '') or '')).lower()
    stages = [label for label, terms in LIFECYCLE_STAGE_MAP.items() if any(t in combined for t in terms)]
    return '; '.join(stages) if stages else 'Not specified'


def extract_assessment_or_mitigation(paper):
    combined = ((paper.get('title', '') or '') + ' ' + (paper.get('abstract', '') or '')).lower()
    assess_terms = ['assess', 'detect', 'evaluat', 'audit', 'measur', 'quantif', 'identif', 'characteriz', 'examin', 'analyz', 'investigat']
    mitigate_terms = ['mitigat', 'debias', 'reduc', 'eliminat', 'correct', 'remov', 'address', 'alleviat', 'improv fairness', 'fairness-aware', 'fair machine', 'reweigh', 'adversarial debias']
    has_assess = any(t in combined for t in assess_terms)
    has_mitigate = any(t in combined for t in mitigate_terms)
    if has_assess and has_mitigate:
        return 'Both'
    elif has_mitigate:
        return 'Mitigation'
    elif has_assess:
        return 'Assessment'
    return 'Not specified'


def extract_approach_method(paper):
    abstract = (paper.get('abstract', '') or '').lower()
    approaches = [label for label, terms in APPROACH_MAP.items() if any(t in abstract for t in terms)]
    return '; '.join(approaches) if approaches else 'Not specified'


def extract_clinical_setting(paper):
    combined = ((paper.get('title', '') or '') + ' ' + (paper.get('abstract', '') or '')).lower()
    settings = [label for label, terms in CLINICAL_SETTING_MAP.items() if any(t in combined for t in terms)]
    return '; '.join(settings) if settings else 'Not specified'


def extract_key_findings(paper):
    abstract = (paper.get('abstract', '') or '')
    if not abstract:
        return 'No abstract available'
    lower = abstract.lower()
    for marker in ['CONCLUSION:', 'CONCLUSIONS:', 'RESULTS:', 'FINDINGS:', 'DISCUSSION:']:
        idx = lower.find(marker.lower())
        if idx != -1:
            finding = abstract[idx:].strip()
            sentences = re.split(r'(?<=[.!?])\s+', finding)
            result = ' '.join(sentences[:3])
            if len(result) > 500:
                result = result[:500] + '...'
            return result
    sentences = re.split(r'(?<=[.!?])\s+', abstract)
    if len(sentences) >= 2:
        result = ' '.join(sentences[-2:])
    else:
        result = abstract
    if len(result) > 500:
        result = result[:500] + '...'
    return result


def extract_all_columns(paper):
    """Extract all structured columns for an included paper. Modifies in-place."""
    paper['study_type'] = extract_study_type(paper)
    paper['ai_ml_method'] = extract_ai_ml_method(paper)
    paper['health_domain'] = extract_health_domain(paper)
    paper['bias_axes'] = extract_bias_axes(paper)
    paper['lifecycle_stage'] = extract_lifecycle_stage(paper)
    paper['assessment_or_mitigation'] = extract_assessment_or_mitigation(paper)
    paper['approach_method'] = extract_approach_method(paper)
    paper['clinical_setting'] = extract_clinical_setting(paper)
    paper['key_findings'] = extract_key_findings(paper)


# ============================================================
# EXCEL BUILDING
# ============================================================

def build_excel(included, excluded, query_stats, meta, config):
    """
    Build a standardized Excel workbook for any database.

    config keys:
        db_name:        Full name (e.g., 'PubMed/MEDLINE')
        db_label:       Short label (e.g., 'PubMed')
        header_color:   Hex color for header row
        alt_row_color:  Hex color for alternating rows
        id_field:       Paper dict key for database ID (e.g., 'pmid', 'doi')
        id_label:       Column header for that ID (e.g., 'PMID', 'DOI')
        output_path:    Path to save the Excel file
        date:           Date string
        search_desc:    Description of search method
    """
    wb = Workbook()
    wb.remove(wb.active)

    h_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    h_fill = PatternFill(start_color=config['header_color'], end_color=config['header_color'], fill_type='solid')
    h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_align = Alignment(vertical='top', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    alt_fill = PatternFill(start_color=config['alt_row_color'], end_color=config['alt_row_color'], fill_type='solid')

    headers = [
        ('No.', 5), (config['id_label'], 12), ('Title', 65), ('Authors', 20),
        ('Year', 6), ('Journal/Venue', 35), ('DOI', 30), ('URL', 35),
        ('Abstract', 80),
        ('Full-Text Status', 18),
        ('Study Type', 20), ('AI/ML Method', 30), ('Health Domain', 25),
        ('Bias Axes Assessed\n(Q1)', 30),
        ('AI Lifecycle Stage\n(Q2)', 25),
        ('Assessment vs Mitigation\n(Q2)', 22),
        ('Approach/Method', 35),
        ('Clinical Setting/Context\n(Q3)', 25),
        ('Key Findings', 50),
        ('Keywords/MeSH', 35), ('Publication Type', 20), ('Notes', 20),
    ]

    # --- Summary sheet ---
    ws_s = wb.create_sheet('Summary', index=0)
    rows = [
        [f'SYSTEMATIC REVIEW: {config["db_name"]} Screening Results'],
        [''],
        ['Review Title:', 'Approaches for Assessing and Mitigating Algorithmic Bias in Health AI'],
        ['Database:', config['db_name']],
        ['Search Method:', config.get('search_desc', '')],
        ['Date:', config['date']],
        ['Queries:', f'{len(query_stats)} queries'],
        [''],
        ['SCREENING FLOW:', ''],
        ['  Total Unique Papers Found:', str(meta.get('total_unique', ''))],
        ['  Passed Title+Abstract:', str(meta.get('ta_included', ''))],
        ['  Full-Text Screened:', str(meta.get('ft_screened', 0))],
        ['  Full-Text Not Available (passed through):', str(meta.get('ft_unavailable', 0))],
        ['  Full-Text Excluded:', str(meta.get('ft_excluded', 0))],
        [''],
        ['FINAL RESULTS:', ''],
        ['  INCLUDED:', str(len(included))],
        ['  EXCLUDED:', str(len(excluded))],
        [''],
        ['PRIMARY RESEARCH QUESTION:'],
        ['', 'What approaches are used to assess or mitigate algorithmic bias in health AI?'],
        [''],
        ['SECONDARY RESEARCH QUESTIONS:'],
        ['Q1:', 'Are certain bias axes (race/ethnicity, gender, SES, language, age, disability) more commonly assessed?'],
        ['Q2:', 'At what point of the AI development lifecycle does assessment vs. mitigation happen?'],
        ['Q3:', 'How do approaches vary by clinical setting or context?'],
        [''],
        ['INCLUSION:', 'Papers about bias/fairness in ML/AI algorithms applied to health (central topic)'],
        ['EXCLUSION:', 'Human cognitive biases, statistical bias without AI/ML, bias mentioned in passing only'],
    ]
    for ri, rd in enumerate(rows, 1):
        for ci, v in enumerate(rd, 1):
            cell = ws_s.cell(row=ri, column=ci, value=v)
            if ri == 1:
                cell.font = Font(bold=True, size=14, color=config['header_color'])
            elif v == '  INCLUDED:':
                cell.font = Font(bold=True, size=12, color='006600')
            elif 'EXCLUDED' in str(v) and ci == 1:
                cell.font = Font(bold=True, size=12, color='CC0000')
            elif v in ['FINAL RESULTS:', 'SCREENING FLOW:']:
                cell.font = Font(bold=True, size=12, color=config['header_color'])
            elif ci == 1 and ':' in str(v):
                cell.font = Font(bold=True)
    ws_s.column_dimensions['A'].width = 35
    ws_s.column_dimensions['B'].width = 85

    # --- Included Papers ---
    ws = wb.create_sheet('Included_Papers')
    for ci, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    id_field = config['id_field']
    for idx, p in enumerate(included, 1):
        r = idx + 1
        vals = [
            idx, p.get(id_field, ''), p.get('title', ''), p.get('authors', ''),
            p.get('year', ''), p.get('journal', ''), p.get('doi', ''), p.get('url', ''),
            (p.get('abstract', '') or '')[:32000],
            p.get('ft_status', 'N/A'),
            p.get('study_type', ''), p.get('ai_ml_method', ''),
            p.get('health_domain', ''),
            p.get('bias_axes', ''), p.get('lifecycle_stage', ''),
            p.get('assessment_or_mitigation', ''),
            p.get('approach_method', ''),
            p.get('clinical_setting', ''),
            p.get('key_findings', ''),
            '; '.join(filter(None, [p.get('keywords', ''), p.get('mesh_terms', '')])),
            p.get('pub_types', ''), ''
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=str(v) if v else '')
            cell.alignment = c_align
            cell.border = border
        if idx % 2 == 0:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=r, column=ci).fill = alt_fill

    # --- Excluded Papers ---
    ws_e = wb.create_sheet('Excluded_Papers')
    eh = [('No.', 5), (config['id_label'], 12), ('Title', 65), ('Year', 6),
          ('Journal/Venue', 35), ('Exclusion Stage', 15), ('Reason', 50)]
    for ci, (h, w) in enumerate(eh, 1):
        cell = ws_e.cell(row=1, column=ci, value=h)
        cell.font = h_font
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = h_align
        cell.border = border
        ws_e.column_dimensions[get_column_letter(ci)].width = w
    ws_e.freeze_panes = 'A2'

    for idx, p in enumerate(excluded, 1):
        vals = [idx, p.get(id_field, ''), p.get('title', ''),
                p.get('year', ''), p.get('journal', ''),
                p.get('exclusion_stage', ''), p.get('exclusion_reason', '')]
        for ci, v in enumerate(vals, 1):
            cell = ws_e.cell(row=idx + 1, column=ci, value=str(v) if v else '')
            cell.alignment = c_align
            cell.border = border

    # --- Search Strategy ---
    ws_q = wb.create_sheet('Search_Strategy')
    qh = [('ID', 5), ('Label', 40), ('Query/Search', 80), ('Results', 15), ('New Unique', 12), ('Cumulative', 12)]
    for ci, (h, w) in enumerate(qh, 1):
        cell = ws_q.cell(row=1, column=ci, value=h)
        cell.font = h_font
        cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
        cell.alignment = h_align
        cell.border = border
        ws_q.column_dimensions[get_column_letter(ci)].width = w
    ws_q.freeze_panes = 'A2'

    for idx, qs in enumerate(query_stats, 1):
        vals = [qs.get('id', ''), qs.get('label', ''), qs.get('query', qs.get('search', '')),
                qs.get('results', qs.get('total_in_pubmed', qs.get('results_found', ''))),
                qs.get('new_unique', ''), qs.get('cumulative', '')]
        for ci, v in enumerate(vals, 1):
            cell = ws_q.cell(row=idx + 1, column=ci, value=v)
            cell.alignment = c_align
            cell.border = border

    wb.save(config['output_path'])
    print(f"\nSaved: {config['output_path']}")
    print(f"Sheets: {wb.sheetnames}")
    return config['output_path']


def print_column_summary(included):
    """Print extraction summary for included papers."""
    print(f"\n{'='*70}")
    print("COLUMN EXTRACTION SUMMARY")
    print(f"{'='*70}")
    for col in ['study_type', 'ai_ml_method', 'health_domain', 'bias_axes',
                'lifecycle_stage', 'assessment_or_mitigation', 'approach_method', 'clinical_setting']:
        vals = {}
        for p in included:
            for v in p.get(col, 'Not specified').split('; '):
                v = v.strip()
                if v:
                    vals[v] = vals.get(v, 0) + 1
        print(f"\n{col}:")
        for v, c in sorted(vals.items(), key=lambda x: -x[1])[:10]:
            print(f"  {c:>4}  {v}")
